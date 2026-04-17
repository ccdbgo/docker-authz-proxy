#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
docker-authz-proxy 完整测试用例生成器 v3
在 v2 基础上对所有测试步骤进行详细描述，包含具体命令、参数、验证方法和检查点
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# 样式常量
# ─────────────────────────────────────────────────────────────────────────────
HEADERS = ["测试用例编号", "测试用例名称", "测试目的", "前提条件", "测试步骤", "预期结果"]
COL_WIDTHS = [18, 36, 44, 50, 90, 60]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

SHEET_COLORS = {
    "TC-SETUP":  "E8F4FD",
    "TC-ROOT":   "D5E8D4",
    "TC-SUDO":   "FFF2CC",
    "TC-DGRP":   "FCE4D6",
    "TC-BOB":    "EAD1DC",
    "TC-ALICE":  "CFE2F3",
    "TC-CROSS":  "E2EFDA",
    "TC-PERF":   "F4CCCC",
    "TC-ERR":    "EAD9FF",
    "TC-AUDIT":  "D9EAD3",
    "TC-DEPLOY": "FFF9C4",
}

thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")


def row_fill(tc_id):
    for prefix, color in SHEET_COLORS.items():
        if tc_id.startswith(prefix):
            return PatternFill("solid", fgColor=color)
    return PatternFill("solid", fgColor="FFFFFF")


def add_sheet(wb, title, cases, first=False):
    ws = wb.active if first else wb.create_sheet(title)
    if first:
        ws.title = title
    ws.freeze_panes = "A2"
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30
    for r, case in enumerate(cases, 2):
        fill = row_fill(case[0])
        for col, val in enumerate(case, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill
            c.border = BORDER
            c.alignment = WRAP
        ws.row_dimensions[r].height = 130
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# 公共前提条件
# ─────────────────────────────────────────────────────────────────────────────
PRE_COMMON = (
    "系统环境：\n"
    "• Linux 主机（Ubuntu 22.04 / CentOS 8+）已安装 Docker Engine 24.x\n"
    "• docker-authz-proxy 已编译并部署到 /usr/local/bin/docker-authz-proxy\n"
    "• 已创建用户：root(uid=0)、test-sudo(uid=1001,sudo组)、"
    "test-docker-g(uid=1002,docker组)、bob(uid=1003)、alice(uid=1004)\n"
    "• policy.yaml 已按各用户策略配置并加载\n"
    "• quota.yaml 已按各用户配额配置并加载\n"
    "• 代理服务已启动（systemctl status docker-authz-proxy 显示 active）\n"
    "• 各用户 socket 已创建于 /run/docker-authz/<username>.sock\n"
    "• 各用户已设置 DOCKER_HOST=unix:///run/docker-authz/<username>.sock\n"
    "• 审计日志目录 /var/log/docker-authz/ 已创建且可写"
)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 0: 环境搭建与策略配置
# ─────────────────────────────────────────────────────────────────────────────
setup_cases = [
    (
        "TC-SETUP-001",
        "创建测试用户 test-sudo",
        "创建 test-sudo 用户并加入 sudo 组，验证用户创建成功",
        "Linux 系统，当前以 root 身份操作",
        "1. 创建用户并设置 shell：\n"
        "   # useradd -m -s /bin/bash -u 1001 test-sudo\n"
        "   # echo 'test-sudo:Test1234!' | chpasswd\n\n"
        "2. 将用户加入 sudo 组：\n"
        "   # usermod -aG sudo test-sudo\n\n"
        "3. 验证用户信息：\n"
        "   # id test-sudo\n"
        "   # cat /etc/passwd | grep test-sudo\n\n"
        "4. 验证主目录已创建：\n"
        "   # ls -la /home/test-sudo\n\n"
        "5. 验证 sudo 权限：\n"
        "   # su - test-sudo -c 'sudo whoami'",
        "• id 输出包含 uid=1001(test-sudo) gid=1001(test-sudo) groups=1001(test-sudo),27(sudo)\n"
        "• /home/test-sudo 目录存在，权限为 700\n"
        "• sudo whoami 返回 root（需输入密码）"
    ),
    (
        "TC-SETUP-002",
        "创建测试用户 test-docker-g",
        "创建 test-docker-g 用户并加入 docker 组，验证用户创建成功",
        "Linux 系统，当前以 root 身份操作；docker 组已存在",
        "1. 创建用户：\n"
        "   # useradd -m -s /bin/bash -u 1002 test-docker-g\n"
        "   # echo 'test-docker-g:Test1234!' | chpasswd\n\n"
        "2. 将用户加入 docker 组：\n"
        "   # usermod -aG docker test-docker-g\n\n"
        "3. 验证用户信息：\n"
        "   # id test-docker-g\n\n"
        "4. 验证 docker 组成员：\n"
        "   # getent group docker",
        "• id 输出包含 uid=1002(test-docker-g) 且 groups 包含 docker\n"
        "• getent group docker 输出包含 test-docker-g"
    ),
    (
        "TC-SETUP-003",
        "创建普通测试用户 bob 和 alice",
        "创建普通用户 bob 和 alice，不加入任何特权组",
        "Linux 系统，当前以 root 身份操作",
        "1. 创建 bob 用户：\n"
        "   # useradd -m -s /bin/bash -u 1003 bob\n"
        "   # echo 'bob:Test1234!' | chpasswd\n\n"
        "2. 创建 alice 用户：\n"
        "   # useradd -m -s /bin/bash -u 1004 alice\n"
        "   # echo 'alice:Test1234!' | chpasswd\n\n"
        "3. 验证两个用户：\n"
        "   # id bob\n"
        "   # id alice\n\n"
        "4. 确认不在特权组：\n"
        "   # groups bob\n"
        "   # groups alice",
        "• bob: uid=1003(bob) gid=1003(bob) groups=1003(bob)（无 sudo/docker 组）\n"
        "• alice: uid=1004(alice) gid=1004(alice) groups=1004(alice)（无 sudo/docker 组）"
    ),
    (
        "TC-SETUP-004",
        "配置 policy.yaml - 完整访问控制策略",
        "为所有用户配置差异化访问控制策略并验证加载成功",
        "policy.yaml 文件路径：/etc/docker-authz/policy.yaml，文件可写",
        "1. 编写 policy.yaml 内容：\n"
        "   # cat > /etc/docker-authz/policy.yaml << 'EOF'\n"
        "   version: 1\n"
        "   default_action: allow\n"
        "   deny_rules:\n"
        "     - users: [test-sudo]\n"
        "       actions: [push, swarm, plugin, secret, config]\n"
        "     - users: [test-docker-g]\n"
        "       actions: [build, push, commit, load, save, swarm, plugin, secret, config]\n"
        "     - users: [bob]\n"
        "       actions: [exec, build, push, commit, load, save, swarm, plugin, secret, config]\n"
        "     - users: [alice]\n"
        "       actions: [exec, build, push, commit, load, save, prune, swarm, plugin, secret, config]\n"
        "   EOF\n\n"
        "2. 验证文件语法：\n"
        "   # python3 -c \"import yaml; yaml.safe_load(open('/etc/docker-authz/policy.yaml'))\"\n\n"
        "3. 发送 SIGHUP 重载策略：\n"
        "   # kill -HUP $(pgrep docker-authz-proxy)\n\n"
        "4. 查看代理日志确认加载成功：\n"
        "   # journalctl -u docker-authz-proxy -n 20 | grep -i policy",
        "• python3 验证无报错\n"
        "• 代理日志出现 'policy reloaded' 或 'policy loaded' 字样\n"
        "• 日志无 'error' 或 'failed' 字样"
    ),
    (
        "TC-SETUP-005",
        "配置 quota.yaml - 各用户资源配额",
        "为各用户配置差异化资源配额并验证加载成功",
        "quota.yaml 文件路径：/etc/docker-authz/quota.yaml，文件可写",
        "1. 编写 quota.yaml 内容：\n"
        "   # cat > /etc/docker-authz/quota.yaml << 'EOF'\n"
        "   version: 1\n"
        "   defaults:\n"
        "     cpu_cores: 2.0\n"
        "     mem_mb: 2048\n"
        "     max_containers: 5\n"
        "   users:\n"
        "     root:\n"
        "       cpu_cores: 0\n"
        "       mem_mb: 0\n"
        "       max_containers: 0\n"
        "     test-sudo:\n"
        "       cpu_cores: 4.0\n"
        "       mem_mb: 4096\n"
        "       max_containers: 10\n"
        "     test-docker-g:\n"
        "       cpu_cores: 2.0\n"
        "       mem_mb: 2048\n"
        "       max_containers: 5\n"
        "     bob:\n"
        "       cpu_cores: 1.0\n"
        "       mem_mb: 1024\n"
        "       max_containers: 3\n"
        "     alice:\n"
        "       cpu_cores: 1.0\n"
        "       mem_mb: 512\n"
        "       max_containers: 2\n"
        "   EOF\n\n"
        "2. 重启代理使配额生效：\n"
        "   # systemctl restart docker-authz-proxy\n\n"
        "3. 查看代理日志确认配额加载：\n"
        "   # journalctl -u docker-authz-proxy -n 20 | grep -i quota",
        "• 代理重启成功，状态为 active(running)\n"
        "• 日志显示配额配置已加载\n"
        "• 无解析错误"
    ),
    (
        "TC-SETUP-006",
        "验证各用户 socket 已创建",
        "确认代理为所有测试用户创建了独立的 Unix socket，权限正确",
        "代理服务已启动",
        "1. 列出 socket 目录：\n"
        "   # ls -la /run/docker-authz/\n\n"
        "2. 逐一检查每个 socket 的权限和所有者：\n"
        "   # stat /run/docker-authz/root.sock\n"
        "   # stat /run/docker-authz/test-sudo.sock\n"
        "   # stat /run/docker-authz/test-docker-g.sock\n"
        "   # stat /run/docker-authz/bob.sock\n"
        "   # stat /run/docker-authz/alice.sock\n\n"
        "3. 验证 socket 类型：\n"
        "   # file /run/docker-authz/bob.sock\n\n"
        "4. 验证普通用户无法访问他人 socket：\n"
        "   # su - bob -c 'ls /run/docker-authz/alice.sock'",
        "• 5 个 socket 文件均存在\n"
        "• 每个 socket 权限为 srw------- (0600)\n"
        "• 所有者分别为对应用户\n"
        "• file 命令输出包含 'socket'\n"
        "• bob 访问 alice.sock 返回 'Permission denied'"
    ),
    (
        "TC-SETUP-007",
        "验证 DOCKER_HOST 环境变量自动配置",
        "确认各用户登录后 DOCKER_HOST 自动指向其专属 socket",
        "profile.d 脚本 /etc/profile.d/docker-authz.sh 已部署",
        "1. 查看 profile.d 脚本内容：\n"
        "   # cat /etc/profile.d/docker-authz.sh\n\n"
        "2. 以 bob 身份登录新 shell 并检查环境变量：\n"
        "   # su - bob -c 'echo $DOCKER_HOST'\n\n"
        "3. 以 alice 身份检查：\n"
        "   # su - alice -c 'echo $DOCKER_HOST'\n\n"
        "4. 以 test-sudo 身份检查：\n"
        "   # su - test-sudo -c 'echo $DOCKER_HOST'\n\n"
        "5. 验证 docker 命令可通过环境变量连接代理：\n"
        "   # su - bob -c 'docker info' 2>&1 | head -5",
        "• bob 的 DOCKER_HOST=unix:///run/docker-authz/bob.sock\n"
        "• alice 的 DOCKER_HOST=unix:///run/docker-authz/alice.sock\n"
        "• test-sudo 的 DOCKER_HOST=unix:///run/docker-authz/test-sudo.sock\n"
        "• docker info 返回正常（不报连接错误）"
    ),
    (
        "TC-SETUP-008",
        "验证代理服务状态和日志",
        "确认代理服务正常运行，日志目录结构正确",
        "代理服务已启动",
        "1. 检查服务状态：\n"
        "   # systemctl status docker-authz-proxy\n\n"
        "2. 检查日志目录结构：\n"
        "   # ls -la /var/log/docker-authz/\n"
        "   # ls -la /var/log/docker-authz/user-operation/\n\n"
        "3. 检查代理进程：\n"
        "   # pgrep -a docker-authz-proxy\n\n"
        "4. 检查代理监听的 socket：\n"
        "   # ls -la /run/docker-authz/\n\n"
        "5. 查看最近启动日志：\n"
        "   # journalctl -u docker-authz-proxy --since '5 minutes ago'",
        "• 服务状态为 active(running)\n"
        "• /var/log/docker-authz/ 目录存在，包含 user-operation/、auth.log 等\n"
        "• 代理进程正在运行\n"
        "• 启动日志无 error 级别错误"
    ),
]  # end setup_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: root 用户 - 全量 Docker 命令遍历
# ─────────────────────────────────────────────────────────────────────────────
PRE_ROOT = PRE_COMMON + "\n• 当前操作用户：root（uid=0）\n• root 无 deny_rules，配额无限制\n• 执行方式：直接以 root 身份登录或 su - root"

root_cases = [
    (
        "TC-ROOT-001",
        "root - docker pull 拉取镜像",
        "验证 root 可以拉取任意镜像，且操作被正确审计",
        PRE_ROOT,
        "1. 切换到 root 用户：\n"
        "   # su - root  （或直接以 root 登录）\n\n"
        "2. 确认 DOCKER_HOST 指向 root socket：\n"
        "   # echo $DOCKER_HOST\n"
        "   # 预期：unix:///run/docker-authz/root.sock\n\n"
        "3. 拉取 nginx 镜像：\n"
        "   # docker pull nginx:latest\n"
        "   # 观察拉取进度输出\n\n"
        "4. 拉取 alpine 镜像：\n"
        "   # docker pull alpine:3.18\n\n"
        "5. 验证镜像已存在：\n"
        "   # docker images | grep -E 'nginx|alpine'\n\n"
        "6. 检查审计日志：\n"
        "   # tail -5 /var/log/docker-authz/user-operation/root.log | python3 -m json.tool",
        "• DOCKER_HOST 正确指向 root.sock\n"
        "• nginx:latest 和 alpine:3.18 均出现在 docker images 列表中\n"
        "• 审计日志包含两条记录：action=pull, result=allow, uid=0, user=root\n"
        "• 日志字段完整：time、user、uid、action、uri、result、status_code=200、latency_ms>0"
    ),
    (
        "TC-ROOT-002",
        "root - docker images 查看所有镜像（无过滤）",
        "验证 root 的 docker images 返回系统中所有用户的镜像，无任何过滤",
        PRE_ROOT + "\n• alice 已执行 docker pull python:3.11-slim\n• bob 已执行 docker pull node:18-alpine",
        "1. 以 root 身份执行：\n"
        "   # docker images\n"
        "   # 记录输出的镜像列表\n\n"
        "2. 查看包含中间层的完整列表：\n"
        "   # docker images --all\n\n"
        "3. 按仓库名过滤验证：\n"
        "   # docker images python\n"
        "   # docker images node\n\n"
        "4. 对比 alice 视角（应只看到自己的镜像）：\n"
        "   # su - alice -c 'docker images'\n\n"
        "5. 检查审计日志：\n"
        "   # tail -3 /var/log/docker-authz/user-operation/root.log | python3 -m json.tool",
        "• root 的 docker images 包含 nginx、alpine、python:3.11-slim、node:18-alpine 等所有镜像\n"
        "• alice 的 docker images 只包含 python:3.11-slim（alice 自己拉取的）\n"
        "• 审计日志记录 action=images, result=allow, uid=0"
    ),
    (
        "TC-ROOT-003",
        "root - docker run 创建并启动容器（无配额限制）",
        "验证 root 可以创建容器，且无资源配额限制（NanoCPUs=0, Memory=0）",
        PRE_ROOT + "\n• nginx:latest 镜像已存在",
        "1. 以 root 身份创建容器：\n"
        "   # docker run -d --name root-nginx nginx:latest\n"
        "   # 记录返回的容器 ID\n\n"
        "2. 验证容器正在运行：\n"
        "   # docker ps | grep root-nginx\n\n"
        "3. 检查容器资源限制（应为 0 即无限制）：\n"
        "   # docker inspect root-nginx | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     hc=d['HostConfig']\n"
        "     print('NanoCPUs:', hc['NanoCPUs'])\n"
        "     print('Memory:', hc['Memory'])\"\n\n"
        "4. 检查系统标签：\n"
        "   # docker inspect root-nginx | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     labels=d['Config']['Labels']\n"
        "     for k,v in labels.items():\n"
        "       if 'authz' in k: print(k,'=',v)\"\n\n"
        "5. 检查审计日志：\n"
        "   # grep 'create_container\\|start' /var/log/docker-authz/user-operation/root.log | tail -4",
        "• docker ps 显示 root-nginx 状态为 Up\n"
        "• NanoCPUs=0（无 CPU 限制），Memory=0（无内存限制）\n"
        "• 系统标签：system.authz.owner.uid=0，system.authz.owner=root，system.authz.created_by=docker\n"
        "• 审计日志记录 action=create_container 和 action=start，result=allow"
    ),
    (
        "TC-ROOT-004",
        "root - docker ps 查看所有容器（无过滤）",
        "验证 root 的 docker ps 返回所有用户的容器，无任何过滤",
        PRE_ROOT + "\n• alice 已创建容器（docker run -d --name alice-app alpine sleep 3600）\n• bob 已创建容器（docker run -d --name bob-app alpine sleep 3600）\n• root-nginx 容器正在运行",
        "1. 以 root 身份查看运行中的容器：\n"
        "   # docker ps\n"
        "   # 记录所有容器名称\n\n"
        "2. 查看所有容器（包括已停止的）：\n"
        "   # docker ps -a\n\n"
        "3. 以格式化方式查看容器名和所有者标签：\n"
        "   # docker ps --format 'table {{.Names}}\\t{{.Status}}\\t{{.Labels}}'\n\n"
        "4. 对比 bob 视角（应只看到自己的容器）：\n"
        "   # su - bob -c 'docker ps'\n\n"
        "5. 验证 root 能看到带前缀的容器名：\n"
        "   # docker ps | grep -E 'user-1003|user-1004'",
        "• root 的 docker ps 包含 root-nginx、user-1003-bob-app、user-1004-alice-app 等所有容器\n"
        "• bob 的 docker ps 只包含 user-1003-bob-app（显示为 bob-app）\n"
        "• root 能看到带 user-<uid>- 前缀的实际容器名\n"
        "• 审计日志记录 action=ps, result=allow, uid=0"
    ),
    (
        "TC-ROOT-005",
        "root - docker exec 进入任意用户容器",
        "验证 root 可以 exec 进入任意用户的容器（不受所有权限制）",
        PRE_ROOT + "\n• root-nginx 容器正在运行\n• alice 的容器 user-1004-alice-app 正在运行",
        "1. exec 进入 root 自己的容器：\n"
        "   # docker exec root-nginx echo 'hello from root-nginx'\n"
        "   # docker exec root-nginx ls /etc/nginx/\n\n"
        "2. exec 进入 alice 的容器（跨用户操作）：\n"
        "   # docker exec user-1004-alice-app echo 'root accessing alice container'\n"
        "   # docker exec user-1004-alice-app id\n\n"
        "3. 以交互方式 exec（非阻塞验证）：\n"
        "   # docker exec root-nginx sh -c 'nginx -v 2>&1'\n\n"
        "4. 检查审计日志：\n"
        "   # grep 'exec' /var/log/docker-authz/user-operation/root.log | tail -4",
        "• 步骤 1：返回 'hello from root-nginx' 和 nginx 配置文件列表\n"
        "• 步骤 2：成功执行，返回 'root accessing alice container' 和容器内 id 信息\n"
        "• 步骤 3：返回 nginx 版本信息\n"
        "• 审计日志记录 action=exec, result=allow, uid=0"
    ),
    (
        "TC-ROOT-006",
        "root - docker stop/start/restart 容器",
        "验证 root 可以停止、启动、重启任意容器，并验证状态变化",
        PRE_ROOT + "\n• root-nginx 容器正在运行",
        "1. 停止容器并验证状态：\n"
        "   # docker stop root-nginx\n"
        "   # docker ps -a | grep root-nginx\n"
        "   # 预期状态：Exited\n\n"
        "2. 启动容器并验证状态：\n"
        "   # docker start root-nginx\n"
        "   # docker ps | grep root-nginx\n"
        "   # 预期状态：Up\n\n"
        "3. 重启容器并验证状态：\n"
        "   # docker restart root-nginx\n"
        "   # sleep 2\n"
        "   # docker ps | grep root-nginx\n"
        "   # 预期状态：Up（重启后运行时间重置）\n\n"
        "4. 检查审计日志：\n"
        "   # grep -E 'stop|start|restart' /var/log/docker-authz/user-operation/root.log | tail -6",
        "• stop 后 docker ps -a 显示 Exited(0)\n"
        "• start 后 docker ps 显示 Up\n"
        "• restart 后 docker ps 显示 Up，运行时间为几秒\n"
        "• 审计日志记录 action=stop、action=start、action=restart，result=allow"
    ),
    (
        "TC-ROOT-007",
        "root - docker logs/stats/top 查看容器信息",
        "验证 root 可以查看任意容器的日志、资源统计和进程列表",
        PRE_ROOT + "\n• root-nginx 容器正在运行",
        "1. 查看容器日志：\n"
        "   # docker logs root-nginx\n"
        "   # docker logs --tail 10 root-nginx\n"
        "   # docker logs --since 1m root-nginx\n\n"
        "2. 查看容器资源统计（单次快照）：\n"
        "   # docker stats root-nginx --no-stream\n"
        "   # 记录 CPU%、MEM USAGE、NET I/O 等字段\n\n"
        "3. 查看容器内进程：\n"
        "   # docker top root-nginx\n"
        "   # docker top root-nginx aux\n\n"
        "4. 检查审计日志：\n"
        "   # grep -E '\"action\":\"logs\"' /var/log/docker-authz/user-operation/root.log | tail -3",
        "• docker logs 返回 nginx 访问日志或启动日志\n"
        "• docker stats 返回包含 CPU%、MEM USAGE/LIMIT、NET I/O、BLOCK I/O 的表格\n"
        "• docker top 返回容器内进程列表（包含 nginx master/worker 进程）\n"
        "• 审计日志记录 action=logs, result=allow"
    ),
    (
        "TC-ROOT-008",
        "root - docker inspect 检查容器/镜像，验证系统标签",
        "验证 root 可以 inspect 任意容器和镜像，且系统标签字段正确",
        PRE_ROOT + "\n• root-nginx 容器正在运行\n• nginx:latest 镜像已存在",
        "1. inspect 容器并提取关键字段：\n"
        "   # docker inspect root-nginx | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     print('ID:', d['Id'][:12])\n"
        "     print('Status:', d['State']['Status'])\n"
        "     print('Image:', d['Config']['Image'])\n"
        "     labels=d['Config']['Labels']\n"
        "     print('Labels:')\n"
        "     for k,v in labels.items(): print(' ', k, '=', v)\"\n\n"
        "2. inspect 镜像：\n"
        "   # docker inspect nginx:latest | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     print('ID:', d['Id'][:20])\n"
        "     print('Created:', d['Created'])\n"
        "     print('Size:', d['Size'])\"\n\n"
        "3. inspect alice 的容器（跨用户）：\n"
        "   # docker inspect user-1004-alice-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     labels=d['Config']['Labels']\n"
        "     for k,v in labels.items():\n"
        "       if 'authz' in k: print(k,'=',v)\"",
        "• root-nginx 的 Labels 包含：\n"
        "  - system.authz.owner.uid=0\n"
        "  - system.authz.owner=root\n"
        "  - system.authz.created_by=docker\n"
        "• alice 容器的 Labels 包含：\n"
        "  - system.authz.owner.uid=1004\n"
        "  - system.authz.owner=alice\n"
        "• inspect 镜像返回完整 JSON 信息"
    ),
    (
        "TC-ROOT-009",
        "root - docker cp 文件复制（宿主机↔容器）",
        "验证 root 可以在容器和宿主机之间双向复制文件",
        PRE_ROOT + "\n• root-nginx 容器正在运行",
        "1. 准备测试文件：\n"
        "   # echo 'docker-authz-proxy test file' > /tmp/root-test.txt\n"
        "   # cat /tmp/root-test.txt\n\n"
        "2. 从宿主机复制文件到容器：\n"
        "   # docker cp /tmp/root-test.txt root-nginx:/tmp/root-test.txt\n"
        "   # 验证文件已复制到容器内：\n"
        "   # docker exec root-nginx cat /tmp/root-test.txt\n\n"
        "3. 从容器复制文件到宿主机：\n"
        "   # docker cp root-nginx:/etc/nginx/nginx.conf /tmp/nginx-conf-backup.txt\n"
        "   # cat /tmp/nginx-conf-backup.txt | head -5\n\n"
        "4. 复制目录：\n"
        "   # docker cp root-nginx:/etc/nginx/ /tmp/nginx-backup/\n"
        "   # ls /tmp/nginx-backup/\n\n"
        "5. 检查审计日志：\n"
        "   # grep 'cp' /var/log/docker-authz/user-operation/root.log | tail -3",
        "• 步骤 2：容器内 cat 返回 'docker-authz-proxy test file'\n"
        "• 步骤 3：/tmp/nginx-conf-backup.txt 包含 nginx 配置内容\n"
        "• 步骤 4：/tmp/nginx-backup/ 目录包含 nginx 配置文件\n"
        "• 审计日志记录 action=cp, result=allow"
    ),
    (
        "TC-ROOT-010",
        "root - docker commit 提交容器为镜像",
        "验证 root 可以将容器提交为新镜像，并验证镜像可用",
        PRE_ROOT + "\n• root-nginx 容器正在运行",
        "1. 在容器内做修改（创建标记文件）：\n"
        "   # docker exec root-nginx sh -c 'echo committed > /tmp/commit-marker'\n\n"
        "2. 提交容器为新镜像：\n"
        "   # docker commit -m 'test commit by root' -a 'root' root-nginx root-nginx-custom:v1\n"
        "   # 记录返回的镜像 ID\n\n"
        "3. 验证新镜像已创建：\n"
        "   # docker images root-nginx-custom\n\n"
        "4. 验证新镜像包含修改内容：\n"
        "   # docker run --rm root-nginx-custom:v1 cat /tmp/commit-marker\n\n"
        "5. 检查镜像历史：\n"
        "   # docker history root-nginx-custom:v1\n\n"
        "6. 检查审计日志：\n"
        "   # grep 'commit' /var/log/docker-authz/user-operation/root.log | tail -2",
        "• docker images 显示 root-nginx-custom:v1 已创建\n"
        "• 运行新镜像返回 'committed'\n"
        "• docker history 显示最新一层包含 commit 信息\n"
        "• 审计日志记录 action=commit, result=allow"
    ),

    # ── root cases 11-20 ──────────────────────────────────────────────────────
    (
        "TC-ROOT-011",
        "root - docker build 构建镜像",
        "验证 root 可以构建 Docker 镜像，并验证构建产物",
        PRE_ROOT,
        "1. 创建构建目录和 Dockerfile：\n"
        "   # mkdir -p /tmp/root-build\n"
        "   # cat > /tmp/root-build/Dockerfile << 'EOF'\n"
        "   FROM alpine:3.18\n"
        "   RUN echo 'built by root' > /root-build-marker.txt\n"
        "   CMD [\"cat\", \"/root-build-marker.txt\"]\n"
        "   EOF\n\n"
        "2. 构建镜像：\n"
        "   # docker build -t root-test-img:v1 /tmp/root-build/\n"
        "   # 观察构建输出（Step 1/2, Step 2/2）\n\n"
        "3. 验证镜像已创建：\n"
        "   # docker images root-test-img\n\n"
        "4. 运行镜像验证内容：\n"
        "   # docker run --rm root-test-img:v1\n\n"
        "5. 检查审计日志：\n"
        "   # grep 'build' /var/log/docker-authz/user-operation/root.log | tail -2",
        "• 构建输出显示 Successfully built <image_id> 和 Successfully tagged root-test-img:v1\n"
        "• docker images 显示 root-test-img:v1\n"
        "• 运行镜像输出 'built by root'\n"
        "• 审计日志记录 action=build, result=allow"
    ),
    (
        "TC-ROOT-012",
        "root - docker tag/save/load 镜像标签与导出导入",
        "验证 root 可以为镜像打标签、导出为 tar 文件、再导入恢复",
        PRE_ROOT + "\n• root-test-img:v1 镜像已存在",
        "1. 为镜像打标签：\n"
        "   # docker tag root-test-img:v1 root-test-img:v2\n"
        "   # docker images root-test-img\n\n"
        "2. 导出镜像为 tar 文件：\n"
        "   # docker save root-test-img:v1 -o /tmp/root-test-img.tar\n"
        "   # ls -lh /tmp/root-test-img.tar\n\n"
        "3. 删除 v2 标签（保留 v1 用于验证）：\n"
        "   # docker rmi root-test-img:v2\n\n"
        "4. 从 tar 文件导入镜像：\n"
        "   # docker load -i /tmp/root-test-img.tar\n"
        "   # docker images root-test-img\n\n"
        "5. 检查审计日志：\n"
        "   # grep -E 'tag|save|load' /var/log/docker-authz/user-operation/root.log | tail -5",
        "• docker tag 成功，docker images 显示 v1 和 v2 两个标签\n"
        "• /tmp/root-test-img.tar 文件存在，大小合理（几 MB）\n"
        "• docker load 输出 'Loaded image: root-test-img:v1'\n"
        "• 审计日志记录 action=tag、action=save、action=load，result=allow"
    ),
    (
        "TC-ROOT-013",
        "root - docker rm/rmi 删除容器和镜像",
        "验证 root 可以删除任意用户的容器和镜像",
        PRE_ROOT + "\n• root-nginx 容器已停止\n• root-test-img:v1 镜像已存在",
        "1. 停止并删除容器：\n"
        "   # docker stop root-nginx\n"
        "   # docker rm root-nginx\n"
        "   # docker ps -a | grep root-nginx\n"
        "   # 预期：无输出（容器已删除）\n\n"
        "2. 强制删除运行中的容器：\n"
        "   # docker run -d --name root-rm-test alpine sleep 3600\n"
        "   # docker rm -f root-rm-test\n"
        "   # docker ps -a | grep root-rm-test\n\n"
        "3. 删除镜像：\n"
        "   # docker rmi root-test-img:v1\n"
        "   # docker images root-test-img\n\n"
        "4. 删除 alice 的镜像（跨用户，root 有权限）：\n"
        "   # docker rmi <alice-image-id>\n\n"
        "5. 检查审计日志：\n"
        "   # grep -E '\"action\":\"rm\"' /var/log/docker-authz/user-operation/root.log | tail -3",
        "• docker ps -a 中 root-nginx 不再出现\n"
        "• docker rm -f 成功删除运行中的容器\n"
        "• docker images 中 root-test-img 不再出现\n"
        "• 审计日志记录 action=rm、action=rmi，result=allow"
    ),
    (
        "TC-ROOT-014",
        "root - docker network 完整管理（ls/create/inspect/connect/disconnect/rm）",
        "验证 root 可以完整管理所有网络，且 network ls 无过滤",
        PRE_ROOT + "\n• alice 已创建网络（alice_u1004_alice-net）\n• bob 已创建网络（bob_u1003_bob-net）",
        "1. 查看所有网络（应包含所有用户的网络）：\n"
        "   # docker network ls\n"
        "   # docker network ls --format 'table {{.ID}}\\t{{.Name}}\\t{{.Driver}}'\n\n"
        "2. 创建自定义网络：\n"
        "   # docker network create --driver bridge --subnet 172.30.0.0/24 root-net\n"
        "   # docker network inspect root-net\n\n"
        "3. 创建容器并连接到网络：\n"
        "   # docker run -d --name root-net-test nginx\n"
        "   # docker network connect root-net root-net-test\n"
        "   # docker inspect root-net-test | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     nets=d['NetworkSettings']['Networks']\n"
        "     print('Networks:', list(nets.keys()))\"\n\n"
        "4. 断开网络连接：\n"
        "   # docker network disconnect root-net root-net-test\n\n"
        "5. 删除网络和容器：\n"
        "   # docker rm -f root-net-test\n"
        "   # docker network rm root-net\n\n"
        "6. 检查审计日志：\n"
        "   # grep 'network' /var/log/docker-authz/user-operation/root.log | tail -6",
        "• network ls 包含 alice_u1004_alice-net、bob_u1003_bob-net 等所有用户的网络\n"
        "• root-net 创建成功，subnet 为 172.30.0.0/24\n"
        "• connect 后容器 Networks 包含 root-net\n"
        "• disconnect 和 rm 均成功\n"
        "• 审计日志记录 network_ls、network_create、network_connect、network_disconnect、network_rm"
    ),
    (
        "TC-ROOT-015",
        "root - docker volume 完整管理（ls/create/inspect/rm）",
        "验证 root 可以完整管理所有卷，且 volume ls 无过滤",
        PRE_ROOT + "\n• alice 已创建卷（alice-vol）\n• bob 已创建卷（bob-vol）",
        "1. 查看所有卷（应包含所有用户的卷）：\n"
        "   # docker volume ls\n"
        "   # docker volume ls --format 'table {{.Name}}\\t{{.Driver}}'\n\n"
        "2. 创建命名卷：\n"
        "   # docker volume create root-vol\n"
        "   # docker volume inspect root-vol\n\n"
        "3. 使用卷创建容器并写入数据：\n"
        "   # docker run --rm -v root-vol:/data alpine sh -c 'echo root-data > /data/test.txt'\n"
        "   # docker run --rm -v root-vol:/data alpine cat /data/test.txt\n\n"
        "4. 删除卷：\n"
        "   # docker volume rm root-vol\n"
        "   # docker volume ls | grep root-vol\n\n"
        "5. 检查审计日志：\n"
        "   # grep 'volume' /var/log/docker-authz/user-operation/root.log | tail -5",
        "• volume ls 包含 alice-vol、bob-vol 等所有用户的卷\n"
        "• root-vol 创建成功，inspect 显示 Driver=local\n"
        "• 容器读取卷数据返回 'root-data'\n"
        "• volume rm 成功，volume ls 中不再出现 root-vol\n"
        "• 审计日志记录 volume_ls、volume_create、volume_rm"
    ),
    (
        "TC-ROOT-016",
        "root - docker system info/version/df/events",
        "验证 root 可以查看系统信息、版本、磁盘使用和事件流",
        PRE_ROOT,
        "1. 查看 Docker 系统信息：\n"
        "   # docker info\n"
        "   # docker info --format '{{.ServerVersion}}'\n"
        "   # docker info --format '{{.Containers}}'\n\n"
        "2. 查看 Docker 版本：\n"
        "   # docker version\n"
        "   # docker version --format '{{.Server.Version}}'\n\n"
        "3. 查看磁盘使用情况：\n"
        "   # docker system df\n"
        "   # docker system df -v\n\n"
        "4. 监听事件流（后台运行 5 秒）：\n"
        "   # timeout 5 docker events --since 5m &\n"
        "   # docker run --rm alpine echo 'trigger event'\n"
        "   # wait\n\n"
        "5. 检查审计日志：\n"
        "   # grep -E 'info|df|events' /var/log/docker-authz/user-operation/root.log | tail -5",
        "• docker info 返回完整系统信息（Containers、Images、ServerVersion 等）\n"
        "• docker version 返回 Client 和 Server 版本信息\n"
        "• docker system df 返回镜像、容器、卷的磁盘使用统计\n"
        "• events 输出包含容器 create、start、die、destroy 事件\n"
        "• 审计日志记录 action=info、action=df、action=events"
    ),
    (
        "TC-ROOT-017",
        "root - docker system prune 系统清理",
        "验证 root 可以执行系统清理，清除已停止容器和悬空镜像",
        PRE_ROOT + "\n• 存在已停止的容器和悬空镜像（通过 docker build 产生）",
        "1. 创建一些待清理的资源：\n"
        "   # docker run --name root-stopped-1 alpine echo done\n"
        "   # docker run --name root-stopped-2 alpine echo done\n"
        "   # 记录已停止容器数量：docker ps -a | grep Exited | wc -l\n\n"
        "2. 查看清理前的磁盘使用：\n"
        "   # docker system df\n\n"
        "3. 执行系统清理（不清理卷）：\n"
        "   # docker system prune -f\n"
        "   # 观察输出（显示删除的资源和释放的空间）\n\n"
        "4. 验证清理结果：\n"
        "   # docker ps -a | grep Exited\n"
        "   # docker images -f dangling=true\n\n"
        "5. 检查审计日志：\n"
        "   # grep 'prune' /var/log/docker-authz/user-operation/root.log | tail -3",
        "• prune 输出显示 'Deleted Containers:' 和 'Total reclaimed space:'\n"
        "• docker ps -a 中已停止容器被清除\n"
        "• docker images -f dangling=true 无输出（悬空镜像已清除）\n"
        "• 审计日志记录 action=prune, result=allow"
    ),
    (
        "TC-ROOT-018",
        "root - docker pause/unpause/kill/wait/rename",
        "验证 root 可以使用容器控制命令：暂停、恢复、强制终止、等待、重命名",
        PRE_ROOT,
        "1. 创建测试容器：\n"
        "   # docker run -d --name root-ctrl-test nginx\n\n"
        "2. 暂停容器并验证状态：\n"
        "   # docker pause root-ctrl-test\n"
        "   # docker ps | grep root-ctrl-test\n"
        "   # 预期状态：Up (Paused)\n\n"
        "3. 恢复容器并验证状态：\n"
        "   # docker unpause root-ctrl-test\n"
        "   # docker ps | grep root-ctrl-test\n"
        "   # 预期状态：Up（无 Paused）\n\n"
        "4. 重命名容器：\n"
        "   # docker rename root-ctrl-test root-ctrl-renamed\n"
        "   # docker ps | grep root-ctrl-renamed\n\n"
        "5. 强制终止容器：\n"
        "   # docker kill root-ctrl-renamed\n"
        "   # docker ps -a | grep root-ctrl-renamed\n"
        "   # 预期状态：Exited\n\n"
        "6. 等待容器退出并获取退出码：\n"
        "   # docker run -d --name root-wait-test alpine sleep 2\n"
        "   # docker wait root-wait-test\n"
        "   # 预期输出：0\n\n"
        "7. 清理：\n"
        "   # docker rm root-ctrl-renamed root-wait-test",
        "• pause 后状态为 Up (Paused)\n"
        "• unpause 后状态恢复为 Up\n"
        "• rename 后 docker ps 显示新名称 root-ctrl-renamed\n"
        "• kill 后状态为 Exited(137)\n"
        "• wait 返回退出码 0\n"
        "• 审计日志记录 stop(pause/unpause/kill)、stop(rename)、logs(wait)"
    ),
    (
        "TC-ROOT-019",
        "root - docker export/import/diff/history/port",
        "验证 root 可以使用容器诊断和镜像操作命令",
        PRE_ROOT,
        "1. 创建测试容器并修改文件系统：\n"
        "   # docker run -d --name root-diag -p 8082:80 nginx\n"
        "   # docker exec root-diag sh -c 'echo test > /tmp/diag-file'\n\n"
        "2. 查看文件系统变更：\n"
        "   # docker diff root-diag\n"
        "   # 预期：A /tmp/diag-file\n\n"
        "3. 查看端口映射：\n"
        "   # docker port root-diag\n"
        "   # 预期：80/tcp -> 0.0.0.0:8082\n\n"
        "4. 查看镜像历史：\n"
        "   # docker history nginx:latest\n"
        "   # docker history --no-trunc nginx:latest | head -5\n\n"
        "5. 导出容器文件系统：\n"
        "   # docker export root-diag -o /tmp/root-diag-export.tar\n"
        "   # ls -lh /tmp/root-diag-export.tar\n\n"
        "6. 导入为新镜像：\n"
        "   # docker import /tmp/root-diag-export.tar root-imported:v1\n"
        "   # docker images root-imported\n\n"
        "7. 清理：\n"
        "   # docker rm -f root-diag\n"
        "   # docker rmi root-imported:v1",
        "• docker diff 输出包含 A /tmp/diag-file\n"
        "• docker port 输出 80/tcp -> 0.0.0.0:8082\n"
        "• docker history 显示镜像层历史\n"
        "• /tmp/root-diag-export.tar 文件存在（几十 MB）\n"
        "• docker images 显示 root-imported:v1"
    ),
    (
        "TC-ROOT-020",
        "root - docker update/search/login 及系统标签完整验证",
        "验证 root 可以动态更新容器资源、搜索镜像，并完整验证系统标签防篡改",
        PRE_ROOT,
        "1. 动态更新容器资源限制：\n"
        "   # docker run -d --name root-update-test nginx\n"
        "   # docker update --cpus=0.5 --memory=256m root-update-test\n"
        "   # docker inspect root-update-test | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     hc=d['HostConfig']\n"
        "     print('NanoCPUs:', hc['NanoCPUs'])\n"
        "     print('Memory:', hc['Memory'])\"\n\n"
        "2. 搜索 Docker Hub 镜像：\n"
        "   # docker search nginx --limit 5\n"
        "   # docker search --filter is-official=true python --limit 3\n\n"
        "3. 验证系统标签防篡改（用户尝试覆盖 owner.uid）：\n"
        "   # docker run -d --name root-tamper-test \\\n"
        "       -l system.authz.owner.uid=9999 \\\n"
        "       -l system.authz.owner=hacker \\\n"
        "       nginx\n"
        "   # docker inspect root-tamper-test | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     labels=d['Config']['Labels']\n"
        "     for k,v in labels.items():\n"
        "       if 'authz' in k: print(k,'=',v)\"\n\n"
        "4. 清理：\n"
        "   # docker rm -f root-update-test root-tamper-test",
        "• update 后 NanoCPUs=500000000，Memory=268435456\n"
        "• docker search 返回镜像搜索结果列表\n"
        "• 系统标签防篡改：system.authz.owner.uid=0（代理注入值），用户指定的 9999 被覆盖\n"
        "• system.authz.owner 值为 hacker,root（用户值在前，代理追加在后）"
    ),
    (
        "TC-ROOT-021",
        "root - 操作其他用户容器（所有权绕过验证）",
        "验证 root 可以操作任意用户的容器，不受所有权限制",
        PRE_ROOT + "\n• bob 已创建容器 user-1003-bob-app（运行中）\n• alice 已创建容器 user-1004-alice-app（运行中）",
        "1. root 停止 bob 的容器：\n"
        "   # docker stop user-1003-bob-app\n"
        "   # docker ps -a | grep user-1003-bob-app\n\n"
        "2. root 启动 bob 的容器：\n"
        "   # docker start user-1003-bob-app\n\n"
        "3. root exec 进入 alice 的容器：\n"
        "   # docker exec user-1004-alice-app echo 'root can access alice container'\n\n"
        "4. root 删除 alice 的容器：\n"
        "   # docker rm -f user-1004-alice-app\n"
        "   # docker ps -a | grep user-1004-alice-app\n\n"
        "5. 检查审计日志（root 操作他人容器的记录）：\n"
        "   # grep -E 'stop|start|exec|rm' /var/log/docker-authz/user-operation/root.log | tail -8",
        "• 所有操作均成功，root 不受所有权限制\n"
        "• docker ps -a 中 user-1004-alice-app 不再出现\n"
        "• 审计日志记录所有操作，result=allow，uid=0\n"
        "• 对比：bob 尝试操作 alice 容器会返回 403"
    ),
]  # end root_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: test-sudo 用户
# 策略：禁止 push、swarm、plugin、secret、config
# 配额：cpu=4, mem=4096MB, max_containers=10
# ─────────────────────────────────────────────────────────────────────────────
PRE_SUDO = (
    PRE_COMMON + "\n"
    "• 当前操作用户：test-sudo（uid=1001，sudo 组）\n"
    "• 执行方式：su - test-sudo 后使用 sudo docker <cmd>\n"
    "• deny_rules：禁止 push、swarm、plugin、secret、config\n"
    "• 配额：cpu_cores=4, mem_mb=4096, max_containers=10\n"
    "• sudo 执行时 eUID=0，loginUID=1001（代理以 loginUID 识别真实身份）"
)

sudo_cases = [
    (
        "TC-SUDO-001",
        "test-sudo - 身份识别验证（sudo 双重身份）",
        "验证代理正确识别 sudo 用户的真实 UID（loginUID=1001），而非有效 UID（0）",
        PRE_SUDO,
        "1. 切换到 test-sudo 用户：\n"
        "   # su - test-sudo\n\n"
        "2. 确认 DOCKER_HOST：\n"
        "   $ echo $DOCKER_HOST\n"
        "   # 预期：unix:///run/docker-authz/test-sudo.sock\n\n"
        "3. 以 sudo 方式执行 docker ps：\n"
        "   $ sudo docker ps\n\n"
        "4. 查看认证日志（验证双重身份记录）：\n"
        "   # tail -5 /var/log/docker-authz/auth.log | python3 -m json.tool\n\n"
        "5. 查看操作日志（验证以真实用户记录）：\n"
        "   # tail -3 /var/log/docker-authz/user-operation/test-sudo.log | python3 -m json.tool",
        "• DOCKER_HOST 正确指向 test-sudo.sock\n"
        "• docker ps 正常返回（只显示 test-sudo 自己的容器）\n"
        "• auth.log 记录：effective_uid=0，real_uid=1001，switched_identity=true，username=test-sudo\n"
        "• 操作日志记录在 test-sudo.log 中（而非 root.log），uid=1001"
    ),
    (
        "TC-SUDO-002",
        "test-sudo - docker pull/images（允许）",
        "验证 test-sudo 可以拉取镜像，且 docker images 只显示自己的镜像",
        PRE_SUDO + "\n• alice 已拉取 python:3.11-slim",
        "1. 以 test-sudo 身份拉取镜像：\n"
        "   $ sudo docker pull redis:7\n"
        "   # 观察拉取进度\n\n"
        "2. 查看镜像列表：\n"
        "   $ sudo docker images\n"
        "   # 记录显示的镜像\n\n"
        "3. 验证 alice 的私有镜像不可见：\n"
        "   $ sudo docker images | grep python\n"
        "   # 预期：无输出（python:3.11-slim 是 alice 的私有镜像）\n\n"
        "4. 检查审计日志：\n"
        "   # tail -5 /var/log/docker-authz/user-operation/test-sudo.log | python3 -m json.tool",
        "• redis:7 拉取成功\n"
        "• docker images 只显示 test-sudo 自己拉取/构建的镜像和公共镜像\n"
        "• python:3.11-slim（alice 的镜像）不可见\n"
        "• 审计日志记录 action=pull, result=allow, uid=1001, user=test-sudo"
    ),
    (
        "TC-SUDO-003",
        "test-sudo - docker run（允许，验证配额注入 4CPU/4096MB）",
        "验证 test-sudo 创建容器时代理自动注入 4CPU/4096MB 配额",
        PRE_SUDO + "\n• redis:7 镜像已存在",
        "1. 创建容器（不指定资源限制）：\n"
        "   $ sudo docker run -d --name sudo-app redis:7\n\n"
        "2. 验证容器正在运行：\n"
        "   $ sudo docker ps | grep sudo-app\n\n"
        "3. 检查代理注入的资源限制：\n"
        "   $ sudo docker inspect sudo-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     hc=d['HostConfig']\n"
        "     print('NanoCPUs:', hc['NanoCPUs'], '(期望: 4000000000)')\n"
        "     print('Memory:', hc['Memory'], '(期望: 4294967296)'\"\n\n"
        "4. 检查系统标签（应显示真实用户 test-sudo）：\n"
        "   $ sudo docker inspect sudo-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     labels=d['Config']['Labels']\n"
        "     for k,v in labels.items():\n"
        "       if 'authz' in k: print(k,'=',v)\"\n\n"
        "5. 检查审计日志：\n"
        "   # tail -3 /var/log/docker-authz/user-operation/test-sudo.log | python3 -m json.tool",
        "• 容器创建成功，状态为 Up\n"
        "• NanoCPUs=4000000000（4 CPU），Memory=4294967296（4096MB）\n"
        "• 系统标签：system.authz.owner.uid=1001，system.authz.owner=test-sudo\n"
        "• 审计日志记录 uid=1001（真实 UID），而非 0"
    ),
    (
        "TC-SUDO-004",
        "test-sudo - docker ps（只看自己容器）",
        "验证 test-sudo 的 docker ps 只返回自己的容器，bob 的容器不可见",
        PRE_SUDO + "\n• sudo-app 容器运行中\n• bob 已创建容器 user-1003-bob-app",
        "1. 查看运行中的容器：\n"
        "   $ sudo docker ps\n"
        "   # 记录容器列表\n\n"
        "2. 查看所有容器（包括已停止的）：\n"
        "   $ sudo docker ps -a\n\n"
        "3. 验证 bob 的容器不可见：\n"
        "   $ sudo docker ps | grep bob\n"
        "   # 预期：无输出\n\n"
        "4. 验证容器名称前缀已剥离（test-sudo 看到的是 sudo-app 而非 user-1001-sudo-app）：\n"
        "   $ sudo docker ps --format '{{.Names}}'\n\n"
        "5. 以 root 身份验证实际容器名：\n"
        "   # docker ps | grep sudo-app",
        "• test-sudo 的 docker ps 只显示 sudo-app（自己的容器）\n"
        "• bob 的容器 user-1003-bob-app 不可见\n"
        "• test-sudo 看到的容器名为 sudo-app（前缀已剥离）\n"
        "• root 看到的实际名称为 user-1001-sudo-app"
    ),
    (
        "TC-SUDO-005",
        "test-sudo - docker exec（允许，只能进入自己的容器）",
        "验证 test-sudo 可以 exec 进入自己的容器，但不能进入他人容器",
        PRE_SUDO + "\n• sudo-app 容器运行中\n• bob 的容器 user-1003-bob-app 运行中",
        "1. exec 进入自己的容器（应成功）：\n"
        "   $ sudo docker exec sudo-app redis-cli ping\n"
        "   # 预期：PONG\n\n"
        "2. exec 执行多条命令：\n"
        "   $ sudo docker exec sudo-app sh -c 'redis-cli info server | head -5'\n\n"
        "3. 尝试 exec 进入 bob 的容器（应被拒绝）：\n"
        "   $ sudo docker exec user-1003-bob-app echo hi\n"
        "   # 预期：返回 403 Forbidden\n\n"
        "4. 检查审计日志（允许和拒绝的记录）：\n"
        "   # grep 'exec' /var/log/docker-authz/user-operation/test-sudo.log | tail -4",
        "• 步骤 1：返回 PONG\n"
        "• 步骤 2：返回 redis server 信息\n"
        "• 步骤 3：返回 403 Forbidden，错误信息包含 'ownership' 或 'permission denied'\n"
        "• 审计日志：自己容器的 exec 记录 result=allow；bob 容器的 exec 记录 result=deny, deny_reason=ownership"
    ),
    (
        "TC-SUDO-006",
        "test-sudo - docker stop/start/restart/rm（只能操作自己的容器）",
        "验证 test-sudo 可以管理自己的容器，但不能操作他人容器",
        PRE_SUDO + "\n• sudo-app 容器运行中\n• bob 的容器 user-1003-bob-app 运行中",
        "1. 停止自己的容器：\n"
        "   $ sudo docker stop sudo-app\n"
        "   $ sudo docker ps -a | grep sudo-app\n"
        "   # 预期状态：Exited\n\n"
        "2. 启动自己的容器：\n"
        "   $ sudo docker start sudo-app\n"
        "   $ sudo docker ps | grep sudo-app\n\n"
        "3. 重启自己的容器：\n"
        "   $ sudo docker restart sudo-app\n\n"
        "4. 尝试停止 bob 的容器（应被拒绝）：\n"
        "   $ sudo docker stop user-1003-bob-app\n"
        "   # 预期：返回 403 Forbidden\n\n"
        "5. 停止并删除自己的容器：\n"
        "   $ sudo docker stop sudo-app\n"
        "   $ sudo docker rm sudo-app\n"
        "   $ sudo docker ps -a | grep sudo-app\n"
        "   # 预期：无输出\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'stop|start|restart|rm' /var/log/docker-authz/user-operation/test-sudo.log | tail -8",
        "• 自己容器的 stop/start/restart/rm 均成功\n"
        "• 操作 bob 容器返回 403，日志记录 deny_reason=ownership\n"
        "• docker ps -a 中 sudo-app 不再出现（已删除）"
    ),
    (
        "TC-SUDO-007",
        "test-sudo - docker build（允许）",
        "验证 test-sudo 可以构建 Docker 镜像",
        PRE_SUDO,
        "1. 创建构建目录和 Dockerfile：\n"
        "   $ mkdir -p /tmp/sudo-build\n"
        "   $ cat > /tmp/sudo-build/Dockerfile << 'EOF'\n"
        "   FROM alpine:3.18\n"
        "   RUN echo 'built by test-sudo' > /sudo-build-marker.txt\n"
        "   CMD [\"cat\", \"/sudo-build-marker.txt\"]\n"
        "   EOF\n\n"
        "2. 构建镜像：\n"
        "   $ sudo docker build -t sudo-test-img:v1 /tmp/sudo-build/\n"
        "   # 观察构建输出\n\n"
        "3. 验证镜像已创建：\n"
        "   $ sudo docker images sudo-test-img\n\n"
        "4. 运行镜像验证内容：\n"
        "   $ sudo docker run --rm sudo-test-img:v1\n\n"
        "5. 检查审计日志：\n"
        "   # grep 'build' /var/log/docker-authz/user-operation/test-sudo.log | tail -2",
        "• 构建成功，输出 Successfully tagged sudo-test-img:v1\n"
        "• docker images 显示 sudo-test-img:v1\n"
        "• 运行镜像输出 'built by test-sudo'\n"
        "• 审计日志记录 action=build, result=allow, uid=1001"
    ),
    (
        "TC-SUDO-008",
        "test-sudo - docker push（拒绝，策略禁止）",
        "验证 test-sudo 被策略禁止执行 docker push，返回 403",
        PRE_SUDO + "\n• sudo-test-img:v1 镜像已存在",
        "1. 为镜像打标签（准备推送）：\n"
        "   $ sudo docker tag sudo-test-img:v1 registry.example.com/sudo-test-img:v1\n\n"
        "2. 尝试推送镜像（应被拒绝）：\n"
        "   $ sudo docker push registry.example.com/sudo-test-img:v1\n"
        "   # 记录返回的错误信息\n\n"
        "3. 验证错误码为 403：\n"
        "   # 错误信息应包含 'Forbidden' 或 '403'\n\n"
        "4. 检查审计日志（应记录拒绝）：\n"
        "   # grep 'push' /var/log/docker-authz/user-operation/test-sudo.log | tail -2 | python3 -m json.tool",
        "• docker push 返回错误，包含 403 Forbidden\n"
        "• 审计日志记录：action=push, result=deny, deny_reason=policy, status_code=403, uid=1001"
    ),
    (
        "TC-SUDO-009",
        "test-sudo - docker swarm/plugin/secret/config（拒绝，策略禁止）",
        "验证 test-sudo 被策略禁止执行 swarm、plugin、secret、config 操作",
        PRE_SUDO,
        "1. 尝试初始化 swarm（应被拒绝）：\n"
        "   $ sudo docker swarm init\n"
        "   # 预期：403 Forbidden\n\n"
        "2. 尝试列出 plugin（应被拒绝）：\n"
        "   $ sudo docker plugin ls\n"
        "   # 预期：403 Forbidden\n\n"
        "3. 尝试列出 secret（应被拒绝）：\n"
        "   $ sudo docker secret ls\n"
        "   # 预期：403 Forbidden\n\n"
        "4. 尝试列出 config（应被拒绝）：\n"
        "   $ sudo docker config ls\n"
        "   # 预期：403 Forbidden\n\n"
        "5. 检查审计日志（4 条拒绝记录）：\n"
        "   # grep 'deny' /var/log/docker-authz/user-operation/test-sudo.log | tail -8 | python3 -m json.tool",
        "• 4 个命令均返回 403 Forbidden\n"
        "• 审计日志记录 4 条：action=swarm/plugin/secret/config，result=deny，deny_reason=policy"
    ),
    (
        "TC-SUDO-010",
        "test-sudo - docker commit/save/load（允许）",
        "验证 test-sudo 可以提交容器为镜像、导出和导入镜像",
        PRE_SUDO + "\n• sudo-test-img:v1 镜像已存在",
        "1. 创建容器并修改：\n"
        "   $ sudo docker run -d --name sudo-commit-test redis:7\n"
        "   $ sudo docker exec sudo-commit-test sh -c 'echo committed > /tmp/marker'\n\n"
        "2. 提交容器为镜像：\n"
        "   $ sudo docker commit sudo-commit-test sudo-committed:v1\n"
        "   $ sudo docker images sudo-committed\n\n"
        "3. 导出镜像为 tar：\n"
        "   $ sudo docker save sudo-test-img:v1 -o /tmp/sudo-img.tar\n"
        "   $ ls -lh /tmp/sudo-img.tar\n\n"
        "4. 删除镜像后重新导入：\n"
        "   $ sudo docker rmi sudo-test-img:v1\n"
        "   $ sudo docker load -i /tmp/sudo-img.tar\n"
        "   $ sudo docker images sudo-test-img\n\n"
        "5. 清理：\n"
        "   $ sudo docker rm -f sudo-commit-test\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'commit|save|load' /var/log/docker-authz/user-operation/test-sudo.log | tail -5",
        "• commit 成功，docker images 显示 sudo-committed:v1\n"
        "• /tmp/sudo-img.tar 文件存在\n"
        "• load 成功恢复 sudo-test-img:v1\n"
        "• 审计日志记录 action=commit/save/load，result=allow"
    ),
    (
        "TC-SUDO-011",
        "test-sudo - docker network/volume 管理（允许，验证名称前缀）",
        "验证 test-sudo 可以管理自己的网络和卷，且名称自动加前缀",
        PRE_SUDO,
        "1. 创建网络：\n"
        "   $ sudo docker network create mynet\n"
        "   $ sudo docker network ls | grep mynet\n"
        "   # test-sudo 看到的名称应为 mynet\n\n"
        "2. 以 root 身份查看实际网络名：\n"
        "   # docker network ls | grep mynet\n"
        "   # 预期实际名称：test-sudo_u1001_mynet\n\n"
        "3. 创建卷：\n"
        "   $ sudo docker volume create myvol\n"
        "   $ sudo docker volume ls | grep myvol\n\n"
        "4. 验证 bob 的网络不可见：\n"
        "   $ sudo docker network ls | grep bob\n"
        "   # 预期：无输出\n\n"
        "5. 删除网络和卷：\n"
        "   $ sudo docker network rm mynet\n"
        "   $ sudo docker volume rm myvol\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'network|volume' /var/log/docker-authz/user-operation/test-sudo.log | tail -6",
        "• test-sudo 看到的网络名为 mynet（前缀已剥离）\n"
        "• root 看到的实际名称为 test-sudo_u1001_mynet\n"
        "• network ls 只显示 test-sudo 自己的网络\n"
        "• volume ls 只显示 test-sudo 自己的卷"
    ),
    (
        "TC-SUDO-012",
        "test-sudo - 容器数量配额限制（max_containers=10）",
        "验证 test-sudo 创建容器超过 10 个时被拒绝",
        PRE_SUDO + "\n• test-sudo 已有 10 个运行中的容器（通过循环创建）",
        "1. 创建 10 个容器（达到上限）：\n"
        "   $ for i in $(seq 1 10); do\n"
        "       sudo docker run -d --name sudo-quota-$i alpine sleep 3600\n"
        "     done\n"
        "   $ sudo docker ps | wc -l\n"
        "   # 预期：11（含标题行）\n\n"
        "2. 尝试创建第 11 个容器（应被拒绝）：\n"
        "   $ sudo docker run -d --name sudo-quota-11 alpine sleep 3600\n"
        "   # 记录错误信息\n\n"
        "3. 验证错误码：\n"
        "   # 错误信息应包含 403 或 429，说明已达容器数量上限\n\n"
        "4. 检查审计日志：\n"
        "   # grep 'quota\\|max_containers\\|deny' /var/log/docker-authz/user-operation/test-sudo.log | tail -3\n\n"
        "5. 清理：\n"
        "   $ for i in $(seq 1 10); do sudo docker rm -f sudo-quota-$i; done",
        "• 前 10 个容器创建成功\n"
        "• 第 11 个容器返回 403 或 429，错误信息说明已达 max_containers=10 上限\n"
        "• 审计日志记录最后一次 create_container 的 result=deny"
    ),
    (
        "TC-SUDO-013",
        "test-sudo - docker logs/stats/top/inspect（允许）",
        "验证 test-sudo 可以查看自己容器的详细信息",
        PRE_SUDO + "\n• 重新创建 sudo-app 容器（sudo docker run -d --name sudo-app redis:7）",
        "1. 查看容器日志：\n"
        "   $ sudo docker logs sudo-app\n"
        "   $ sudo docker logs --tail 20 sudo-app\n\n"
        "2. 查看资源统计：\n"
        "   $ sudo docker stats sudo-app --no-stream\n"
        "   # 记录 CPU%、MEM USAGE/LIMIT\n\n"
        "3. 查看容器进程：\n"
        "   $ sudo docker top sudo-app\n\n"
        "4. 检查容器详情：\n"
        "   $ sudo docker inspect sudo-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     print('Status:', d['State']['Status'])\n"
        "     print('Image:', d['Config']['Image'])\n"
        "     hc=d['HostConfig']\n"
        "     print('NanoCPUs:', hc['NanoCPUs'])\n"
        "     print('Memory:', hc['Memory'])\"\n\n"
        "5. 检查审计日志：\n"
        "   # grep -E 'logs|inspect' /var/log/docker-authz/user-operation/test-sudo.log | tail -5",
        "• docker logs 返回 redis 启动日志\n"
        "• docker stats 显示 MEM LIMIT 为 4GiB（配额注入）\n"
        "• docker top 返回 redis-server 进程\n"
        "• inspect 显示 NanoCPUs=4000000000，Memory=4294967296\n"
        "• 审计日志记录 action=logs/inspect，result=allow"
    ),
    (
        "TC-SUDO-014",
        "test-sudo - docker cp/diff/rename/pause/unpause（允许）",
        "验证 test-sudo 可以使用容器文件操作和控制命令",
        PRE_SUDO + "\n• sudo-app 容器运行中",
        "1. 文件复制测试：\n"
        "   $ echo 'sudo-test' > /tmp/sudo-cp-test.txt\n"
        "   $ sudo docker cp /tmp/sudo-cp-test.txt sudo-app:/tmp/\n"
        "   $ sudo docker exec sudo-app cat /tmp/sudo-cp-test.txt\n\n"
        "2. 查看文件系统变更：\n"
        "   $ sudo docker diff sudo-app\n"
        "   # 预期：A /tmp/sudo-cp-test.txt\n\n"
        "3. 重命名容器：\n"
        "   $ sudo docker rename sudo-app sudo-app-renamed\n"
        "   $ sudo docker ps | grep sudo-app-renamed\n\n"
        "4. 暂停和恢复容器：\n"
        "   $ sudo docker pause sudo-app-renamed\n"
        "   $ sudo docker ps | grep sudo-app-renamed\n"
        "   # 预期状态：Up (Paused)\n"
        "   $ sudo docker unpause sudo-app-renamed\n"
        "   $ sudo docker ps | grep sudo-app-renamed\n"
        "   # 预期状态：Up\n\n"
        "5. 清理：\n"
        "   $ sudo docker rm -f sudo-app-renamed",
        "• cp 后容器内 cat 返回 'sudo-test'\n"
        "• diff 输出包含 A /tmp/sudo-cp-test.txt\n"
        "• rename 成功，docker ps 显示新名称\n"
        "• pause 后状态为 Up (Paused)，unpause 后恢复 Up"
    ),
    (
        "TC-SUDO-015",
        "test-sudo - docker system prune/df/info（允许）",
        "验证 test-sudo 可以执行系统清理和查看系统信息",
        PRE_SUDO,
        "1. 查看系统信息：\n"
        "   $ sudo docker info\n"
        "   $ sudo docker version\n\n"
        "2. 查看磁盘使用：\n"
        "   $ sudo docker system df\n\n"
        "3. 创建一些待清理资源：\n"
        "   $ sudo docker run --name sudo-prune-1 alpine echo done\n"
        "   $ sudo docker run --name sudo-prune-2 alpine echo done\n\n"
        "4. 执行系统清理：\n"
        "   $ sudo docker system prune -f\n"
        "   # 观察输出\n\n"
        "5. 验证清理结果：\n"
        "   $ sudo docker ps -a | grep Exited\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'prune|info|df' /var/log/docker-authz/user-operation/test-sudo.log | tail -5",
        "• docker info 和 version 返回正常数据\n"
        "• docker system df 返回磁盘使用统计\n"
        "• prune 清理成功，已停止容器被删除\n"
        "• 审计日志记录 action=prune/info/df，result=allow"
    ),
]  # end sudo_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: test-docker-g 用户
# 策略：禁止 build、push、commit、load、save、swarm、plugin、secret、config
# 配额：cpu=2, mem=2048MB, max_containers=5
# ─────────────────────────────────────────────────────────────────────────────
PRE_DGRP = (
    PRE_COMMON + "\n"
    "• 当前操作用户：test-docker-g（uid=1002，docker 组）\n"
    "• 执行方式：su - test-docker-g 后直接执行 docker <cmd>（无需 sudo）\n"
    "• deny_rules：禁止 build、push、commit、load、save、swarm、plugin、secret、config\n"
    "• 配额：cpu_cores=2, mem_mb=2048, max_containers=5"
)

dgrp_cases = [
    (
        "TC-DGRP-001",
        "test-docker-g - docker ps/pull/images（允许，资源隔离）",
        "验证 test-docker-g 可以执行基本查询命令，且资源视图正确隔离",
        PRE_DGRP + "\n• alice 已创建容器和镜像",
        "1. 切换到 test-docker-g 用户：\n"
        "   # su - test-docker-g\n\n"
        "2. 确认 DOCKER_HOST：\n"
        "   $ echo $DOCKER_HOST\n"
        "   # 预期：unix:///run/docker-authz/test-docker-g.sock\n\n"
        "3. 查看容器列表：\n"
        "   $ docker ps\n"
        "   $ docker ps -a\n"
        "   # 验证 alice 的容器不可见\n\n"
        "4. 拉取镜像：\n"
        "   $ docker pull postgres:15\n\n"
        "5. 查看镜像列表：\n"
        "   $ docker images\n"
        "   # 验证只显示自己的镜像和公共镜像\n\n"
        "6. 检查审计日志：\n"
        "   # tail -5 /var/log/docker-authz/user-operation/test-docker-g.log | python3 -m json.tool",
        "• DOCKER_HOST 正确指向 test-docker-g.sock\n"
        "• docker ps 只显示 test-docker-g 自己的容器\n"
        "• postgres:15 拉取成功\n"
        "• docker images 只显示 test-docker-g 的镜像和公共镜像\n"
        "• 审计日志记录 uid=1002, user=test-docker-g"
    ),
    (
        "TC-DGRP-002",
        "test-docker-g - docker run（允许，验证配额注入 2CPU/2048MB）",
        "验证 test-docker-g 创建容器时代理自动注入 2CPU/2048MB 配额",
        PRE_DGRP + "\n• postgres:15 镜像已存在",
        "1. 创建容器（不指定资源限制）：\n"
        "   $ docker run -d --name dgrp-app \\\n"
        "       -e POSTGRES_PASSWORD=test123 \\\n"
        "       postgres:15\n\n"
        "2. 验证容器运行：\n"
        "   $ docker ps | grep dgrp-app\n\n"
        "3. 检查代理注入的资源限制：\n"
        "   $ docker inspect dgrp-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     hc=d['HostConfig']\n"
        "     print('NanoCPUs:', hc['NanoCPUs'], '(期望: 2000000000)')\n"
        "     print('Memory:', hc['Memory'], '(期望: 2147483648)')\"\n\n"
        "4. 检查系统标签：\n"
        "   $ docker inspect dgrp-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     labels=d['Config']['Labels']\n"
        "     for k,v in labels.items():\n"
        "       if 'authz' in k: print(k,'=',v)\"",
        "• 容器创建成功，状态为 Up\n"
        "• NanoCPUs=2000000000（2 CPU），Memory=2147483648（2048MB）\n"
        "• 系统标签：system.authz.owner.uid=1002，system.authz.owner=test-docker-g"
    ),
    (
        "TC-DGRP-003",
        "test-docker-g - docker build（拒绝，策略禁止）",
        "验证 test-docker-g 被策略禁止执行 docker build，返回 403",
        PRE_DGRP,
        "1. 创建 Dockerfile：\n"
        "   $ mkdir -p /tmp/dgrp-build\n"
        "   $ echo 'FROM alpine' > /tmp/dgrp-build/Dockerfile\n\n"
        "2. 尝试构建镜像（应被拒绝）：\n"
        "   $ docker build -t dgrp-img /tmp/dgrp-build/\n"
        "   # 记录错误信息\n\n"
        "3. 验证错误码为 403：\n"
        "   # 错误信息应包含 'Forbidden' 或 '403'\n\n"
        "4. 检查审计日志：\n"
        "   # grep 'build' /var/log/docker-authz/user-operation/test-docker-g.log | tail -2 | python3 -m json.tool",
        "• docker build 返回 403 Forbidden\n"
        "• 审计日志记录：action=build, result=deny, deny_reason=policy, status_code=403, uid=1002"
    ),
    (
        "TC-DGRP-004",
        "test-docker-g - docker push/commit/save/load（拒绝，策略禁止）",
        "验证 test-docker-g 被策略禁止执行 push、commit、save、load 操作",
        PRE_DGRP + "\n• dgrp-app 容器运行中\n• postgres:15 镜像已存在",
        "1. 尝试推送镜像（应被拒绝）：\n"
        "   $ docker push registry.example.com/dgrp-img\n"
        "   # 预期：403 Forbidden\n\n"
        "2. 尝试提交容器（应被拒绝）：\n"
        "   $ docker commit dgrp-app dgrp-committed:v1\n"
        "   # 预期：403 Forbidden\n\n"
        "3. 尝试导出镜像（应被拒绝）：\n"
        "   $ docker save postgres:15 -o /tmp/dgrp-pg.tar\n"
        "   # 预期：403 Forbidden\n\n"
        "4. 尝试导入镜像（应被拒绝）：\n"
        "   $ docker load -i /tmp/some-img.tar\n"
        "   # 预期：403 Forbidden\n\n"
        "5. 检查审计日志（4 条拒绝记录）：\n"
        "   # grep 'deny' /var/log/docker-authz/user-operation/test-docker-g.log | tail -8 | python3 -m json.tool",
        "• 4 个命令均返回 403 Forbidden\n"
        "• 审计日志记录 4 条：action=push/commit/save/load，result=deny，deny_reason=policy"
    ),
    (
        "TC-DGRP-005",
        "test-docker-g - docker exec/logs/stats/inspect/cp（允许）",
        "验证 test-docker-g 可以对自己的容器执行 exec、查看日志、统计、检查和文件复制",
        PRE_DGRP + "\n• dgrp-app 容器运行中",
        "1. exec 进入容器：\n"
        "   $ docker exec dgrp-app echo 'hello from dgrp'\n"
        "   $ docker exec dgrp-app psql -U postgres -c 'SELECT version();'\n\n"
        "2. 查看容器日志：\n"
        "   $ docker logs dgrp-app\n"
        "   $ docker logs --tail 10 dgrp-app\n\n"
        "3. 查看资源统计：\n"
        "   $ docker stats dgrp-app --no-stream\n\n"
        "4. 检查容器详情：\n"
        "   $ docker inspect dgrp-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     print('Status:', d['State']['Status'])\n"
        "     hc=d['HostConfig']\n"
        "     print('NanoCPUs:', hc['NanoCPUs'])\"\n\n"
        "5. 文件复制：\n"
        "   $ echo 'dgrp-test' > /tmp/dgrp-test.txt\n"
        "   $ docker cp /tmp/dgrp-test.txt dgrp-app:/tmp/\n"
        "   $ docker exec dgrp-app cat /tmp/dgrp-test.txt\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'exec|logs|inspect|cp' /var/log/docker-authz/user-operation/test-docker-g.log | tail -6",
        "• exec 返回 'hello from dgrp' 和 PostgreSQL 版本信息\n"
        "• docker logs 返回 postgres 启动日志\n"
        "• docker stats 显示 MEM LIMIT 为 2GiB\n"
        "• 文件复制成功，容器内 cat 返回 'dgrp-test'\n"
        "• 审计日志记录 result=allow"
    ),
    (
        "TC-DGRP-006",
        "test-docker-g - docker tag/rmi（允许，只能操作自己的镜像）",
        "验证 test-docker-g 可以为自己的镜像打标签和删除，但不能删除他人镜像",
        PRE_DGRP + "\n• postgres:15 镜像已存在（test-docker-g 拉取的）\n• alice 有私有镜像",
        "1. 为镜像打标签：\n"
        "   $ docker tag postgres:15 dgrp-pg:custom\n"
        "   $ docker images dgrp-pg\n\n"
        "2. 删除自己的标签：\n"
        "   $ docker rmi dgrp-pg:custom\n"
        "   $ docker images dgrp-pg\n"
        "   # 预期：无输出\n\n"
        "3. 尝试删除 alice 的镜像（应被拒绝）：\n"
        "   # 先以 root 获取 alice 镜像 ID：\n"
        "   # docker images | grep alice\n"
        "   $ docker rmi <alice-image-id>\n"
        "   # 预期：403 Forbidden\n\n"
        "4. 检查审计日志：\n"
        "   # grep -E 'tag|rmi' /var/log/docker-authz/user-operation/test-docker-g.log | tail -4",
        "• docker tag 成功，docker images 显示 dgrp-pg:custom\n"
        "• docker rmi 自己的镜像成功\n"
        "• 删除 alice 镜像返回 403，日志记录 deny_reason=ownership"
    ),
    (
        "TC-DGRP-007",
        "test-docker-g - docker network/volume 管理（允许）",
        "验证 test-docker-g 可以管理自己的网络和卷，且只能看到自己的资源",
        PRE_DGRP + "\n• alice 已创建网络和卷",
        "1. 创建网络：\n"
        "   $ docker network create dgrp-net\n"
        "   $ docker network ls\n"
        "   # 验证只显示自己的网络\n\n"
        "2. 检查网络名称前缀（以 root 验证）：\n"
        "   # docker network ls | grep dgrp\n"
        "   # 预期实际名称：test-docker-g_u1002_dgrp-net\n\n"
        "3. 创建卷：\n"
        "   $ docker volume create dgrp-vol\n"
        "   $ docker volume ls\n"
        "   # 验证只显示自己的卷\n\n"
        "4. 使用卷创建容器：\n"
        "   $ docker run --rm -v dgrp-vol:/data alpine sh -c 'echo dgrp-data > /data/test.txt'\n"
        "   $ docker run --rm -v dgrp-vol:/data alpine cat /data/test.txt\n\n"
        "5. 删除网络和卷：\n"
        "   $ docker network rm dgrp-net\n"
        "   $ docker volume rm dgrp-vol\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'network|volume' /var/log/docker-authz/user-operation/test-docker-g.log | tail -6",
        "• network ls 只显示 test-docker-g 自己的网络\n"
        "• 实际网络名为 test-docker-g_u1002_dgrp-net\n"
        "• volume ls 只显示 test-docker-g 自己的卷\n"
        "• 卷数据读写正常"
    ),
    (
        "TC-DGRP-008",
        "test-docker-g - docker stop/rm/rename/pause（允许，只能操作自己的容器）",
        "验证 test-docker-g 可以管理自己的容器，但不能操作他人容器",
        PRE_DGRP + "\n• dgrp-app 容器运行中\n• alice 的容器 user-1004-alice-app 运行中",
        "1. 暂停自己的容器：\n"
        "   $ docker pause dgrp-app\n"
        "   $ docker ps | grep dgrp-app\n"
        "   # 预期状态：Up (Paused)\n\n"
        "2. 恢复容器：\n"
        "   $ docker unpause dgrp-app\n\n"
        "3. 重命名容器：\n"
        "   $ docker rename dgrp-app dgrp-app-v2\n"
        "   $ docker ps | grep dgrp-app-v2\n\n"
        "4. 停止并删除自己的容器：\n"
        "   $ docker stop dgrp-app-v2\n"
        "   $ docker rm dgrp-app-v2\n\n"
        "5. 尝试停止 alice 的容器（应被拒绝）：\n"
        "   $ docker stop user-1004-alice-app\n"
        "   # 预期：403 Forbidden\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'stop|rm|rename|pause' /var/log/docker-authz/user-operation/test-docker-g.log | tail -8",
        "• 自己容器的所有操作均成功\n"
        "• 操作 alice 容器返回 403，日志记录 deny_reason=ownership\n"
        "• 审计日志记录所有操作的 result"
    ),
    (
        "TC-DGRP-009",
        "test-docker-g - docker swarm/plugin/secret/config（拒绝）",
        "验证 test-docker-g 被策略禁止执行 swarm、plugin、secret、config 操作",
        PRE_DGRP,
        "1. 尝试 swarm init（应被拒绝）：\n"
        "   $ docker swarm init\n"
        "   # 预期：403 Forbidden\n\n"
        "2. 尝试 plugin ls（应被拒绝）：\n"
        "   $ docker plugin ls\n"
        "   # 预期：403 Forbidden\n\n"
        "3. 尝试 secret ls（应被拒绝）：\n"
        "   $ docker secret ls\n"
        "   # 预期：403 Forbidden\n\n"
        "4. 尝试 config ls（应被拒绝）：\n"
        "   $ docker config ls\n"
        "   # 预期：403 Forbidden\n\n"
        "5. 检查审计日志：\n"
        "   # grep 'deny' /var/log/docker-authz/user-operation/test-docker-g.log | tail -8 | python3 -m json.tool",
        "• 4 个命令均返回 403 Forbidden\n"
        "• 审计日志记录 4 条拒绝记录，deny_reason=policy"
    ),
    (
        "TC-DGRP-010",
        "test-docker-g - 容器数量配额限制（max_containers=5）",
        "验证 test-docker-g 创建容器超过 5 个时被拒绝",
        PRE_DGRP + "\n• test-docker-g 已有 5 个运行中的容器",
        "1. 创建 5 个容器（达到上限）：\n"
        "   $ for i in $(seq 1 5); do\n"
        "       docker run -d --name dgrp-quota-$i alpine sleep 3600\n"
        "     done\n"
        "   $ docker ps | wc -l\n"
        "   # 预期：6（含标题行）\n\n"
        "2. 尝试创建第 6 个容器（应被拒绝）：\n"
        "   $ docker run -d --name dgrp-quota-6 alpine sleep 3600\n"
        "   # 记录错误信息\n\n"
        "3. 验证错误码：\n"
        "   # 错误信息应包含 403 或 429\n\n"
        "4. 清理：\n"
        "   $ for i in $(seq 1 5); do docker rm -f dgrp-quota-$i; done",
        "• 前 5 个容器创建成功\n"
        "• 第 6 个容器返回 403 或 429，说明已达 max_containers=5 上限\n"
        "• 审计日志记录最后一次 create_container 的 result=deny"
    ),
]  # end dgrp_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4: bob 用户
# 策略：禁止 exec、build、push、commit、load、save、swarm、plugin、secret、config
# 配额：cpu=1, mem=1024MB, max_containers=3
# ─────────────────────────────────────────────────────────────────────────────
PRE_BOB = (
    PRE_COMMON + "\n"
    "• 当前操作用户：bob（uid=1003，普通用户）\n"
    "• 执行方式：su - bob 后直接执行 docker <cmd>\n"
    "• deny_rules：禁止 exec、build、push、commit、load、save、swarm、plugin、secret、config\n"
    "• 配额：cpu_cores=1, mem_mb=1024, max_containers=3"
)

bob_cases = [
    (
        "TC-BOB-001",
        "bob - docker ps/pull/images（允许，资源隔离）",
        "验证 bob 可以执行基本查询命令，且资源视图正确隔离",
        PRE_BOB + "\n• alice 已创建容器和镜像",
        "1. 切换到 bob 用户：\n"
        "   # su - bob\n\n"
        "2. 确认 DOCKER_HOST：\n"
        "   $ echo $DOCKER_HOST\n"
        "   # 预期：unix:///run/docker-authz/bob.sock\n\n"
        "3. 查看容器列表（应只显示自己的容器）：\n"
        "   $ docker ps\n"
        "   $ docker ps -a\n\n"
        "4. 拉取镜像：\n"
        "   $ docker pull node:18-alpine\n\n"
        "5. 查看镜像列表（应只显示自己的镜像和公共镜像）：\n"
        "   $ docker images\n"
        "   # 验证 alice 的私有镜像不可见\n\n"
        "6. 检查审计日志：\n"
        "   # tail -5 /var/log/docker-authz/user-operation/bob.log | python3 -m json.tool",
        "• DOCKER_HOST 正确指向 bob.sock\n"
        "• docker ps 只显示 bob 自己的容器\n"
        "• node:18-alpine 拉取成功\n"
        "• docker images 只显示 bob 的镜像和公共镜像\n"
        "• 审计日志记录 uid=1003, user=bob"
    ),
    (
        "TC-BOB-002",
        "bob - docker run（允许，验证配额注入 1CPU/1024MB）",
        "验证 bob 创建容器时代理自动注入 1CPU/1024MB 配额",
        PRE_BOB + "\n• node:18-alpine 镜像已存在",
        "1. 创建容器（不指定资源限制）：\n"
        "   $ docker run -d --name bob-app node:18-alpine sleep 3600\n\n"
        "2. 验证容器运行：\n"
        "   $ docker ps | grep bob-app\n\n"
        "3. 检查代理注入的资源限制：\n"
        "   $ docker inspect bob-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     hc=d['HostConfig']\n"
        "     print('NanoCPUs:', hc['NanoCPUs'], '(期望: 1000000000)')\n"
        "     print('Memory:', hc['Memory'], '(期望: 1073741824)')\"\n\n"
        "4. 检查系统标签：\n"
        "   $ docker inspect bob-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     labels=d['Config']['Labels']\n"
        "     for k,v in labels.items():\n"
        "       if 'authz' in k: print(k,'=',v)\"",
        "• 容器创建成功，状态为 Up\n"
        "• NanoCPUs=1000000000（1 CPU），Memory=1073741824（1024MB）\n"
        "• 系统标签：system.authz.owner.uid=1003，system.authz.owner=bob"
    ),
    (
        "TC-BOB-003",
        "bob - docker exec（拒绝，策略禁止）",
        "验证 bob 被策略禁止执行 docker exec 和 docker attach",
        PRE_BOB + "\n• bob-app 容器运行中",
        "1. 尝试 exec 进入自己的容器（应被拒绝）：\n"
        "   $ docker exec bob-app echo hello\n"
        "   # 记录错误信息，预期：403 Forbidden\n\n"
        "2. 尝试 exec 执行命令（应被拒绝）：\n"
        "   $ docker exec bob-app sh -c 'ls /'\n"
        "   # 预期：403 Forbidden\n\n"
        "3. 尝试 attach 容器（应被拒绝）：\n"
        "   $ timeout 3 docker attach bob-app\n"
        "   # 预期：403 Forbidden\n\n"
        "4. 检查审计日志（3 条拒绝记录）：\n"
        "   # grep 'exec' /var/log/docker-authz/user-operation/bob.log | tail -4 | python3 -m json.tool",
        "• 3 个命令均返回 403 Forbidden\n"
        "• 审计日志记录：action=exec, result=deny, deny_reason=policy, status_code=403, uid=1003"
    ),
    (
        "TC-BOB-004",
        "bob - docker build/push/commit/save/load（拒绝，策略禁止）",
        "验证 bob 被策略禁止执行 build、push、commit、save、load 操作",
        PRE_BOB + "\n• bob-app 容器运行中\n• node:18-alpine 镜像已存在",
        "1. 尝试构建镜像（应被拒绝）：\n"
        "   $ mkdir -p /tmp/bob-build && echo 'FROM alpine' > /tmp/bob-build/Dockerfile\n"
        "   $ docker build -t bob-img /tmp/bob-build/\n"
        "   # 预期：403 Forbidden\n\n"
        "2. 尝试推送镜像（应被拒绝）：\n"
        "   $ docker push registry.example.com/bob-img\n"
        "   # 预期：403 Forbidden\n\n"
        "3. 尝试提交容器（应被拒绝）：\n"
        "   $ docker commit bob-app bob-committed:v1\n"
        "   # 预期：403 Forbidden\n\n"
        "4. 尝试导出镜像（应被拒绝）：\n"
        "   $ docker save node:18-alpine -o /tmp/bob-node.tar\n"
        "   # 预期：403 Forbidden\n\n"
        "5. 尝试导入镜像（应被拒绝）：\n"
        "   $ docker load -i /tmp/some-img.tar\n"
        "   # 预期：403 Forbidden\n\n"
        "6. 检查审计日志（5 条拒绝记录）：\n"
        "   # grep 'deny' /var/log/docker-authz/user-operation/bob.log | tail -10 | python3 -m json.tool",
        "• 5 个命令均返回 403 Forbidden\n"
        "• 审计日志记录 5 条：action=build/push/commit/save/load，result=deny，deny_reason=policy"
    ),
    (
        "TC-BOB-005",
        "bob - docker stop/start/restart/rm（允许，只能操作自己的容器）",
        "验证 bob 可以管理自己的容器生命周期，但不能操作他人容器",
        PRE_BOB + "\n• bob-app 容器运行中\n• alice 的容器 user-1004-alice-app 运行中",
        "1. 停止自己的容器：\n"
        "   $ docker stop bob-app\n"
        "   $ docker ps -a | grep bob-app\n"
        "   # 预期状态：Exited\n\n"
        "2. 启动自己的容器：\n"
        "   $ docker start bob-app\n"
        "   $ docker ps | grep bob-app\n"
        "   # 预期状态：Up\n\n"
        "3. 重启自己的容器：\n"
        "   $ docker restart bob-app\n"
        "   $ docker ps | grep bob-app\n\n"
        "4. 尝试停止 alice 的容器（应被拒绝）：\n"
        "   $ docker stop user-1004-alice-app\n"
        "   # 预期：403 Forbidden\n\n"
        "5. 停止并删除自己的容器：\n"
        "   $ docker stop bob-app && docker rm bob-app\n"
        "   $ docker ps -a | grep bob-app\n"
        "   # 预期：无输出\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'stop|start|restart|rm' /var/log/docker-authz/user-operation/bob.log | tail -8",
        "• 自己容器的 stop/start/restart/rm 均成功\n"
        "• 操作 alice 容器返回 403，日志记录 deny_reason=ownership\n"
        "• docker ps -a 中 bob-app 不再出现"
    ),
    (
        "TC-BOB-006",
        "bob - docker logs/stats/top/inspect（允许）",
        "验证 bob 可以查看自己容器的详细信息",
        PRE_BOB + "\n• 重新创建 bob-app（docker run -d --name bob-app node:18-alpine sleep 3600）",
        "1. 查看容器日志：\n"
        "   $ docker logs bob-app\n"
        "   $ docker logs --tail 20 bob-app\n\n"
        "2. 查看资源统计：\n"
        "   $ docker stats bob-app --no-stream\n"
        "   # 记录 MEM USAGE/LIMIT（应显示 1GiB 限制）\n\n"
        "3. 查看容器进程：\n"
        "   $ docker top bob-app\n\n"
        "4. 检查容器详情：\n"
        "   $ docker inspect bob-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     print('Status:', d['State']['Status'])\n"
        "     hc=d['HostConfig']\n"
        "     print('NanoCPUs:', hc['NanoCPUs'])\n"
        "     print('Memory:', hc['Memory'])\"\n\n"
        "5. 检查审计日志：\n"
        "   # grep -E 'logs|inspect' /var/log/docker-authz/user-operation/bob.log | tail -5",
        "• docker logs 返回容器日志\n"
        "• docker stats 显示 MEM LIMIT 为 1GiB（配额注入）\n"
        "• docker top 返回进程列表\n"
        "• inspect 显示 NanoCPUs=1000000000，Memory=1073741824"
    ),
    (
        "TC-BOB-007",
        "bob - docker cp/diff/rename/pause/kill（允许）",
        "验证 bob 可以使用容器文件操作和控制命令",
        PRE_BOB + "\n• bob-app 容器运行中",
        "1. 文件复制：\n"
        "   $ echo 'bob-test-data' > /tmp/bob-test.txt\n"
        "   $ docker cp /tmp/bob-test.txt bob-app:/tmp/\n"
        "   $ docker logs bob-app 2>&1 | head -3\n"
        "   # 注：exec 被禁止，无法直接验证容器内文件，通过 cp 返回码验证\n\n"
        "2. 查看文件系统变更：\n"
        "   $ docker diff bob-app\n"
        "   # 预期：A /tmp/bob-test.txt\n\n"
        "3. 重命名容器：\n"
        "   $ docker rename bob-app bob-app-v2\n"
        "   $ docker ps | grep bob-app-v2\n\n"
        "4. 暂停和恢复容器：\n"
        "   $ docker pause bob-app-v2\n"
        "   $ docker ps | grep bob-app-v2\n"
        "   # 预期状态：Up (Paused)\n"
        "   $ docker unpause bob-app-v2\n\n"
        "5. 强制终止容器：\n"
        "   $ docker kill bob-app-v2\n"
        "   $ docker ps -a | grep bob-app-v2\n"
        "   # 预期状态：Exited(137)\n\n"
        "6. 清理：\n"
        "   $ docker rm bob-app-v2",
        "• cp 命令返回码为 0（成功）\n"
        "• diff 输出包含 A /tmp/bob-test.txt\n"
        "• rename 成功，docker ps 显示新名称\n"
        "• pause 后状态为 Up (Paused)，unpause 后恢复\n"
        "• kill 后状态为 Exited(137)"
    ),
    (
        "TC-BOB-008",
        "bob - docker tag/rmi（允许，只能操作自己的镜像）",
        "验证 bob 可以为自己的镜像打标签和删除",
        PRE_BOB + "\n• node:18-alpine 镜像已存在（bob 拉取的）",
        "1. 为镜像打标签：\n"
        "   $ docker tag node:18-alpine bob-node:custom\n"
        "   $ docker images bob-node\n\n"
        "2. 删除自己的标签：\n"
        "   $ docker rmi bob-node:custom\n"
        "   $ docker images bob-node\n"
        "   # 预期：无输出\n\n"
        "3. 尝试删除 alice 的镜像（应被拒绝）：\n"
        "   $ docker rmi <alice-image-id>\n"
        "   # 预期：403 Forbidden\n\n"
        "4. 检查审计日志：\n"
        "   # grep -E 'tag|rmi' /var/log/docker-authz/user-operation/bob.log | tail -4",
        "• docker tag 成功\n"
        "• docker rmi 自己的镜像成功\n"
        "• 删除 alice 镜像返回 403，日志记录 deny_reason=ownership"
    ),
    (
        "TC-BOB-009",
        "bob - docker network/volume 管理（允许）",
        "验证 bob 可以管理自己的网络和卷，且只能看到自己的资源",
        PRE_BOB,
        "1. 创建网络：\n"
        "   $ docker network create bob-net\n"
        "   $ docker network ls\n"
        "   # 验证只显示自己的网络\n\n"
        "2. 检查网络名称前缀（以 root 验证）：\n"
        "   # docker network ls | grep bob\n"
        "   # 预期实际名称：bob_u1003_bob-net\n\n"
        "3. 创建卷：\n"
        "   $ docker volume create bob-vol\n"
        "   $ docker volume ls\n\n"
        "4. 使用卷创建容器：\n"
        "   $ docker run -d --name bob-vol-test -v bob-vol:/data alpine sleep 60\n"
        "   $ docker ps | grep bob-vol-test\n\n"
        "5. 删除网络和卷：\n"
        "   $ docker rm -f bob-vol-test\n"
        "   $ docker network rm bob-net\n"
        "   $ docker volume rm bob-vol\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'network|volume' /var/log/docker-authz/user-operation/bob.log | tail -6",
        "• network ls 只显示 bob 自己的网络\n"
        "• 实际网络名为 bob_u1003_bob-net\n"
        "• volume ls 只显示 bob 自己的卷\n"
        "• 容器使用卷创建成功"
    ),
    (
        "TC-BOB-010",
        "bob - CPU/内存配额超限被拒绝",
        "验证 bob 请求超过配额的 CPU 或内存时被拒绝",
        PRE_BOB,
        "1. 尝试请求超过 1 CPU（应被拒绝）：\n"
        "   $ docker run --cpus=2 alpine echo hi\n"
        "   # 预期：403 Forbidden，说明 CPU 超出配额\n\n"
        "2. 尝试请求超过 1024MB 内存（应被拒绝）：\n"
        "   $ docker run -m 2g alpine echo hi\n"
        "   # 预期：403 Forbidden，说明内存超出配额\n\n"
        "3. 验证在配额内的请求被允许：\n"
        "   $ docker run --cpus=0.5 -m 512m --rm alpine echo 'within quota'\n"
        "   # 预期：成功执行，输出 'within quota'\n\n"
        "4. 检查审计日志：\n"
        "   # grep 'deny' /var/log/docker-authz/user-operation/bob.log | tail -4 | python3 -m json.tool",
        "• 超过 1 CPU 的请求返回 403\n"
        "• 超过 1024MB 的请求返回 403\n"
        "• 在配额内的请求（0.5 CPU, 512MB）成功执行\n"
        "• 审计日志记录超限请求的 result=deny"
    ),
    (
        "TC-BOB-011",
        "bob - 容器数量配额限制（max_containers=3）",
        "验证 bob 创建容器超过 3 个时被拒绝",
        PRE_BOB,
        "1. 创建 3 个容器（达到上限）：\n"
        "   $ docker run -d --name bob-quota-1 alpine sleep 3600\n"
        "   $ docker run -d --name bob-quota-2 alpine sleep 3600\n"
        "   $ docker run -d --name bob-quota-3 alpine sleep 3600\n"
        "   $ docker ps | wc -l\n"
        "   # 预期：4（含标题行）\n\n"
        "2. 尝试创建第 4 个容器（应被拒绝）：\n"
        "   $ docker run -d --name bob-quota-4 alpine sleep 3600\n"
        "   # 记录错误信息，预期：403 或 429\n\n"
        "3. 验证已有 3 个容器仍在运行：\n"
        "   $ docker ps | grep bob-quota\n\n"
        "4. 检查审计日志：\n"
        "   # grep 'deny' /var/log/docker-authz/user-operation/bob.log | tail -2 | python3 -m json.tool\n\n"
        "5. 清理：\n"
        "   $ docker rm -f bob-quota-1 bob-quota-2 bob-quota-3",
        "• 前 3 个容器创建成功\n"
        "• 第 4 个容器返回 403 或 429，说明已达 max_containers=3 上限\n"
        "• 审计日志记录最后一次 create_container 的 result=deny"
    ),
    (
        "TC-BOB-012",
        "bob - docker system prune/df/info（允许）",
        "验证 bob 可以执行系统清理和查看系统信息",
        PRE_BOB,
        "1. 查看系统信息：\n"
        "   $ docker info\n"
        "   $ docker version\n\n"
        "2. 查看磁盘使用：\n"
        "   $ docker system df\n\n"
        "3. 创建待清理资源：\n"
        "   $ docker run --name bob-prune-1 alpine echo done\n"
        "   $ docker run --name bob-prune-2 alpine echo done\n\n"
        "4. 执行系统清理：\n"
        "   $ docker system prune -f\n"
        "   # 观察输出\n\n"
        "5. 验证清理结果：\n"
        "   $ docker ps -a | grep Exited\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'prune|info|df' /var/log/docker-authz/user-operation/bob.log | tail -5",
        "• docker info 和 version 返回正常数据\n"
        "• prune 清理成功，已停止容器被删除\n"
        "• 审计日志记录 action=prune/info/df，result=allow"
    ),
    (
        "TC-BOB-013",
        "bob - docker swarm/plugin/secret/config（拒绝）",
        "验证 bob 被策略禁止执行 swarm、plugin、secret、config 操作",
        PRE_BOB,
        "1. 尝试 swarm init（应被拒绝）：\n"
        "   $ docker swarm init\n"
        "   # 预期：403 Forbidden\n\n"
        "2. 尝试 plugin ls（应被拒绝）：\n"
        "   $ docker plugin ls\n"
        "   # 预期：403 Forbidden\n\n"
        "3. 尝试 secret ls（应被拒绝）：\n"
        "   $ docker secret ls\n"
        "   # 预期：403 Forbidden\n\n"
        "4. 检查审计日志：\n"
        "   # grep 'deny' /var/log/docker-authz/user-operation/bob.log | tail -6 | python3 -m json.tool",
        "• 3 个命令均返回 403 Forbidden\n"
        "• 审计日志记录 3 条拒绝记录，deny_reason=policy"
    ),
]  # end bob_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5: alice 用户（最严格策略）
# 策略：禁止 exec、build、push、commit、load、save、prune、swarm、plugin、secret、config
# 配额：cpu=1, mem=512MB, max_containers=2
# ─────────────────────────────────────────────────────────────────────────────
PRE_ALICE = (
    PRE_COMMON + "\n"
    "• 当前操作用户：alice（uid=1004，普通用户）\n"
    "• 执行方式：su - alice 后直接执行 docker <cmd>\n"
    "• deny_rules：禁止 exec、build、push、commit、load、save、prune、swarm、plugin、secret、config\n"
    "• 配额：cpu_cores=1, mem_mb=512, max_containers=2"
)

alice_cases = [
    (
        "TC-ALICE-001",
        "alice - docker ps/pull/images（允许，资源隔离）",
        "验证 alice 可以执行基本查询命令，且资源视图正确隔离",
        PRE_ALICE + "\n• bob 已创建容器和镜像",
        "1. 切换到 alice 用户：\n"
        "   # su - alice\n\n"
        "2. 确认 DOCKER_HOST：\n"
        "   $ echo $DOCKER_HOST\n"
        "   # 预期：unix:///run/docker-authz/alice.sock\n\n"
        "3. 查看容器列表（应只显示自己的容器）：\n"
        "   $ docker ps\n"
        "   $ docker ps -a\n"
        "   # 验证 bob 的容器不可见\n\n"
        "4. 拉取镜像：\n"
        "   $ docker pull python:3.11-slim\n\n"
        "5. 查看镜像列表（应只显示自己的镜像和公共镜像）：\n"
        "   $ docker images\n"
        "   # 验证 bob 的私有镜像不可见\n\n"
        "6. 检查审计日志：\n"
        "   # tail -5 /var/log/docker-authz/user-operation/alice.log | python3 -m json.tool",
        "• DOCKER_HOST 正确指向 alice.sock\n"
        "• docker ps 只显示 alice 自己的容器\n"
        "• python:3.11-slim 拉取成功\n"
        "• docker images 只显示 alice 的镜像和公共镜像\n"
        "• 审计日志记录 uid=1004, user=alice"
    ),
    (
        "TC-ALICE-002",
        "alice - docker run（允许，验证配额注入 1CPU/512MB）",
        "验证 alice 创建容器时代理自动注入 1CPU/512MB 配额",
        PRE_ALICE + "\n• python:3.11-slim 镜像已存在",
        "1. 创建容器（不指定资源限制）：\n"
        "   $ docker run -d --name alice-app python:3.11-slim sleep 3600\n\n"
        "2. 验证容器运行：\n"
        "   $ docker ps | grep alice-app\n\n"
        "3. 检查代理注入的资源限制：\n"
        "   $ docker inspect alice-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     hc=d['HostConfig']\n"
        "     print('NanoCPUs:', hc['NanoCPUs'], '(期望: 1000000000)')\n"
        "     print('Memory:', hc['Memory'], '(期望: 536870912)')\"\n\n"
        "4. 检查系统标签：\n"
        "   $ docker inspect alice-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     labels=d['Config']['Labels']\n"
        "     for k,v in labels.items():\n"
        "       if 'authz' in k: print(k,'=',v)\"",
        "• 容器创建成功，状态为 Up\n"
        "• NanoCPUs=1000000000（1 CPU），Memory=536870912（512MB）\n"
        "• 系统标签：system.authz.owner.uid=1004，system.authz.owner=alice"
    ),
    (
        "TC-ALICE-003",
        "alice - docker exec/build/push/commit/save/load/prune（全部拒绝）",
        "验证 alice 被策略禁止执行所有受限操作",
        PRE_ALICE + "\n• alice-app 容器运行中\n• python:3.11-slim 镜像已存在",
        "1. 尝试 exec（应被拒绝）：\n"
        "   $ docker exec alice-app echo hello\n"
        "   # 预期：403 Forbidden\n\n"
        "2. 尝试 build（应被拒绝）：\n"
        "   $ mkdir -p /tmp/alice-build && echo 'FROM alpine' > /tmp/alice-build/Dockerfile\n"
        "   $ docker build -t alice-img /tmp/alice-build/\n"
        "   # 预期：403 Forbidden\n\n"
        "3. 尝试 push（应被拒绝）：\n"
        "   $ docker push registry.example.com/alice-img\n"
        "   # 预期：403 Forbidden\n\n"
        "4. 尝试 commit（应被拒绝）：\n"
        "   $ docker commit alice-app alice-committed:v1\n"
        "   # 预期：403 Forbidden\n\n"
        "5. 尝试 save（应被拒绝）：\n"
        "   $ docker save python:3.11-slim -o /tmp/alice-py.tar\n"
        "   # 预期：403 Forbidden\n\n"
        "6. 尝试 load（应被拒绝）：\n"
        "   $ docker load -i /tmp/some-img.tar\n"
        "   # 预期：403 Forbidden\n\n"
        "7. 尝试 system prune（应被拒绝）：\n"
        "   $ docker system prune -f\n"
        "   # 预期：403 Forbidden\n\n"
        "8. 检查审计日志（7 条拒绝记录）：\n"
        "   # grep 'deny' /var/log/docker-authz/user-operation/alice.log | tail -14 | python3 -m json.tool",
        "• 7 个命令均返回 403 Forbidden\n"
        "• 审计日志记录 7 条：action=exec/build/push/commit/save/load/prune，result=deny，deny_reason=policy"
    ),
    (
        "TC-ALICE-004",
        "alice - docker stop/start/restart/rm（允许，只能操作自己的容器）",
        "验证 alice 可以管理自己的容器生命周期，但不能操作他人容器",
        PRE_ALICE + "\n• alice-app 容器运行中\n• bob 的容器 user-1003-bob-app 运行中",
        "1. 停止自己的容器：\n"
        "   $ docker stop alice-app\n"
        "   $ docker ps -a | grep alice-app\n"
        "   # 预期状态：Exited\n\n"
        "2. 启动自己的容器：\n"
        "   $ docker start alice-app\n"
        "   $ docker ps | grep alice-app\n\n"
        "3. 重启自己的容器：\n"
        "   $ docker restart alice-app\n\n"
        "4. 尝试停止 bob 的容器（应被拒绝）：\n"
        "   $ docker stop user-1003-bob-app\n"
        "   # 预期：403 Forbidden\n\n"
        "5. 停止并删除自己的容器：\n"
        "   $ docker stop alice-app && docker rm alice-app\n"
        "   $ docker ps -a | grep alice-app\n"
        "   # 预期：无输出\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'stop|start|restart|rm' /var/log/docker-authz/user-operation/alice.log | tail -8",
        "• 自己容器的 stop/start/restart/rm 均成功\n"
        "• 操作 bob 容器返回 403，日志记录 deny_reason=ownership"
    ),
    (
        "TC-ALICE-005",
        "alice - docker logs/stats/inspect/cp/diff（允许）",
        "验证 alice 可以查看自己容器的信息和执行文件操作",
        PRE_ALICE + "\n• 重新创建 alice-app（docker run -d --name alice-app python:3.11-slim sleep 3600）",
        "1. 查看容器日志：\n"
        "   $ docker logs alice-app\n"
        "   $ docker logs --tail 10 alice-app\n\n"
        "2. 查看资源统计：\n"
        "   $ docker stats alice-app --no-stream\n"
        "   # 记录 MEM USAGE/LIMIT（应显示 512MiB 限制）\n\n"
        "3. 检查容器详情：\n"
        "   $ docker inspect alice-app | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     hc=d['HostConfig']\n"
        "     print('NanoCPUs:', hc['NanoCPUs'])\n"
        "     print('Memory:', hc['Memory'])\"\n\n"
        "4. 文件复制（注意：exec 被禁止，无法验证容器内文件）：\n"
        "   $ echo 'alice-data' > /tmp/alice-test.txt\n"
        "   $ docker cp /tmp/alice-test.txt alice-app:/tmp/\n"
        "   # 验证 cp 命令返回码为 0\n\n"
        "5. 查看文件系统变更：\n"
        "   $ docker diff alice-app\n"
        "   # 预期：A /tmp/alice-test.txt\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'logs|inspect|cp' /var/log/docker-authz/user-operation/alice.log | tail -5",
        "• docker logs 返回容器日志\n"
        "• docker stats 显示 MEM LIMIT 为 512MiB（配额注入）\n"
        "• inspect 显示 NanoCPUs=1000000000，Memory=536870912\n"
        "• cp 命令返回码为 0\n"
        "• diff 输出包含 A /tmp/alice-test.txt"
    ),
    (
        "TC-ALICE-006",
        "alice - docker tag/rmi（允许，只能操作自己的镜像）",
        "验证 alice 可以为自己的镜像打标签和删除",
        PRE_ALICE + "\n• python:3.11-slim 镜像已存在（alice 拉取的）",
        "1. 为镜像打标签：\n"
        "   $ docker tag python:3.11-slim alice-py:custom\n"
        "   $ docker images alice-py\n\n"
        "2. 删除自己的标签：\n"
        "   $ docker rmi alice-py:custom\n"
        "   $ docker images alice-py\n"
        "   # 预期：无输出\n\n"
        "3. 尝试删除 bob 的镜像（应被拒绝）：\n"
        "   $ docker rmi <bob-image-id>\n"
        "   # 预期：403 Forbidden\n\n"
        "4. 检查审计日志：\n"
        "   # grep -E 'tag|rmi' /var/log/docker-authz/user-operation/alice.log | tail -4",
        "• docker tag 成功\n"
        "• docker rmi 自己的镜像成功\n"
        "• 删除 bob 镜像返回 403，日志记录 deny_reason=ownership"
    ),
    (
        "TC-ALICE-007",
        "alice - docker network/volume 管理（允许）",
        "验证 alice 可以管理自己的网络和卷",
        PRE_ALICE,
        "1. 创建网络：\n"
        "   $ docker network create alice-net\n"
        "   $ docker network ls\n"
        "   # 验证只显示自己的网络\n\n"
        "2. 检查网络名称前缀（以 root 验证）：\n"
        "   # docker network ls | grep alice\n"
        "   # 预期实际名称：alice_u1004_alice-net\n\n"
        "3. 创建卷：\n"
        "   $ docker volume create alice-vol\n"
        "   $ docker volume ls\n\n"
        "4. 删除网络和卷：\n"
        "   $ docker network rm alice-net\n"
        "   $ docker volume rm alice-vol\n\n"
        "5. 检查审计日志：\n"
        "   # grep -E 'network|volume' /var/log/docker-authz/user-operation/alice.log | tail -6",
        "• network ls 只显示 alice 自己的网络\n"
        "• 实际网络名为 alice_u1004_alice-net\n"
        "• volume ls 只显示 alice 自己的卷\n"
        "• 删除操作均成功"
    ),
    (
        "TC-ALICE-008",
        "alice - 内存配额超限被拒绝（max 512MB）",
        "验证 alice 请求超过 512MB 内存时被拒绝",
        PRE_ALICE,
        "1. 尝试请求超过 512MB 内存（应被拒绝）：\n"
        "   $ docker run -m 1g alpine echo hi\n"
        "   # 预期：403 Forbidden\n\n"
        "2. 尝试请求超过 1 CPU（应被拒绝）：\n"
        "   $ docker run --cpus=2 alpine echo hi\n"
        "   # 预期：403 Forbidden\n\n"
        "3. 验证在配额内的请求被允许：\n"
        "   $ docker run --cpus=0.5 -m 256m --rm alpine echo 'within quota'\n"
        "   # 预期：成功执行，输出 'within quota'\n\n"
        "4. 检查审计日志：\n"
        "   # grep 'deny' /var/log/docker-authz/user-operation/alice.log | tail -4 | python3 -m json.tool",
        "• 超过 512MB 的请求返回 403\n"
        "• 超过 1 CPU 的请求返回 403\n"
        "• 在配额内的请求成功执行\n"
        "• 审计日志记录超限请求的 result=deny"
    ),
    (
        "TC-ALICE-009",
        "alice - 容器数量配额限制（max_containers=2）",
        "验证 alice 创建容器超过 2 个时被拒绝",
        PRE_ALICE,
        "1. 创建 2 个容器（达到上限）：\n"
        "   $ docker run -d --name alice-quota-1 alpine sleep 3600\n"
        "   $ docker run -d --name alice-quota-2 alpine sleep 3600\n"
        "   $ docker ps | wc -l\n"
        "   # 预期：3（含标题行）\n\n"
        "2. 尝试创建第 3 个容器（应被拒绝）：\n"
        "   $ docker run -d --name alice-quota-3 alpine sleep 3600\n"
        "   # 记录错误信息，预期：403 或 429\n\n"
        "3. 验证已有 2 个容器仍在运行：\n"
        "   $ docker ps | grep alice-quota\n\n"
        "4. 检查审计日志：\n"
        "   # grep 'deny' /var/log/docker-authz/user-operation/alice.log | tail -2 | python3 -m json.tool\n\n"
        "5. 清理：\n"
        "   $ docker rm -f alice-quota-1 alice-quota-2",
        "• 前 2 个容器创建成功\n"
        "• 第 3 个容器返回 403 或 429，说明已达 max_containers=2 上限\n"
        "• 审计日志记录最后一次 create_container 的 result=deny"
    ),
    (
        "TC-ALICE-010",
        "alice - 系统标签防篡改验证",
        "验证 alice 无法通过指定标签覆盖 system.authz.owner.uid",
        PRE_ALICE,
        "1. 尝试创建容器时覆盖系统标签：\n"
        "   $ docker run -d --name alice-tamper \\\n"
        "       -l system.authz.owner.uid=9999 \\\n"
        "       -l system.authz.owner=hacker \\\n"
        "       alpine sleep 3600\n\n"
        "2. 检查实际标签值：\n"
        "   $ docker inspect alice-tamper | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     labels=d['Config']['Labels']\n"
        "     print('所有 authz 标签：')\n"
        "     for k,v in labels.items():\n"
        "       if 'authz' in k: print(' ', k, '=', v)\"\n\n"
        "3. 验证 owner.uid 为真实值（1004）而非伪造值（9999）：\n"
        "   $ docker inspect alice-tamper | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     uid=d['Config']['Labels'].get('system.authz.owner.uid')\n"
        "     print('owner.uid:', uid)\n"
        "     assert uid == '1004', f'防篡改失败！uid={uid}'\n"
        "     print('防篡改验证通过')\"\n\n"
        "4. 清理：\n"
        "   $ docker rm -f alice-tamper",
        "• 容器创建成功（不因标签被拒绝）\n"
        "• system.authz.owner.uid=1004（代理注入的真实值），用户指定的 9999 被覆盖\n"
        "• system.authz.owner 值为 hacker,alice（用户值在前，代理追加在后）\n"
        "• Python 断言通过，输出 '防篡改验证通过'"
    ),
    (
        "TC-ALICE-011",
        "alice - docker info/version/search（允许）",
        "验证 alice 可以查看系统信息和搜索镜像",
        PRE_ALICE,
        "1. 查看系统信息：\n"
        "   $ docker info\n"
        "   $ docker version\n\n"
        "2. 搜索 Docker Hub 镜像：\n"
        "   $ docker search python --limit 5\n"
        "   $ docker search --filter is-official=true python --limit 3\n\n"
        "3. 查看镜像历史：\n"
        "   $ docker history python:3.11-slim\n\n"
        "4. 检查审计日志：\n"
        "   # grep -E 'info|search' /var/log/docker-authz/user-operation/alice.log | tail -4",
        "• docker info 和 version 返回正常数据\n"
        "• docker search 返回搜索结果列表\n"
        "• docker history 返回镜像层历史\n"
        "• 审计日志记录 action=info/search，result=allow"
    ),
    (
        "TC-ALICE-012",
        "alice - docker pause/kill/rename（允许）",
        "验证 alice 可以使用容器控制命令",
        PRE_ALICE + "\n• alice-app 容器运行中",
        "1. 暂停容器：\n"
        "   $ docker pause alice-app\n"
        "   $ docker ps | grep alice-app\n"
        "   # 预期状态：Up (Paused)\n\n"
        "2. 恢复容器：\n"
        "   $ docker unpause alice-app\n"
        "   $ docker ps | grep alice-app\n"
        "   # 预期状态：Up\n\n"
        "3. 重命名容器：\n"
        "   $ docker rename alice-app alice-app-v2\n"
        "   $ docker ps | grep alice-app-v2\n\n"
        "4. 强制终止容器：\n"
        "   $ docker kill alice-app-v2\n"
        "   $ docker ps -a | grep alice-app-v2\n"
        "   # 预期状态：Exited(137)\n\n"
        "5. 清理：\n"
        "   $ docker rm alice-app-v2\n\n"
        "6. 检查审计日志：\n"
        "   # grep -E 'pause|kill|rename' /var/log/docker-authz/user-operation/alice.log | tail -5",
        "• pause 后状态为 Up (Paused)\n"
        "• unpause 后恢复 Up\n"
        "• rename 成功\n"
        "• kill 后状态为 Exited(137)\n"
        "• 审计日志记录所有操作，result=allow"
    ),
]  # end alice_cases

PRE_CROSS = (
    "• 代理服务运行中\n"
    "• 用户 bob (uid=1003) 和 alice (uid=1004) 均已创建\n"
    "• 各用户通过各自 socket 连接代理\n"
    "• 测试前确认两用户均无残留容器"
)

cross_cases = [
    (
        "TC-CROSS-001",
        "跨用户容器不可见性验证",
        "验证 bob 无法看到 alice 的容器，alice 无法看到 bob 的容器",
        PRE_CROSS,
        "1. alice 创建容器：\n"
        "   $ su - alice -c 'docker run -d --name alice-secret nginx:alpine'\n\n"
        "2. bob 列出容器：\n"
        "   $ su - bob -c 'docker ps -a'\n"
        "   # 预期：列表中不出现 alice-secret 或 alice-secret 的任何变体\n\n"
        "3. bob 尝试直接访问 alice 容器：\n"
        "   $ su - bob -c 'docker inspect alice-secret'\n"
        "   # 预期：Error: No such container: alice-secret\n\n"
        "4. 验证实际存储的容器名（root 视角）：\n"
        "   # docker ps -a --format '{{.Names}}'\n"
        "   # 预期：显示 user-1004-alice-secret（带前缀）\n\n"
        "5. alice 列出自己的容器：\n"
        "   $ su - alice -c 'docker ps -a'\n"
        "   # 预期：显示 alice-secret（前缀已过滤）\n\n"
        "6. 清理：\n"
        "   $ su - alice -c 'docker rm -f alice-secret'",
        "• bob 的 docker ps 不显示 alice 的容器\n"
        "• bob inspect alice 容器返回 404\n"
        "• root 视角可见带前缀的真实名称\n"
        "• alice 自己可见（前缀过滤后）"
    ),
    (
        "TC-CROSS-002",
        "跨用户容器操作拒绝验证",
        "验证 bob 无法 stop/rm/exec 属于 alice 的容器",
        PRE_CROSS,
        "1. alice 启动容器：\n"
        "   $ su - alice -c 'docker run -d --name alice-web nginx:alpine'\n\n"
        "2. bob 尝试停止 alice 容器：\n"
        "   $ su - bob -c 'docker stop alice-web'\n"
        "   # 预期：Error: No such container: alice-web\n\n"
        "3. bob 尝试删除 alice 容器：\n"
        "   $ su - bob -c 'docker rm alice-web'\n"
        "   # 预期：Error: No such container: alice-web\n\n"
        "4. bob 尝试 exec 进入 alice 容器：\n"
        "   $ su - bob -c 'docker exec alice-web ls'\n"
        "   # 预期：Error: No such container: alice-web\n\n"
        "5. 验证 alice 容器仍在运行：\n"
        "   $ su - alice -c 'docker ps | grep alice-web'\n"
        "   # 预期：Up\n\n"
        "6. 检查审计日志（bob 的操作被拒绝）：\n"
        "   # grep 'alice-web' /var/log/docker-authz/user-operation/bob.log\n"
        "   # 预期：result=deny 或无记录（容器不存在）\n\n"
        "7. 清理：\n"
        "   $ su - alice -c 'docker rm -f alice-web'",
        "• bob 所有跨用户操作均返回容器不存在\n"
        "• alice 容器不受影响，持续运行\n"
        "• 审计日志记录 bob 的越权尝试"
    ),
    (
        "TC-CROSS-003",
        "跨用户网络隔离验证",
        "验证 bob 无法看到或使用 alice 的网络",
        PRE_CROSS,
        "1. alice 创建网络：\n"
        "   $ su - alice -c 'docker network create alice-net'\n\n"
        "2. bob 列出网络：\n"
        "   $ su - bob -c 'docker network ls'\n"
        "   # 预期：不显示 alice-net 或其带前缀版本\n\n"
        "3. bob 尝试连接 alice 网络：\n"
        "   $ su - bob -c 'docker network connect alice-net $(docker run -d nginx:alpine)'\n"
        "   # 预期：Error: network alice-net not found\n\n"
        "4. 验证 alice 网络真实名称（root 视角）：\n"
        "   # docker network ls | grep alice\n"
        "   # 预期：显示 alice_u1004_alice-net\n\n"
        "5. alice 查看自己的网络：\n"
        "   $ su - alice -c 'docker network ls'\n"
        "   # 预期：显示 alice-net（前缀过滤）\n\n"
        "6. 清理：\n"
        "   $ su - alice -c 'docker network rm alice-net'",
        "• bob 网络列表不含 alice 的网络\n"
        "• bob 连接 alice 网络失败\n"
        "• alice 自己可见（前缀过滤后）"
    ),
    (
        "TC-CROSS-004",
        "跨用户卷隔离验证",
        "验证 bob 无法看到或使用 alice 的数据卷",
        PRE_CROSS,
        "1. alice 创建卷：\n"
        "   $ su - alice -c 'docker volume create alice-data'\n\n"
        "2. bob 列出卷：\n"
        "   $ su - bob -c 'docker volume ls'\n"
        "   # 预期：不显示 alice-data\n\n"
        "3. bob 尝试挂载 alice 的卷：\n"
        "   $ su - bob -c 'docker run --rm -v alice-data:/mnt alpine ls /mnt'\n"
        "   # 预期：Error: volume alice-data not found 或创建了新的 bob 卷\n\n"
        "4. 验证 alice 卷真实名称（root 视角）：\n"
        "   # docker volume ls | grep alice\n"
        "   # 预期：显示 alice_u1004_alice-data\n\n"
        "5. alice 查看自己的卷：\n"
        "   $ su - alice -c 'docker volume ls'\n"
        "   # 预期：显示 alice-data（前缀过滤）\n\n"
        "6. 清理：\n"
        "   $ su - alice -c 'docker volume rm alice-data'",
        "• bob 卷列表不含 alice 的卷\n"
        "• bob 无法挂载 alice 的卷\n"
        "• alice 自己可见（前缀过滤后）"
    ),
    (
        "TC-CROSS-005",
        "容器所有权标签防篡改验证",
        "验证用户无法通过标签伪造所有权",
        PRE_CROSS,
        "1. bob 尝试创建带伪造 owner 标签的容器：\n"
        "   $ su - bob -c 'docker run -d \\\n"
        "       --label system.authz.owner.uid=1004 \\\n"
        "       --name bob-fake nginx:alpine'\n\n"
        "2. 检查实际标签（代理应覆盖）：\n"
        "   # docker inspect user-1003-bob-fake | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     labels = d['Config']['Labels']\n"
        "     print('owner.uid:', labels.get('system.authz.owner.uid'))\n"
        "     print('期望: 1003 (bob 的真实 uid)')\"\n\n"
        "3. 验证 alice 无法看到该容器：\n"
        "   $ su - alice -c 'docker ps -a | grep bob-fake'\n"
        "   # 预期：无输出\n\n"
        "4. 验证 bob 可以看到该容器（所有权正确）：\n"
        "   $ su - bob -c 'docker ps -a | grep bob-fake'\n"
        "   # 预期：显示 bob-fake\n\n"
        "5. 清理：\n"
        "   $ su - bob -c 'docker rm -f bob-fake'",
        "• 代理覆盖伪造的 owner.uid 为真实 uid=1003\n"
        "• alice 无法看到 bob 的容器\n"
        "• bob 正常可见自己的容器"
    ),
    (
        "TC-CROSS-006",
        "root 用户可见所有用户容器",
        "验证 root 可以管理所有用户的容器（带前缀）",
        PRE_CROSS + "\n• 以 root 身份操作",
        "1. bob 和 alice 各创建一个容器：\n"
        "   $ su - bob -c 'docker run -d --name bob-svc nginx:alpine'\n"
        "   $ su - alice -c 'docker run -d --name alice-svc nginx:alpine'\n\n"
        "2. root 列出所有容器：\n"
        "   # docker ps -a\n"
        "   # 预期：显示 user-1003-bob-svc 和 user-1004-alice-svc\n\n"
        "3. root 停止 bob 的容器：\n"
        "   # docker stop user-1003-bob-svc\n"
        "   # 预期：成功\n\n"
        "4. root 删除 alice 的容器：\n"
        "   # docker rm -f user-1004-alice-svc\n"
        "   # 预期：成功\n\n"
        "5. 验证操作结果：\n"
        "   # docker ps -a | grep -E 'bob-svc|alice-svc'\n"
        "   # 预期：bob-svc 状态 Exited，alice-svc 已删除\n\n"
        "6. 清理：\n"
        "   # docker rm user-1003-bob-svc",
        "• root 可见所有用户容器（带前缀）\n"
        "• root 可以操作任意用户的容器\n"
        "• 操作成功，无权限拒绝"
    ),
    (
        "TC-CROSS-007",
        "sudo 用户可见所有容器验证",
        "验证 test-sudo 用户（eUID=0）可以看到所有容器",
        PRE_CROSS + "\n• test-sudo 用户已配置",
        "1. bob 创建容器：\n"
        "   $ su - bob -c 'docker run -d --name bob-visible nginx:alpine'\n\n"
        "2. test-sudo 列出所有容器：\n"
        "   $ su - test-sudo -c 'sudo docker ps -a'\n"
        "   # 预期：显示 user-1003-bob-visible（root 视角）\n\n"
        "3. 验证 test-sudo 的 loginUID 被正确识别：\n"
        "   # grep 'test-sudo' /var/log/docker-authz/auth.log | tail -3\n"
        "   # 预期：effective_uid=0, real_uid=1001\n\n"
        "4. 清理：\n"
        "   $ su - bob -c 'docker rm -f bob-visible'",
        "• test-sudo 以 sudo 运行时可见所有容器\n"
        "• auth.log 记录 effective_uid=0, real_uid=1001\n"
        "• 容器名带前缀（root 视角）"
    ),
    (
        "TC-CROSS-008",
        "并发多用户操作隔离验证",
        "验证多用户同时操作时互不干扰",
        PRE_CROSS,
        "1. 并发启动多个用户的容器：\n"
        "   $ su - bob -c 'docker run -d --name bob-c1 nginx:alpine' &\n"
        "   $ su - alice -c 'docker run -d --name alice-c1 nginx:alpine' &\n"
        "   $ su - bob -c 'docker run -d --name bob-c2 nginx:alpine' &\n"
        "   $ wait\n"
        "   # 等待所有后台任务完成\n\n"
        "2. 验证各用户只看到自己的容器：\n"
        "   $ su - bob -c 'docker ps -a'\n"
        "   # 预期：bob-c1, bob-c2（不含 alice-c1）\n"
        "   $ su - alice -c 'docker ps -a'\n"
        "   # 预期：alice-c1（不含 bob-c1, bob-c2）\n\n"
        "3. 验证容器数量：\n"
        "   $ su - bob -c 'docker ps -a | grep -c bob'\n"
        "   # 预期：2\n"
        "   $ su - alice -c 'docker ps -a | grep -c alice'\n"
        "   # 预期：1\n\n"
        "4. 清理：\n"
        "   $ su - bob -c 'docker rm -f bob-c1 bob-c2'\n"
        "   $ su - alice -c 'docker rm -f alice-c1'",
        "• 并发操作后各用户容器隔离正确\n"
        "• bob 看到 2 个容器，alice 看到 1 个\n"
        "• 无容器泄露到其他用户"
    ),
    (
        "TC-CROSS-009",
        "镜像列表隔离验证",
        "验证普通用户只能看到公共镜像，不泄露其他用户私有镜像信息",
        PRE_CROSS,
        "1. bob 列出镜像：\n"
        "   $ su - bob -c 'docker images'\n"
        "   # 记录当前镜像列表\n\n"
        "2. alice 列出镜像：\n"
        "   $ su - alice -c 'docker images'\n"
        "   # 预期：与 bob 看到的公共镜像相同\n\n"
        "3. root 列出所有镜像：\n"
        "   # docker images\n"
        "   # 预期：包含所有镜像（含私有）\n\n"
        "4. 验证镜像列表一致性：\n"
        "   $ su - bob -c 'docker images -q' > /tmp/bob_images.txt\n"
        "   $ su - alice -c 'docker images -q' > /tmp/alice_images.txt\n"
        "   $ diff /tmp/bob_images.txt /tmp/alice_images.txt\n"
        "   # 预期：无差异（公共镜像相同）",
        "• bob 和 alice 看到相同的公共镜像列表\n"
        "• 无私有镜像信息泄露\n"
        "• root 可见所有镜像"
    ),
    (
        "TC-CROSS-010",
        "SQLite 所有权数据库一致性验证",
        "验证容器所有权数据库在多用户操作后保持一致",
        PRE_CROSS + "\n• 代理使用 SQLite 存储容器所有权",
        "1. bob 创建容器：\n"
        "   $ su - bob -c 'docker run -d --name bob-db-test nginx:alpine'\n\n"
        "2. 查询 SQLite 数据库：\n"
        "   # sqlite3 /var/lib/docker-authz/ownership.db \\\n"
        "       \"SELECT container_name, owner_uid FROM containers WHERE owner_uid=1003;\"\n"
        "   # 预期：user-1003-bob-db-test | 1003\n\n"
        "3. alice 创建容器：\n"
        "   $ su - alice -c 'docker run -d --name alice-db-test nginx:alpine'\n\n"
        "4. 再次查询数据库：\n"
        "   # sqlite3 /var/lib/docker-authz/ownership.db \\\n"
        "       \"SELECT container_name, owner_uid FROM containers ORDER BY owner_uid;\"\n"
        "   # 预期：两条记录，uid 分别为 1003 和 1004\n\n"
        "5. bob 删除容器后验证数据库清理：\n"
        "   $ su - bob -c 'docker rm -f bob-db-test'\n"
        "   # sqlite3 /var/lib/docker-authz/ownership.db \\\n"
        "       \"SELECT * FROM containers WHERE container_name='user-1003-bob-db-test';\"\n"
        "   # 预期：无记录（已清理）\n\n"
        "6. 清理：\n"
        "   $ su - alice -c 'docker rm -f alice-db-test'",
        "• SQLite 正确记录容器所有权\n"
        "• 删除容器后数据库记录同步清理\n"
        "• 多用户记录互不干扰"
    ),
    (
        "TC-CROSS-011",
        "docker-group 用户与普通用户隔离验证",
        "验证 test-docker-g 与 bob/alice 之间的容器隔离",
        PRE_CROSS + "\n• test-docker-g 用户已配置",
        "1. test-docker-g 创建容器：\n"
        "   $ su - test-docker-g -c 'docker run -d --name dgrp-isolated nginx:alpine'\n\n"
        "2. bob 尝试访问该容器：\n"
        "   $ su - bob -c 'docker inspect dgrp-isolated'\n"
        "   # 预期：Error: No such container\n\n"
        "3. alice 尝试访问该容器：\n"
        "   $ su - alice -c 'docker inspect dgrp-isolated'\n"
        "   # 预期：Error: No such container\n\n"
        "4. test-docker-g 可以看到自己的容器：\n"
        "   $ su - test-docker-g -c 'docker ps -a | grep dgrp-isolated'\n"
        "   # 预期：显示 dgrp-isolated\n\n"
        "5. 清理：\n"
        "   $ su - test-docker-g -c 'docker rm -f dgrp-isolated'",
        "• bob 和 alice 无法访问 test-docker-g 的容器\n"
        "• test-docker-g 正常可见自己的容器\n"
        "• 隔离在所有用户类型间生效"
    ),
    (
        "TC-CROSS-012",
        "响应过滤前缀剥离一致性验证",
        "验证所有用户的响应过滤均正确剥离前缀",
        PRE_CROSS,
        "1. 各用户创建同名容器（不同前缀）：\n"
        "   $ su - bob -c 'docker run -d --name myapp nginx:alpine'\n"
        "   $ su - alice -c 'docker run -d --name myapp nginx:alpine'\n\n"
        "2. bob 查看容器名：\n"
        "   $ su - bob -c 'docker ps --format {{.Names}}'\n"
        "   # 预期：myapp（不含 user-1003- 前缀）\n\n"
        "3. alice 查看容器名：\n"
        "   $ su - alice -c 'docker ps --format {{.Names}}'\n"
        "   # 预期：myapp（不含 user-1004- 前缀）\n\n"
        "4. root 查看所有容器：\n"
        "   # docker ps --format {{.Names}}\n"
        "   # 预期：user-1003-myapp 和 user-1004-myapp（带前缀）\n\n"
        "5. 验证 inspect 响应也过滤前缀：\n"
        "   $ su - bob -c 'docker inspect myapp | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     print(d[\\\"Name\\\"])  # 预期: /myapp\"'\n\n"
        "6. 清理：\n"
        "   $ su - bob -c 'docker rm -f myapp'\n"
        "   $ su - alice -c 'docker rm -f myapp'",
        "• 各用户看到的容器名均已剥离前缀\n"
        "• root 看到带前缀的真实名称\n"
        "• inspect 响应中 Name 字段也正确过滤"
    ),
    (
        "TC-CROSS-013",
        "用户删除后容器孤立处理验证",
        "验证用户被删除后其容器的处理方式",
        PRE_CROSS + "\n• 创建临时用户 tempuser (uid=1099)",
        "1. 创建临时用户并运行容器：\n"
        "   # useradd -u 1099 tempuser\n"
        "   # 为 tempuser 配置代理 socket\n"
        "   $ su - tempuser -c 'docker run -d --name temp-orphan nginx:alpine'\n\n"
        "2. 删除用户：\n"
        "   # userdel tempuser\n\n"
        "3. 验证孤立容器仍存在（root 视角）：\n"
        "   # docker ps -a | grep user-1099\n"
        "   # 预期：user-1099-temp-orphan 仍在运行\n\n"
        "4. 验证 SQLite 中的所有权记录：\n"
        "   # sqlite3 /var/lib/docker-authz/ownership.db \\\n"
        "       \"SELECT * FROM containers WHERE owner_uid=1099;\"\n"
        "   # 预期：记录仍存在\n\n"
        "5. root 清理孤立容器：\n"
        "   # docker rm -f user-1099-temp-orphan\n\n"
        "6. 验证数据库清理：\n"
        "   # sqlite3 /var/lib/docker-authz/ownership.db \\\n"
        "       \"SELECT * FROM containers WHERE owner_uid=1099;\"\n"
        "   # 预期：无记录",
        "• 用户删除后容器变为孤立状态但仍运行\n"
        "• root 可以清理孤立容器\n"
        "• 数据库记录随容器删除而清理"
    ),
]  # end cross_cases

PRE_PERF = (
    "• 代理服务运行中\n"
    "• 测试环境：至少 4 核 CPU，8GB 内存\n"
    "• 已安装 time 命令和 wrk/ab 压测工具\n"
    "• 测试前清理所有残留容器"
)

perf_cases = [
    (
        "TC-PERF-001",
        "代理请求延迟基准测试",
        "测量代理对 Docker API 请求的额外延迟",
        PRE_PERF,
        "1. 测量直连 Docker daemon 延迟（基准）：\n"
        "   # time docker -H unix:///var/run/docker.sock ps\n"
        "   # 记录 real 时间（基准值）\n\n"
        "2. 测量通过代理的延迟：\n"
        "   $ su - bob -c 'time docker ps'\n"
        "   # 记录 real 时间\n\n"
        "3. 重复测试 10 次取平均值：\n"
        "   $ for i in $(seq 1 10); do\n"
        "       su - bob -c 'time docker ps' 2>&1 | grep real\n"
        "     done\n\n"
        "4. 计算平均延迟：\n"
        "   # 手动计算或使用 awk 统计\n"
        "   # 预期：代理额外延迟 < 50ms\n\n"
        "5. 记录结果到测试报告",
        "• 代理额外延迟 < 50ms（P99）\n"
        "• 10 次测试结果稳定，无异常峰值\n"
        "• 基准与代理延迟差值在可接受范围"
    ),
    (
        "TC-PERF-002",
        "并发用户请求吞吐量测试",
        "测试多用户并发操作时代理的吞吐量",
        PRE_PERF,
        "1. 准备并发测试脚本：\n"
        "   # cat > /tmp/concurrent_test.sh << 'EOF'\n"
        "   #!/bin/bash\n"
        "   USER=$1\n"
        "   for i in $(seq 1 20); do\n"
        "     su - $USER -c 'docker ps' > /dev/null 2>&1\n"
        "   done\n"
        "   EOF\n"
        "   # chmod +x /tmp/concurrent_test.sh\n\n"
        "2. 并发运行 4 个用户各 20 次请求：\n"
        "   # time (\n"
        "     /tmp/concurrent_test.sh bob &\n"
        "     /tmp/concurrent_test.sh alice &\n"
        "     /tmp/concurrent_test.sh test-docker-g &\n"
        "     /tmp/concurrent_test.sh test-sudo &\n"
        "     wait\n"
        "   )\n\n"
        "3. 记录总耗时和每秒请求数：\n"
        "   # 总请求数 = 80，计算 QPS = 80 / real_time\n"
        "   # 预期：QPS > 50\n\n"
        "4. 检查代理进程 CPU 和内存：\n"
        "   # ps aux | grep docker-authz-proxy\n"
        "   # 预期：CPU < 50%，内存 < 200MB",
        "• 80 个并发请求全部成功完成\n"
        "• QPS > 50\n"
        "• 代理进程资源占用在合理范围"
    ),
    (
        "TC-PERF-003",
        "大量容器列表响应过滤性能",
        "测试存在大量容器时响应过滤的性能",
        PRE_PERF + "\n• 预先创建 50 个容器（root 视角）",
        "1. 创建 50 个测试容器（root）：\n"
        "   # for i in $(seq 1 25); do\n"
        "     docker run -d --name user-1003-bob-perf-$i nginx:alpine\n"
        "   done\n"
        "   # for i in $(seq 1 25); do\n"
        "     docker run -d --name user-1004-alice-perf-$i nginx:alpine\n"
        "   done\n\n"
        "2. 测量 bob 列出容器的时间：\n"
        "   $ time su - bob -c 'docker ps -a'\n"
        "   # 预期：< 2 秒，显示 25 个 bob 容器\n\n"
        "3. 测量 alice 列出容器的时间：\n"
        "   $ time su - alice -c 'docker ps -a'\n"
        "   # 预期：< 2 秒，显示 25 个 alice 容器\n\n"
        "4. 验证过滤正确性：\n"
        "   $ su - bob -c 'docker ps -a | grep -c bob-perf'\n"
        "   # 预期：25\n"
        "   $ su - alice -c 'docker ps -a | grep -c alice-perf'\n"
        "   # 预期：25\n\n"
        "5. 清理：\n"
        "   # docker rm -f $(docker ps -aq --filter name=perf)",
        "• 50 容器场景下响应时间 < 2 秒\n"
        "• 过滤结果正确（各用户 25 个）\n"
        "• 无内存溢出或超时"
    ),
    (
        "TC-PERF-004",
        "配额检查性能测试",
        "测试资源配额检查对容器创建延迟的影响",
        PRE_PERF,
        "1. 测量无配额限制时容器创建时间（root）：\n"
        "   # time docker run --rm nginx:alpine echo 'test'\n"
        "   # 记录基准时间\n\n"
        "2. 测量有配额限制时容器创建时间（bob）：\n"
        "   $ time su - bob -c 'docker run --rm nginx:alpine echo test'\n"
        "   # 记录时间\n\n"
        "3. 重复 5 次取平均：\n"
        "   $ for i in $(seq 1 5); do\n"
        "     time su - bob -c 'docker run --rm nginx:alpine echo test' 2>&1 | grep real\n"
        "   done\n\n"
        "4. 计算配额检查额外开销：\n"
        "   # 预期：配额检查额外延迟 < 20ms",
        "• 配额检查额外延迟 < 20ms\n"
        "• 5 次测试结果稳定\n"
        "• 容器创建成功率 100%"
    ),
    (
        "TC-PERF-005",
        "审计日志写入性能测试",
        "测试高频操作时审计日志写入对性能的影响",
        PRE_PERF,
        "1. 执行 100 次 docker ps 操作：\n"
        "   $ time for i in $(seq 1 100); do\n"
        "     su - bob -c 'docker ps' > /dev/null\n"
        "   done\n\n"
        "2. 检查审计日志文件大小增长：\n"
        "   # ls -lh /var/log/docker-authz/user-operation/bob.log\n"
        "   # wc -l /var/log/docker-authz/user-operation/bob.log\n"
        "   # 预期：100 条新增记录\n\n"
        "3. 验证日志写入无丢失：\n"
        "   # grep -c '\"action\":\"ps\"' /var/log/docker-authz/user-operation/bob.log\n"
        "   # 预期：>= 100\n\n"
        "4. 检查磁盘 I/O：\n"
        "   # iostat -x 1 5 | grep -A2 'Device'\n"
        "   # 预期：I/O 等待 < 10%",
        "• 100 次操作全部记录到审计日志\n"
        "• 日志写入无丢失\n"
        "• 磁盘 I/O 在合理范围"
    ),
    (
        "TC-PERF-006",
        "代理重启恢复时间测试",
        "测试代理服务重启后恢复正常服务的时间",
        PRE_PERF,
        "1. 记录重启前状态：\n"
        "   # systemctl status docker-authz-proxy\n\n"
        "2. 重启代理服务并计时：\n"
        "   # time systemctl restart docker-authz-proxy\n\n"
        "3. 等待服务就绪并测试：\n"
        "   # for i in $(seq 1 10); do\n"
        "     su - bob -c 'docker ps' && echo 'Ready after ${i}s' && break\n"
        "     sleep 1\n"
        "   done\n\n"
        "4. 验证重启后功能正常：\n"
        "   $ su - bob -c 'docker run --rm nginx:alpine echo ok'\n"
        "   # 预期：ok\n\n"
        "5. 检查重启后日志：\n"
        "   # journalctl -u docker-authz-proxy --since '1 min ago' | tail -20",
        "• 代理重启时间 < 5 秒\n"
        "• 重启后立即可以处理请求\n"
        "• 功能完全恢复，无状态丢失"
    ),
    (
        "TC-PERF-007",
        "长时间运行稳定性测试",
        "测试代理在持续负载下的稳定性（内存泄漏检测）",
        PRE_PERF + "\n• 测试持续时间：30 分钟",
        "1. 记录初始内存使用：\n"
        "   # ps aux | grep docker-authz-proxy | awk '{print $6}'\n"
        "   # 记录 RSS 值（KB）\n\n"
        "2. 启动持续负载脚本（后台运行）：\n"
        "   # cat > /tmp/load_test.sh << 'EOF'\n"
        "   #!/bin/bash\n"
        "   END=$((SECONDS+1800))\n"
        "   while [ $SECONDS -lt $END ]; do\n"
        "     su - bob -c 'docker ps' > /dev/null\n"
        "     su - alice -c 'docker ps' > /dev/null\n"
        "     sleep 0.1\n"
        "   done\n"
        "   EOF\n"
        "   # bash /tmp/load_test.sh &\n\n"
        "3. 每 5 分钟记录一次内存：\n"
        "   # for i in $(seq 1 6); do\n"
        "     sleep 300\n"
        "     ps aux | grep docker-authz-proxy | awk '{print NR\": \"$6\" KB\"}'\n"
        "   done\n\n"
        "4. 测试结束后对比内存：\n"
        "   # 预期：内存增长 < 10%（无明显泄漏）",
        "• 30 分钟持续运行无崩溃\n"
        "• 内存增长 < 10%\n"
        "• 所有请求正常响应"
    ),
    (
        "TC-PERF-008",
        "大文件 docker cp 性能测试",
        "测试通过代理进行大文件传输的性能",
        PRE_PERF,
        "1. 创建测试容器和大文件：\n"
        "   $ su - bob -c 'docker run -d --name bob-cp-test nginx:alpine'\n"
        "   # dd if=/dev/urandom of=/tmp/testfile_100mb bs=1M count=100\n\n"
        "2. 测试文件复制到容器的性能：\n"
        "   $ time su - bob -c 'docker cp /tmp/testfile_100mb bob-cp-test:/tmp/'\n"
        "   # 记录时间和速度\n\n"
        "3. 测试从容器复制文件的性能：\n"
        "   $ time su - bob -c 'docker cp bob-cp-test:/tmp/testfile_100mb /tmp/out_100mb'\n"
        "   # 记录时间和速度\n\n"
        "4. 验证文件完整性：\n"
        "   # md5sum /tmp/testfile_100mb /tmp/out_100mb\n"
        "   # 预期：两个 MD5 值相同\n\n"
        "5. 清理：\n"
        "   $ su - bob -c 'docker rm -f bob-cp-test'\n"
        "   # rm /tmp/testfile_100mb /tmp/out_100mb",
        "• 100MB 文件传输速度 > 100MB/s\n"
        "• 文件 MD5 校验一致\n"
        "• 代理无超时或中断"
    ),
    (
        "TC-PERF-009",
        "策略文件热重载性能测试",
        "测试策略文件更新后代理热重载的时间",
        PRE_PERF + "\n• 代理支持 SIGHUP 热重载",
        "1. 记录当前策略生效状态：\n"
        "   $ su - bob -c 'docker exec $(docker run -d nginx:alpine) ls'\n"
        "   # 预期：被拒绝（bob 无 exec 权限）\n\n"
        "2. 修改策略文件（临时允许 bob exec）：\n"
        "   # cp /etc/docker-authz/policy.yaml /etc/docker-authz/policy.yaml.bak\n"
        "   # 编辑 policy.yaml，为 bob 添加 exec 权限\n\n"
        "3. 发送热重载信号并计时：\n"
        "   # time kill -HUP $(pgrep docker-authz-proxy)\n\n"
        "4. 验证新策略立即生效：\n"
        "   $ su - bob -c 'docker run -d --name bob-exec-test nginx:alpine'\n"
        "   $ su - bob -c 'docker exec bob-exec-test ls'\n"
        "   # 预期：成功（新策略生效）\n\n"
        "5. 恢复原策略：\n"
        "   # cp /etc/docker-authz/policy.yaml.bak /etc/docker-authz/policy.yaml\n"
        "   # kill -HUP $(pgrep docker-authz-proxy)\n\n"
        "6. 清理：\n"
        "   $ su - bob -c 'docker rm -f bob-exec-test'",
        "• 热重载时间 < 1 秒\n"
        "• 新策略立即生效\n"
        "• 热重载期间无服务中断"
    ),
    (
        "TC-PERF-010",
        "quota.yaml 配额限制压力测试",
        "测试在配额边界条件下的系统行为",
        PRE_PERF,
        "1. bob 快速创建容器至配额上限（max_containers=3）：\n"
        "   $ su - bob -c 'docker run -d --name bob-q1 nginx:alpine'\n"
        "   $ su - bob -c 'docker run -d --name bob-q2 nginx:alpine'\n"
        "   $ su - bob -c 'docker run -d --name bob-q3 nginx:alpine'\n"
        "   # 预期：3 个均成功\n\n"
        "2. 尝试创建第 4 个容器：\n"
        "   $ su - bob -c 'docker run -d --name bob-q4 nginx:alpine'\n"
        "   # 预期：被拒绝，提示超出配额\n\n"
        "3. 并发尝试创建多个容器（竞争条件测试）：\n"
        "   $ su - bob -c 'docker rm bob-q3'\n"
        "   $ su - bob -c 'docker run -d --name bob-race1 nginx:alpine' &\n"
        "   $ su - bob -c 'docker run -d --name bob-race2 nginx:alpine' &\n"
        "   $ wait\n"
        "   # 预期：只有一个成功（配额=3，已有2个）\n\n"
        "4. 验证最终容器数量：\n"
        "   $ su - bob -c 'docker ps -a | grep -c bob'\n"
        "   # 预期：<= 3\n\n"
        "5. 清理：\n"
        "   $ su - bob -c 'docker rm -f bob-q1 bob-q2 bob-race1 bob-race2 2>/dev/null'",
        "• 配额上限严格执行\n"
        "• 并发创建时配额检查无竞争条件\n"
        "• 最终容器数不超过 max_containers=3"
    ),
]  # end perf_cases

PRE_ERR = (
    "• 代理服务运行中\n"
    "• 测试用户已创建并配置\n"
    "• 测试前确认系统处于正常状态"
)

err_cases = [
    (
        "TC-ERR-001",
        "无效容器名称处理",
        "验证代理对无效容器名称的处理",
        PRE_ERR,
        "1. 尝试创建名称含特殊字符的容器：\n"
        "   $ su - bob -c 'docker run -d --name \"bob/invalid\" nginx:alpine'\n"
        "   # 预期：Docker 返回无效名称错误\n\n"
        "2. 尝试创建名称过长的容器：\n"
        "   $ su - bob -c 'docker run -d --name $(python3 -c \"print(\\\"a\\\"*300)\") nginx:alpine'\n"
        "   # 预期：错误，名称过长\n\n"
        "3. 尝试创建空名称容器：\n"
        "   $ su - bob -c 'docker run -d --name \"\" nginx:alpine'\n"
        "   # 预期：错误，名称不能为空\n\n"
        "4. 验证代理不崩溃：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回（代理仍在运行）",
        "• 无效名称被 Docker 或代理拒绝\n"
        "• 代理不崩溃，继续正常服务\n"
        "• 错误信息清晰"
    ),
    (
        "TC-ERR-002",
        "超出内存配额的容器创建",
        "验证超出内存配额时的错误处理",
        PRE_ERR + "\n• alice 内存配额：512MB",
        "1. alice 尝试创建超出内存限制的容器：\n"
        "   $ su - alice -c 'docker run -d --memory=1g --name alice-oom nginx:alpine'\n"
        "   # 预期：被拒绝，提示超出内存配额\n\n"
        "2. 验证错误信息包含配额信息：\n"
        "   # 预期错误：memory limit exceeds quota (512MB)\n\n"
        "3. alice 尝试不指定内存限制（代理注入配额）：\n"
        "   $ su - alice -c 'docker run -d --name alice-ok nginx:alpine'\n"
        "   # 预期：成功，代理注入 512MB 限制\n\n"
        "4. 验证注入的内存限制：\n"
        "   # docker inspect user-1004-alice-ok | python3 -c \"\n"
        "     import json,sys; d=json.load(sys.stdin)[0]\n"
        "     print('Memory:', d['HostConfig']['Memory'], '(期望: 536870912)')\"\n\n"
        "5. 清理：\n"
        "   $ su - alice -c 'docker rm -f alice-ok'",
        "• 超出配额的请求被拒绝\n"
        "• 错误信息包含配额限制值\n"
        "• 正常请求代理注入正确配额"
    ),
    (
        "TC-ERR-003",
        "代理 socket 权限错误处理",
        "验证无权限访问 socket 时的错误处理",
        PRE_ERR,
        "1. 创建无 socket 权限的用户：\n"
        "   # useradd -u 1099 noperm-user\n"
        "   # 不为该用户创建 socket\n\n"
        "2. 该用户尝试执行 docker 命令：\n"
        "   $ su - noperm-user -c 'docker ps'\n"
        "   # 预期：Cannot connect to the Docker daemon\n\n"
        "3. 验证错误不泄露系统信息：\n"
        "   # 预期：错误信息不包含内部路径或配置\n\n"
        "4. 验证代理正常运行（其他用户不受影响）：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回\n\n"
        "5. 清理：\n"
        "   # userdel noperm-user",
        "• 无权限用户无法连接代理\n"
        "• 错误信息不泄露内部信息\n"
        "• 其他用户不受影响"
    ),
    (
        "TC-ERR-004",
        "Docker daemon 不可用时的处理",
        "验证 Docker daemon 停止时代理的错误处理",
        PRE_ERR + "\n• 需要 root 权限停止 Docker daemon",
        "1. 停止 Docker daemon：\n"
        "   # systemctl stop docker\n\n"
        "2. bob 尝试执行 docker 命令：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：错误，无法连接到 Docker daemon\n\n"
        "3. 验证代理返回合适的错误码：\n"
        "   # 预期：HTTP 503 或类似错误\n\n"
        "4. 重启 Docker daemon：\n"
        "   # systemctl start docker\n"
        "   # sleep 3\n\n"
        "5. 验证代理自动恢复：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回",
        "• Docker daemon 不可用时返回清晰错误\n"
        "• 代理不崩溃\n"
        "• Docker daemon 恢复后代理自动恢复"
    ),
    (
        "TC-ERR-005",
        "policy.yaml 格式错误处理",
        "验证策略文件格式错误时的处理",
        PRE_ERR,
        "1. 备份策略文件：\n"
        "   # cp /etc/docker-authz/policy.yaml /etc/docker-authz/policy.yaml.bak\n\n"
        "2. 写入无效 YAML：\n"
        "   # echo 'invalid: yaml: content: [' > /etc/docker-authz/policy.yaml\n\n"
        "3. 重启代理服务：\n"
        "   # systemctl restart docker-authz-proxy\n"
        "   # 预期：启动失败，日志包含 YAML 解析错误\n\n"
        "4. 检查错误日志：\n"
        "   # journalctl -u docker-authz-proxy --since '1 min ago'\n"
        "   # 预期：YAML parse error 相关信息\n\n"
        "5. 恢复策略文件：\n"
        "   # cp /etc/docker-authz/policy.yaml.bak /etc/docker-authz/policy.yaml\n"
        "   # systemctl restart docker-authz-proxy\n"
        "   # systemctl status docker-authz-proxy\n"
        "   # 预期：active (running)",
        "• 无效策略文件导致代理启动失败\n"
        "• 错误日志包含具体解析错误\n"
        "• 恢复正确配置后代理正常启动"
    ),
    (
        "TC-ERR-006",
        "quota.yaml 格式错误处理",
        "验证配额文件格式错误时的处理",
        PRE_ERR,
        "1. 备份配额文件：\n"
        "   # cp /etc/docker-authz/quota.yaml /etc/docker-authz/quota.yaml.bak\n\n"
        "2. 写入无效配额值：\n"
        "   # cat > /etc/docker-authz/quota.yaml << 'EOF'\n"
        "   defaults:\n"
        "     cpu_cores: -1\n"
        "     mem_mb: 'invalid'\n"
        "   EOF\n\n"
        "3. 重启代理：\n"
        "   # systemctl restart docker-authz-proxy\n"
        "   # 预期：启动失败或使用默认值\n\n"
        "4. 检查日志：\n"
        "   # journalctl -u docker-authz-proxy --since '1 min ago'\n\n"
        "5. 恢复配额文件：\n"
        "   # cp /etc/docker-authz/quota.yaml.bak /etc/docker-authz/quota.yaml\n"
        "   # systemctl restart docker-authz-proxy",
        "• 无效配额值导致启动失败或使用安全默认值\n"
        "• 错误日志包含具体错误信息\n"
        "• 恢复后正常运行"
    ),
    (
        "TC-ERR-007",
        "磁盘空间不足时审计日志处理",
        "验证磁盘空间不足时审计日志的处理方式",
        PRE_ERR + "\n• 需要模拟磁盘空间不足",
        "1. 检查当前日志目录磁盘使用：\n"
        "   # df -h /var/log/docker-authz/\n\n"
        "2. 模拟磁盘空间不足（创建大文件占用空间）：\n"
        "   # dd if=/dev/zero of=/var/log/docker-authz/fill_disk bs=1M count=1000\n"
        "   # 注意：此操作可能影响系统，谨慎执行\n\n"
        "3. bob 执行 docker 操作：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：操作成功（代理不因日志失败而拒绝请求）\n\n"
        "4. 检查代理日志中的警告：\n"
        "   # journalctl -u docker-authz-proxy --since '1 min ago' | grep -i 'disk\\|log\\|write'\n\n"
        "5. 清理：\n"
        "   # rm /var/log/docker-authz/fill_disk",
        "• 磁盘空间不足时代理继续服务（降级处理）\n"
        "• 代理日志记录磁盘空间警告\n"
        "• 操作不因日志失败而被拒绝"
    ),
    (
        "TC-ERR-008",
        "恶意 JSON 请求处理",
        "验证代理对畸形 JSON 请求的处理",
        PRE_ERR,
        "1. 发送畸形 JSON 到代理 socket：\n"
        "   # echo -e 'POST /v1.41/containers/create HTTP/1.1\\r\\nHost: localhost\\r\\nContent-Type: application/json\\r\\nContent-Length: 10\\r\\n\\r\\n{invalid}' | \\\n"
        "     nc -U /var/run/docker-authz/bob.sock\n"
        "   # 预期：400 Bad Request\n\n"
        "2. 发送超大 JSON 请求：\n"
        "   # python3 -c \"print('{\\\"Image\\\": \\\"' + 'a'*100000 + '\\\"}')\" | \\\n"
        "     curl -s -X POST --unix-socket /var/run/docker-authz/bob.sock \\\n"
        "     -H 'Content-Type: application/json' \\\n"
        "     http://localhost/v1.41/containers/create -d @-\n"
        "   # 预期：400 或 413 错误\n\n"
        "3. 验证代理不崩溃：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回",
        "• 畸形 JSON 返回 400 错误\n"
        "• 超大请求被拒绝\n"
        "• 代理不崩溃，继续正常服务"
    ),
    (
        "TC-ERR-009",
        "并发请求竞争条件测试",
        "验证并发请求时代理的线程安全性",
        PRE_ERR,
        "1. 并发发送 50 个请求：\n"
        "   # for i in $(seq 1 50); do\n"
        "     su - bob -c 'docker ps' > /dev/null 2>&1 &\n"
        "   done\n"
        "   # wait\n"
        "   # echo '所有请求完成'\n\n"
        "2. 验证代理仍在运行：\n"
        "   # systemctl status docker-authz-proxy | grep Active\n"
        "   # 预期：active (running)\n\n"
        "3. 验证功能正常：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回\n\n"
        "4. 检查是否有错误日志：\n"
        "   # journalctl -u docker-authz-proxy --since '2 min ago' | grep -i error",
        "• 50 个并发请求全部处理\n"
        "• 代理无崩溃或死锁\n"
        "• 无竞争条件导致的错误"
    ),
    (
        "TC-ERR-010",
        "容器名称前缀冲突处理",
        "验证用户创建的容器名与系统前缀冲突时的处理",
        PRE_ERR,
        "1. bob 尝试创建名称包含其他用户前缀的容器：\n"
        "   $ su - bob -c 'docker run -d --name user-1004-alice-fake nginx:alpine'\n"
        "   # 预期：被拒绝或前缀被重新处理\n\n"
        "2. 验证 alice 无法看到该容器（如果创建成功）：\n"
        "   $ su - alice -c 'docker ps -a | grep alice-fake'\n"
        "   # 预期：无输出\n\n"
        "3. 验证 bob 的容器名处理：\n"
        "   # docker ps -a | grep bob\n"
        "   # 预期：容器名为 user-1003-user-1004-alice-fake 或被拒绝\n\n"
        "4. 清理：\n"
        "   # docker rm -f $(docker ps -aq --filter name=alice-fake) 2>/dev/null",
        "• 前缀冲突被正确处理\n"
        "• alice 无法看到 bob 创建的伪造容器\n"
        "• 系统行为一致且可预期"
    ),
    (
        "TC-ERR-011",
        "SO_PEERCRED 身份验证失败处理",
        "验证 SO_PEERCRED 无法获取时的处理",
        PRE_ERR + "\n• 需要特殊工具模拟 SO_PEERCRED 失败",
        "1. 使用 socat 模拟无 SO_PEERCRED 的连接：\n"
        "   # socat - UNIX-CONNECT:/var/run/docker-authz/bob.sock\n"
        "   # 发送 HTTP 请求\n"
        "   # 预期：连接被拒绝或返回 403\n\n"
        "2. 检查代理日志：\n"
        "   # journalctl -u docker-authz-proxy --since '1 min ago' | grep -i 'peercred\\|identity\\|auth'\n"
        "   # 预期：记录身份验证失败\n\n"
        "3. 验证正常用户不受影响：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回",
        "• SO_PEERCRED 失败时连接被拒绝\n"
        "• 代理日志记录身份验证失败\n"
        "• 正常用户不受影响"
    ),
    (
        "TC-ERR-012",
        "镜像不存在时的错误处理",
        "验证拉取不存在镜像时的错误处理",
        PRE_ERR,
        "1. bob 尝试运行不存在的镜像：\n"
        "   $ su - bob -c 'docker run --rm nonexistent-image-xyz:latest echo test'\n"
        "   # 预期：Unable to find image 错误\n\n"
        "2. 验证错误信息正确传递：\n"
        "   # 预期：错误信息包含镜像名称\n\n"
        "3. 验证代理不崩溃：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回\n\n"
        "4. 检查审计日志：\n"
        "   # grep 'nonexistent' /var/log/docker-authz/user-operation/bob.log\n"
        "   # 预期：记录操作，result=allow（代理允许，Docker 返回错误）",
        "• 不存在镜像返回清晰错误信息\n"
        "• 代理正确传递 Docker 错误\n"
        "• 审计日志记录操作"
    ),
    (
        "TC-ERR-013",
        "网络配额耗尽处理",
        "验证用户创建网络超出限制时的处理",
        PRE_ERR + "\n• 如果配置了网络数量限制",
        "1. bob 创建多个网络直到限制：\n"
        "   $ for i in $(seq 1 10); do\n"
        "     su - bob -c \"docker network create bob-net-$i\"\n"
        "   done\n"
        "   # 记录哪个网络创建失败\n\n"
        "2. 验证超出限制后被拒绝：\n"
        "   # 预期：超出限制后返回配额错误\n\n"
        "3. 删除部分网络后可以继续创建：\n"
        "   $ su - bob -c 'docker network rm bob-net-1'\n"
        "   $ su - bob -c 'docker network create bob-net-new'\n"
        "   # 预期：成功\n\n"
        "4. 清理：\n"
        "   $ for i in $(seq 2 10); do\n"
        "     su - bob -c \"docker network rm bob-net-$i 2>/dev/null\"\n"
        "   done\n"
        "   $ su - bob -c 'docker network rm bob-net-new 2>/dev/null'",
        "• 网络数量超出限制时被拒绝\n"
        "• 删除后可以继续创建\n"
        "• 错误信息清晰"
    ),
    (
        "TC-ERR-014",
        "代理进程异常终止恢复",
        "验证代理进程被 kill 后 systemd 自动重启",
        PRE_ERR + "\n• 代理由 systemd 管理，配置了 Restart=always",
        "1. 记录代理进程 PID：\n"
        "   # pgrep docker-authz-proxy\n\n"
        "2. 强制终止代理进程：\n"
        "   # kill -9 $(pgrep docker-authz-proxy)\n\n"
        "3. 等待 systemd 重启（通常 < 5 秒）：\n"
        "   # sleep 5\n\n"
        "4. 验证代理已重启：\n"
        "   # systemctl status docker-authz-proxy | grep Active\n"
        "   # pgrep docker-authz-proxy\n"
        "   # 预期：新的 PID，状态 active\n\n"
        "5. 验证功能恢复：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回\n\n"
        "6. 检查 systemd 重启日志：\n"
        "   # journalctl -u docker-authz-proxy --since '2 min ago' | grep -E 'start|restart|kill'",
        "• systemd 在 5 秒内自动重启代理\n"
        "• 重启后功能完全恢复\n"
        "• systemd 日志记录重启事件"
    ),
    (
        "TC-ERR-015",
        "loginUID 读取失败处理",
        "验证无法读取 /proc/loginuid 时的处理",
        PRE_ERR,
        "1. 检查正常情况下 loginUID 读取：\n"
        "   # cat /proc/self/loginuid\n"
        "   # 预期：返回当前用户的 loginUID\n\n"
        "2. 模拟 loginUID 为 4294967295（未登录）：\n"
        "   # 通过 systemd 服务或特殊进程测试\n"
        "   # 预期：代理使用 effective UID 作为备选\n\n"
        "3. 验证代理日志记录：\n"
        "   # journalctl -u docker-authz-proxy --since '1 min ago' | grep -i 'loginuid'\n\n"
        "4. 验证操作仍然被正确授权：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回",
        "• loginUID 不可用时使用备选身份验证\n"
        "• 代理日志记录异常情况\n"
        "• 操作不因此失败"
    ),
    (
        "TC-ERR-016",
        "容器创建时资源参数边界测试",
        "验证极端资源参数值的处理",
        PRE_ERR,
        "1. bob 尝试创建 CPU 为 0 的容器：\n"
        "   $ su - bob -c 'docker run -d --cpus=0 --name bob-zero-cpu nginx:alpine'\n"
        "   # 预期：被拒绝或使用最小值\n\n"
        "2. bob 尝试创建内存为 1MB 的容器：\n"
        "   $ su - bob -c 'docker run -d --memory=1m --name bob-tiny-mem nginx:alpine'\n"
        "   # 预期：被拒绝（低于最小内存要求）或成功\n\n"
        "3. bob 尝试创建超大 CPU 配额容器（超出配额）：\n"
        "   $ su - bob -c 'docker run -d --cpus=100 --name bob-huge-cpu nginx:alpine'\n"
        "   # 预期：被拒绝，超出 1 CPU 配额\n\n"
        "4. 清理：\n"
        "   $ su - bob -c 'docker rm -f bob-zero-cpu bob-tiny-mem bob-huge-cpu 2>/dev/null'",
        "• 极端资源参数被正确处理\n"
        "• 超出配额的请求被拒绝\n"
        "• 错误信息清晰"
    ),
    (
        "TC-ERR-017",
        "审计日志目录权限错误处理",
        "验证审计日志目录权限不足时的处理",
        PRE_ERR,
        "1. 修改审计日志目录权限：\n"
        "   # chmod 000 /var/log/docker-authz/user-operation/\n\n"
        "2. bob 执行 docker 操作：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：操作成功（代理降级处理，不因日志失败拒绝请求）\n\n"
        "3. 检查代理错误日志：\n"
        "   # journalctl -u docker-authz-proxy --since '1 min ago' | grep -i 'log\\|permission\\|write'\n"
        "   # 预期：记录日志写入失败警告\n\n"
        "4. 恢复权限：\n"
        "   # chmod 755 /var/log/docker-authz/user-operation/\n\n"
        "5. 验证日志恢复正常：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # grep 'ps' /var/log/docker-authz/user-operation/bob.log | tail -1",
        "• 日志目录权限不足时代理继续服务\n"
        "• 代理日志记录写入失败警告\n"
        "• 权限恢复后日志正常写入"
    ),
    (
        "TC-ERR-018",
        "网络断开时 docker pull 处理",
        "验证网络不可用时 docker pull 的错误处理",
        PRE_ERR,
        "1. 模拟网络断开（阻断 Docker Hub 访问）：\n"
        "   # iptables -A OUTPUT -d registry-1.docker.io -j DROP\n\n"
        "2. bob 尝试拉取镜像：\n"
        "   $ su - bob -c 'docker pull nginx:latest'\n"
        "   # 预期：网络超时或连接拒绝错误\n\n"
        "3. 验证错误信息正确传递：\n"
        "   # 预期：包含网络错误信息\n\n"
        "4. 恢复网络：\n"
        "   # iptables -D OUTPUT -d registry-1.docker.io -j DROP\n\n"
        "5. 验证代理不崩溃：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回",
        "• 网络不可用时返回清晰错误\n"
        "• 代理正确传递网络错误\n"
        "• 代理不崩溃"
    ),
]  # end err_cases

PRE_AUDIT = (
    "• 代理服务运行中\n"
    "• 审计日志目录：/var/log/docker-authz/\n"
    "• 用户日志：/var/log/docker-authz/user-operation/<username>.log\n"
    "• 集中日志：/var/log/docker-authz/auth.log"
)

audit_cases = [
    (
        "TC-AUDIT-001",
        "审计日志文件创建验证",
        "验证首次操作时自动创建用户审计日志文件",
        PRE_AUDIT + "\n• 删除 bob 的审计日志文件（如存在）",
        "1. 删除 bob 的审计日志（如存在）：\n"
        "   # rm -f /var/log/docker-authz/user-operation/bob.log\n\n"
        "2. bob 执行第一个 docker 命令：\n"
        "   $ su - bob -c 'docker ps'\n\n"
        "3. 验证日志文件已创建：\n"
        "   # ls -la /var/log/docker-authz/user-operation/bob.log\n"
        "   # 预期：文件存在，权限合适\n\n"
        "4. 查看日志内容：\n"
        "   # cat /var/log/docker-authz/user-operation/bob.log\n"
        "   # 预期：JSON 格式，包含 timestamp, user, action, result 字段",
        "• 首次操作后日志文件自动创建\n"
        "• 日志为 JSON 格式\n"
        "• 包含必要字段"
    ),
    (
        "TC-AUDIT-002",
        "允许操作的审计日志记录",
        "验证允许的操作被正确记录到审计日志",
        PRE_AUDIT,
        "1. bob 执行允许的操作：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   $ su - bob -c 'docker run -d --name bob-audit-test nginx:alpine'\n"
        "   $ su - bob -c 'docker stop bob-audit-test'\n\n"
        "2. 查看审计日志：\n"
        "   # tail -5 /var/log/docker-authz/user-operation/bob.log\n\n"
        "3. 验证日志字段：\n"
        "   # tail -3 /var/log/docker-authz/user-operation/bob.log | python3 -c \"\n"
        "     import json,sys\n"
        "     for line in sys.stdin:\n"
        "       d = json.loads(line.strip())\n"
        "       print('timestamp:', d.get('timestamp'))\n"
        "       print('user:', d.get('user'))\n"
        "       print('action:', d.get('action'))\n"
        "       print('result:', d.get('result'))\n"
        "       print('---')\"\n\n"
        "4. 清理：\n"
        "   $ su - bob -c 'docker rm bob-audit-test'",
        "• 每个操作均有对应日志记录\n"
        "• result=allow\n"
        "• timestamp, user, action 字段完整"
    ),
    (
        "TC-AUDIT-003",
        "拒绝操作的审计日志记录",
        "验证被拒绝的操作被正确记录到审计日志",
        PRE_AUDIT,
        "1. bob 尝试被拒绝的操作：\n"
        "   $ su - bob -c 'docker exec $(docker run -d nginx:alpine) ls' 2>&1 || true\n"
        "   $ su - bob -c 'docker build -t test . 2>&1' || true\n\n"
        "2. 查看审计日志中的拒绝记录：\n"
        "   # grep 'deny\\|result.*false' /var/log/docker-authz/user-operation/bob.log | tail -5\n\n"
        "3. 验证拒绝日志字段：\n"
        "   # grep 'deny' /var/log/docker-authz/user-operation/bob.log | tail -1 | python3 -c \"\n"
        "     import json,sys\n"
        "     d = json.loads(sys.stdin.read().strip())\n"
        "     print('result:', d.get('result'))\n"
        "     print('reason:', d.get('reason', 'N/A'))\n"
        "     print('action:', d.get('action'))\"\n\n"
        "4. 验证集中日志也有记录：\n"
        "   # grep 'bob' /var/log/docker-authz/auth.log | grep 'deny' | tail -3",
        "• 拒绝操作有对应日志记录\n"
        "• result=deny\n"
        "• 包含拒绝原因\n"
        "• 集中日志同步记录"
    ),
    (
        "TC-AUDIT-004",
        "审计日志时间戳准确性验证",
        "验证审计日志时间戳与实际操作时间一致",
        PRE_AUDIT,
        "1. 记录操作前时间：\n"
        "   # date +%s\n"
        "   # 记录 Unix 时间戳 T1\n\n"
        "2. bob 执行操作：\n"
        "   $ su - bob -c 'docker ps'\n\n"
        "3. 记录操作后时间：\n"
        "   # date +%s\n"
        "   # 记录 Unix 时间戳 T2\n\n"
        "4. 检查日志时间戳：\n"
        "   # tail -1 /var/log/docker-authz/user-operation/bob.log | python3 -c \"\n"
        "     import json,sys\n"
        "     d = json.loads(sys.stdin.read().strip())\n"
        "     print('log timestamp:', d.get('timestamp'))\"\n\n"
        "5. 验证时间戳在 T1 和 T2 之间：\n"
        "   # 手动对比时间戳\n"
        "   # 预期：误差 < 1 秒",
        "• 日志时间戳与实际操作时间误差 < 1 秒\n"
        "• 时间戳格式标准（ISO 8601 或 Unix 时间戳）"
    ),
    (
        "TC-AUDIT-005",
        "集中审计日志 auth.log 验证",
        "验证所有用户的操作都记录到集中日志",
        PRE_AUDIT,
        "1. 多个用户执行操作：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   $ su - alice -c 'docker ps'\n"
        "   $ su - test-docker-g -c 'docker ps'\n\n"
        "2. 查看集中日志：\n"
        "   # tail -10 /var/log/docker-authz/auth.log\n\n"
        "3. 验证各用户记录：\n"
        "   # grep 'bob' /var/log/docker-authz/auth.log | tail -1\n"
        "   # grep 'alice' /var/log/docker-authz/auth.log | tail -1\n"
        "   # grep 'test-docker-g' /var/log/docker-authz/auth.log | tail -1\n\n"
        "4. 验证日志格式一致性：\n"
        "   # tail -5 /var/log/docker-authz/auth.log | python3 -c \"\n"
        "     import json,sys\n"
        "     for line in sys.stdin:\n"
        "       d = json.loads(line.strip())\n"
        "       assert 'user' in d and 'action' in d and 'result' in d\n"
        "       print('OK:', d['user'], d['action'])\"",
        "• 所有用户操作均记录到 auth.log\n"
        "• 日志格式一致\n"
        "• 包含 user, action, result 字段"
    ),
    (
        "TC-AUDIT-006",
        "sudo 用户审计日志 loginUID 记录",
        "验证 sudo 用户的审计日志正确记录真实用户身份",
        PRE_AUDIT,
        "1. test-sudo 以 sudo 执行 docker 命令：\n"
        "   $ su - test-sudo -c 'sudo docker ps'\n\n"
        "2. 查看 auth.log 中的记录：\n"
        "   # grep 'test-sudo' /var/log/docker-authz/auth.log | tail -3\n\n"
        "3. 验证日志包含双重身份信息：\n"
        "   # grep 'test-sudo' /var/log/docker-authz/auth.log | tail -1 | python3 -c \"\n"
        "     import json,sys\n"
        "     d = json.loads(sys.stdin.read().strip())\n"
        "     print('effective_uid:', d.get('effective_uid'))\n"
        "     print('real_uid:', d.get('real_uid', d.get('login_uid')))\n"
        "     print('期望: effective_uid=0, real_uid=1001')\"\n\n"
        "4. 验证用户日志文件：\n"
        "   # ls /var/log/docker-authz/user-operation/test-sudo.log\n"
        "   # tail -3 /var/log/docker-authz/user-operation/test-sudo.log",
        "• auth.log 记录 effective_uid=0 和 real_uid=1001\n"
        "• test-sudo 有独立的用户日志文件\n"
        "• 真实身份可追溯"
    ),
    (
        "TC-AUDIT-007",
        "审计日志轮转验证",
        "验证审计日志轮转配置正确",
        PRE_AUDIT + "\n• 已配置 logrotate 或类似工具",
        "1. 检查 logrotate 配置：\n"
        "   # cat /etc/logrotate.d/docker-authz\n"
        "   # 预期：配置了 rotate, compress, dateext 等\n\n"
        "2. 手动触发日志轮转：\n"
        "   # logrotate -f /etc/logrotate.d/docker-authz\n\n"
        "3. 验证轮转结果：\n"
        "   # ls -la /var/log/docker-authz/user-operation/\n"
        "   # 预期：存在 bob.log.1 或 bob.log-YYYYMMDD.gz\n\n"
        "4. 验证轮转后新日志正常写入：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # ls -la /var/log/docker-authz/user-operation/bob.log\n"
        "   # 预期：新的 bob.log 文件，包含最新操作",
        "• logrotate 配置正确\n"
        "• 日志轮转成功\n"
        "• 轮转后新日志正常写入"
    ),
    (
        "TC-AUDIT-008",
        "审计日志内容完整性验证",
        "验证审计日志包含足够的信息用于安全审计",
        PRE_AUDIT,
        "1. bob 执行一系列操作：\n"
        "   $ su - bob -c 'docker run -d --name bob-audit nginx:alpine'\n"
        "   $ su - bob -c 'docker stop bob-audit'\n"
        "   $ su - bob -c 'docker rm bob-audit'\n\n"
        "2. 检查日志完整性：\n"
        "   # tail -3 /var/log/docker-authz/user-operation/bob.log | python3 -c \"\n"
        "     import json,sys\n"
        "     required = ['timestamp','user','uid','action','result']\n"
        "     for line in sys.stdin:\n"
        "       d = json.loads(line.strip())\n"
        "       missing = [f for f in required if f not in d]\n"
        "       if missing:\n"
        "         print('缺少字段:', missing)\n"
        "       else:\n"
        "         print('完整:', d['action'], d['result'])\"\n\n"
        "3. 验证容器名称记录：\n"
        "   # grep 'bob-audit' /var/log/docker-authz/user-operation/bob.log | tail -3\n"
        "   # 预期：日志包含容器名称（可能带前缀）\n\n"
        "4. 验证 IP/socket 信息记录：\n"
        "   # tail -1 /var/log/docker-authz/user-operation/bob.log | python3 -c \"\n"
        "     import json,sys; d=json.loads(sys.stdin.read())\n"
        "     print('socket/source:', d.get('source', d.get('socket', 'N/A')))\"",
        "• 日志包含所有必要字段\n"
        "• 容器名称被记录\n"
        "• 操作来源信息被记录"
    ),
    (
        "TC-AUDIT-009",
        "审计日志权限验证",
        "验证审计日志文件权限设置正确，防止普通用户篡改",
        PRE_AUDIT,
        "1. 检查日志目录权限：\n"
        "   # ls -la /var/log/docker-authz/\n"
        "   # 预期：目录属主 root，权限 750 或 755\n\n"
        "2. 检查用户日志文件权限：\n"
        "   # ls -la /var/log/docker-authz/user-operation/\n"
        "   # 预期：文件属主 root，权限 640 或 644\n\n"
        "3. bob 尝试修改自己的日志：\n"
        "   $ su - bob -c 'echo \"tampered\" >> /var/log/docker-authz/user-operation/bob.log'\n"
        "   # 预期：Permission denied\n\n"
        "4. bob 尝试删除自己的日志：\n"
        "   $ su - bob -c 'rm /var/log/docker-authz/user-operation/bob.log'\n"
        "   # 预期：Permission denied\n\n"
        "5. bob 尝试读取 alice 的日志：\n"
        "   $ su - bob -c 'cat /var/log/docker-authz/user-operation/alice.log'\n"
        "   # 预期：Permission denied",
        "• 日志目录权限正确\n"
        "• 普通用户无法修改或删除日志\n"
        "• 用户无法读取其他用户的日志"
    ),
    (
        "TC-AUDIT-010",
        "审计日志查询工具验证",
        "验证可以使用标准工具查询和分析审计日志",
        PRE_AUDIT,
        "1. 统计 bob 的操作次数：\n"
        "   # wc -l /var/log/docker-authz/user-operation/bob.log\n\n"
        "2. 查询特定时间段的操作：\n"
        "   # python3 -c \"\n"
        "     import json\n"
        "     from datetime import datetime\n"
        "     with open('/var/log/docker-authz/user-operation/bob.log') as f:\n"
        "       for line in f:\n"
        "         d = json.loads(line)\n"
        "         print(d.get('timestamp'), d.get('action'), d.get('result'))\"\n\n"
        "3. 统计各操作类型：\n"
        "   # python3 -c \"\n"
        "     import json\n"
        "     from collections import Counter\n"
        "     actions = Counter()\n"
        "     with open('/var/log/docker-authz/user-operation/bob.log') as f:\n"
        "       for line in f:\n"
        "         d = json.loads(line)\n"
        "         actions[d.get('action')] += 1\n"
        "     for a,c in actions.most_common(): print(f'{a}: {c}')\"\n\n"
        "4. 查询被拒绝的操作：\n"
        "   # grep -c 'deny' /var/log/docker-authz/user-operation/bob.log",
        "• 日志可以用标准工具查询\n"
        "• JSON 格式便于程序化分析\n"
        "• 统计功能正常"
    ),
    (
        "TC-AUDIT-011",
        "跨用户操作审计追踪",
        "验证跨用户越权尝试被完整记录",
        PRE_AUDIT,
        "1. bob 尝试访问 alice 的容器：\n"
        "   $ su - alice -c 'docker run -d --name alice-audit-target nginx:alpine'\n"
        "   $ su - bob -c 'docker stop alice-audit-target'\n"
        "   $ su - bob -c 'docker rm alice-audit-target'\n\n"
        "2. 检查 bob 的审计日志：\n"
        "   # grep 'alice-audit-target' /var/log/docker-authz/user-operation/bob.log\n"
        "   # 预期：记录尝试操作，result=deny 或无记录（容器不存在）\n\n"
        "3. 检查集中日志：\n"
        "   # grep 'alice-audit-target' /var/log/docker-authz/auth.log\n\n"
        "4. 验证 alice 的容器未受影响：\n"
        "   $ su - alice -c 'docker ps | grep alice-audit-target'\n"
        "   # 预期：仍在运行\n\n"
        "5. 清理：\n"
        "   $ su - alice -c 'docker rm -f alice-audit-target'",
        "• 越权尝试被记录到审计日志\n"
        "• 集中日志同步记录\n"
        "• alice 容器未受影响"
    ),
    (
        "TC-AUDIT-012",
        "root 操作审计日志验证",
        "验证 root 用户的操作也被审计记录",
        PRE_AUDIT,
        "1. root 执行 docker 操作：\n"
        "   # docker ps\n"
        "   # docker run --rm nginx:alpine echo 'root-audit-test'\n\n"
        "2. 检查 root 审计日志：\n"
        "   # ls /var/log/docker-authz/user-operation/root.log\n"
        "   # tail -3 /var/log/docker-authz/user-operation/root.log\n\n"
        "3. 验证 root 操作记录：\n"
        "   # tail -1 /var/log/docker-authz/user-operation/root.log | python3 -c \"\n"
        "     import json,sys; d=json.loads(sys.stdin.read())\n"
        "     print('user:', d.get('user'))\n"
        "     print('uid:', d.get('uid'))\n"
        "     print('result:', d.get('result'))\"\n"
        "   # 预期：user=root, uid=0, result=allow\n\n"
        "4. 验证集中日志：\n"
        "   # grep 'root' /var/log/docker-authz/auth.log | tail -3",
        "• root 操作被审计记录\n"
        "• uid=0 正确记录\n"
        "• result=allow（root 无限制）"
    ),
    (
        "TC-AUDIT-013",
        "审计日志 JSON 格式验证",
        "验证所有审计日志条目均为有效 JSON 格式",
        PRE_AUDIT,
        "1. 执行多种操作生成日志：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   $ su - bob -c 'docker images'\n"
        "   $ su - alice -c 'docker ps'\n\n"
        "2. 验证 bob 日志 JSON 格式：\n"
        "   # python3 -c \"\n"
        "     import json\n"
        "     errors = 0\n"
        "     with open('/var/log/docker-authz/user-operation/bob.log') as f:\n"
        "       for i, line in enumerate(f, 1):\n"
        "         try:\n"
        "           json.loads(line.strip())\n"
        "         except json.JSONDecodeError as e:\n"
        "           print(f'Line {i}: {e}')\n"
        "           errors += 1\n"
        "     print(f'Total errors: {errors}')\"\n\n"
        "3. 验证 auth.log JSON 格式：\n"
        "   # python3 -c \"\n"
        "     import json\n"
        "     errors = 0\n"
        "     with open('/var/log/docker-authz/auth.log') as f:\n"
        "       for i, line in enumerate(f, 1):\n"
        "         try:\n"
        "           json.loads(line.strip())\n"
        "         except: errors += 1\n"
        "     print(f'auth.log errors: {errors}')\"",
        "• 所有日志条目均为有效 JSON\n"
        "• JSON 解析无错误\n"
        "• 格式一致"
    ),
    (
        "TC-AUDIT-014",
        "审计日志与实际操作一致性验证",
        "验证审计日志记录与实际执行的操作完全一致",
        PRE_AUDIT,
        "1. 清空 bob 日志（备份后）：\n"
        "   # cp /var/log/docker-authz/user-operation/bob.log /tmp/bob.log.bak\n"
        "   # truncate -s 0 /var/log/docker-authz/user-operation/bob.log\n\n"
        "2. 执行精确的操作序列：\n"
        "   $ su - bob -c 'docker ps'                          # 操作1: ps\n"
        "   $ su - bob -c 'docker images'                      # 操作2: images\n"
        "   $ su - bob -c 'docker run -d --name bob-cons nginx:alpine'  # 操作3: create\n"
        "   $ su - bob -c 'docker stop bob-cons'               # 操作4: stop\n"
        "   $ su - bob -c 'docker rm bob-cons'                 # 操作5: rm\n\n"
        "3. 验证日志条数：\n"
        "   # wc -l /var/log/docker-authz/user-operation/bob.log\n"
        "   # 预期：>= 5 条记录\n\n"
        "4. 验证操作顺序：\n"
        "   # python3 -c \"\n"
        "     import json\n"
        "     with open('/var/log/docker-authz/user-operation/bob.log') as f:\n"
        "       for line in f:\n"
        "         d = json.loads(line)\n"
        "         print(d.get('timestamp',''), d.get('action',''), d.get('result',''))\"\n\n"
        "5. 恢复日志：\n"
        "   # cat /tmp/bob.log.bak >> /var/log/docker-authz/user-operation/bob.log",
        "• 日志条数与操作次数一致\n"
        "• 操作顺序与时间戳一致\n"
        "• 每个操作均有对应记录"
    ),
]  # end audit_cases

PRE_DEPLOY = (
    "• 测试环境：Ubuntu 22.04 LTS\n"
    "• Docker Engine 已安装\n"
    "• 具有 root 权限\n"
    "• 代理二进制文件已编译"
)

deploy_cases = [
    (
        "TC-DEPLOY-001",
        "代理服务安装验证",
        "验证代理服务可以正确安装和启动",
        PRE_DEPLOY,
        "1. 安装代理二进制文件：\n"
        "   # cp docker-authz-proxy /usr/local/bin/\n"
        "   # chmod +x /usr/local/bin/docker-authz-proxy\n\n"
        "2. 创建配置目录：\n"
        "   # mkdir -p /etc/docker-authz\n"
        "   # mkdir -p /var/log/docker-authz/user-operation\n"
        "   # mkdir -p /var/lib/docker-authz\n\n"
        "3. 部署配置文件：\n"
        "   # cp policy.yaml /etc/docker-authz/\n"
        "   # cp quota.yaml /etc/docker-authz/\n\n"
        "4. 创建 systemd 服务文件：\n"
        "   # cp docker-authz-proxy.service /etc/systemd/system/\n"
        "   # systemctl daemon-reload\n\n"
        "5. 启动服务：\n"
        "   # systemctl enable docker-authz-proxy\n"
        "   # systemctl start docker-authz-proxy\n"
        "   # systemctl status docker-authz-proxy\n"
        "   # 预期：active (running)",
        "• 二进制文件安装成功\n"
        "• 配置文件部署成功\n"
        "• systemd 服务启动成功，状态 active"
    ),
    (
        "TC-DEPLOY-002",
        "用户 socket 自动创建验证",
        "验证代理启动后自动为已配置用户创建 socket",
        PRE_DEPLOY + "\n• 代理服务已启动",
        "1. 检查 socket 目录：\n"
        "   # ls -la /var/run/docker-authz/\n"
        "   # 预期：存在各用户的 .sock 文件\n\n"
        "2. 验证各用户 socket 权限：\n"
        "   # ls -la /var/run/docker-authz/*.sock\n"
        "   # 预期：各 socket 属主为对应用户，权限 600 或 660\n\n"
        "3. 验证 socket 可连接：\n"
        "   # for user in bob alice test-docker-g test-sudo; do\n"
        "     echo -n \"$user socket: \"\n"
        "     su - $user -c 'docker ps > /dev/null 2>&1 && echo OK || echo FAIL'\n"
        "   done\n\n"
        "4. 验证 socket 文件类型：\n"
        "   # file /var/run/docker-authz/bob.sock\n"
        "   # 预期：socket",
        "• 所有用户 socket 自动创建\n"
        "• socket 权限正确\n"
        "• 所有用户可以通过 socket 连接"
    ),
    (
        "TC-DEPLOY-003",
        "代理服务升级验证",
        "验证代理服务可以无缝升级",
        PRE_DEPLOY + "\n• 代理服务已运行",
        "1. 记录当前版本：\n"
        "   # docker-authz-proxy --version\n"
        "   # 记录版本号\n\n"
        "2. 停止服务：\n"
        "   # systemctl stop docker-authz-proxy\n\n"
        "3. 替换二进制文件：\n"
        "   # cp docker-authz-proxy-new /usr/local/bin/docker-authz-proxy\n\n"
        "4. 重启服务：\n"
        "   # systemctl start docker-authz-proxy\n"
        "   # systemctl status docker-authz-proxy\n\n"
        "5. 验证新版本：\n"
        "   # docker-authz-proxy --version\n"
        "   # 预期：新版本号\n\n"
        "6. 验证功能正常：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回",
        "• 升级过程无数据丢失\n"
        "• 新版本正常启动\n"
        "• 升级后功能完全正常"
    ),
    (
        "TC-DEPLOY-004",
        "代理服务卸载验证",
        "验证代理服务可以完整卸载",
        PRE_DEPLOY + "\n• 代理服务已运行",
        "1. 停止并禁用服务：\n"
        "   # systemctl stop docker-authz-proxy\n"
        "   # systemctl disable docker-authz-proxy\n\n"
        "2. 删除服务文件：\n"
        "   # rm /etc/systemd/system/docker-authz-proxy.service\n"
        "   # systemctl daemon-reload\n\n"
        "3. 删除二进制文件：\n"
        "   # rm /usr/local/bin/docker-authz-proxy\n\n"
        "4. 验证 Docker 直连恢复：\n"
        "   # docker ps\n"
        "   # 预期：直连 Docker daemon 成功\n\n"
        "5. 验证 socket 已清理：\n"
        "   # ls /var/run/docker-authz/ 2>/dev/null || echo '目录已清理'",
        "• 服务完整停止和禁用\n"
        "• 文件清理完整\n"
        "• Docker 直连恢复正常"
    ),
    (
        "TC-DEPLOY-005",
        "配置文件热重载验证",
        "验证发送 SIGHUP 信号后配置文件被重新加载",
        PRE_DEPLOY + "\n• 代理服务已运行",
        "1. 记录当前策略：\n"
        "   # cat /etc/docker-authz/policy.yaml\n\n"
        "2. 修改策略文件（添加注释）：\n"
        "   # echo '# Updated at: '$(date) >> /etc/docker-authz/policy.yaml\n\n"
        "3. 发送 SIGHUP 信号：\n"
        "   # kill -HUP $(pgrep docker-authz-proxy)\n\n"
        "4. 检查日志确认重载：\n"
        "   # journalctl -u docker-authz-proxy --since '30 sec ago' | grep -i 'reload\\|config\\|SIGHUP'\n"
        "   # 预期：记录配置重载事件\n\n"
        "5. 验证服务仍在运行：\n"
        "   # systemctl status docker-authz-proxy | grep Active\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回",
        "• SIGHUP 触发配置重载\n"
        "• 日志记录重载事件\n"
        "• 重载后服务正常运行"
    ),
    (
        "TC-DEPLOY-006",
        "多实例部署冲突检测",
        "验证同一主机上不能运行多个代理实例",
        PRE_DEPLOY + "\n• 代理服务已运行",
        "1. 尝试启动第二个代理实例：\n"
        "   # docker-authz-proxy &\n"
        "   # 预期：启动失败，端口或 socket 已被占用\n\n"
        "2. 检查错误信息：\n"
        "   # 预期：address already in use 或类似错误\n\n"
        "3. 验证原实例仍在运行：\n"
        "   # systemctl status docker-authz-proxy | grep Active\n"
        "   # pgrep -c docker-authz-proxy\n"
        "   # 预期：只有 1 个进程\n\n"
        "4. 验证功能正常：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回",
        "• 第二个实例启动失败\n"
        "• 原实例不受影响\n"
        "• 只有一个代理进程运行"
    ),
    (
        "TC-DEPLOY-007",
        "系统启动时代理自动启动验证",
        "验证系统重启后代理服务自动启动",
        PRE_DEPLOY + "\n• 代理已配置为 systemd 服务，enable 状态",
        "1. 确认服务已启用：\n"
        "   # systemctl is-enabled docker-authz-proxy\n"
        "   # 预期：enabled\n\n"
        "2. 重启系统（或模拟）：\n"
        "   # systemctl reboot\n"
        "   # 等待系统重启完成\n\n"
        "3. 重启后验证服务状态：\n"
        "   # systemctl status docker-authz-proxy\n"
        "   # 预期：active (running)\n\n"
        "4. 验证功能正常：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # 预期：正常返回\n\n"
        "5. 检查启动日志：\n"
        "   # journalctl -u docker-authz-proxy -b | head -20\n"
        "   # 预期：启动成功日志",
        "• 系统重启后代理自动启动\n"
        "• 启动后功能完全正常\n"
        "• 启动日志无错误"
    ),
    (
        "TC-DEPLOY-008",
        "代理日志级别配置验证",
        "验证代理日志级别可以通过配置调整",
        PRE_DEPLOY + "\n• 代理支持日志级别配置（debug/info/warn/error）",
        "1. 查看当前日志级别：\n"
        "   # grep -i 'log.level\\|loglevel' /etc/docker-authz/config.yaml 2>/dev/null\n"
        "   # 或查看启动参数\n\n"
        "2. 设置为 debug 级别：\n"
        "   # 修改配置文件或启动参数添加 --log-level=debug\n"
        "   # systemctl restart docker-authz-proxy\n\n"
        "3. 执行操作并检查 debug 日志：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # journalctl -u docker-authz-proxy --since '30 sec ago'\n"
        "   # 预期：包含详细的 debug 信息\n\n"
        "4. 恢复 info 级别：\n"
        "   # 修改配置为 --log-level=info\n"
        "   # systemctl restart docker-authz-proxy",
        "• debug 级别输出详细日志\n"
        "• info 级别输出正常日志\n"
        "• 日志级别切换生效"
    ),
    (
        "TC-DEPLOY-009",
        "代理健康检查接口验证",
        "验证代理提供健康检查接口",
        PRE_DEPLOY + "\n• 代理提供 /health 或类似接口",
        "1. 检查健康检查接口：\n"
        "   # curl -s http://localhost:8080/health\n"
        "   # 或通过 socket：\n"
        "   # curl -s --unix-socket /var/run/docker-authz/admin.sock http://localhost/health\n"
        "   # 预期：{\"status\": \"ok\"} 或类似响应\n\n"
        "2. 停止 Docker daemon 后检查健康状态：\n"
        "   # systemctl stop docker\n"
        "   # curl -s http://localhost:8080/health\n"
        "   # 预期：{\"status\": \"degraded\"} 或类似\n\n"
        "3. 恢复 Docker daemon：\n"
        "   # systemctl start docker\n"
        "   # curl -s http://localhost:8080/health\n"
        "   # 预期：{\"status\": \"ok\"}",
        "• 健康检查接口正常响应\n"
        "• Docker daemon 不可用时状态降级\n"
        "• 恢复后状态正常"
    ),
    (
        "TC-DEPLOY-010",
        "代理监控指标接口验证",
        "验证代理提供 Prometheus 或类似监控指标",
        PRE_DEPLOY + "\n• 代理提供 /metrics 接口",
        "1. 检查监控指标接口：\n"
        "   # curl -s http://localhost:8080/metrics | head -30\n"
        "   # 预期：Prometheus 格式的指标数据\n\n"
        "2. 验证关键指标存在：\n"
        "   # curl -s http://localhost:8080/metrics | grep -E \\\n"
        "     'authz_requests_total|authz_denied_total|authz_latency'\n"
        "   # 预期：找到相关指标\n\n"
        "3. 执行操作后验证指标更新：\n"
        "   $ su - bob -c 'docker ps'\n"
        "   # curl -s http://localhost:8080/metrics | grep 'authz_requests_total'\n"
        "   # 预期：计数增加\n\n"
        "4. 验证拒绝计数：\n"
        "   $ su - bob -c 'docker exec test ls 2>/dev/null || true'\n"
        "   # curl -s http://localhost:8080/metrics | grep 'authz_denied_total'\n"
        "   # 预期：拒绝计数增加",
        "• 监控指标接口正常响应\n"
        "• 关键指标存在\n"
        "• 操作后指标正确更新"
    ),
]  # end deploy_cases

# ─────────────────────────────────────────────
# Workbook generation
# ─────────────────────────────────────────────

from openpyxl import Workbook

ROW_HEIGHT = 130

wb = Workbook()

sheets = [
    ("环境搭建与策略配置", setup_cases),
    ("root用户测试",       root_cases),
    ("test-sudo用户测试",  sudo_cases),
    ("test-docker-g用户测试", dgrp_cases),
    ("bob用户测试",        bob_cases),
    ("alice用户测试",      alice_cases),
    ("跨用户隔离验证",     cross_cases),
    ("性能测试",           perf_cases),
    ("异常与边界测试",     err_cases),
    ("审计日志测试",       audit_cases),
    ("部署与运维测试",     deploy_cases),
]

for i, (title, cases) in enumerate(sheets):
    add_sheet(wb, title, cases, first=(i == 0))

# Summary sheet
ws_sum = wb.create_sheet(title="测试用例汇总", index=0)
sum_headers = ["模块", "用例数量", "Sheet名称"]
for col, h in enumerate(sum_headers, 1):
    cell = ws_sum.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="2F5496")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
ws_sum.column_dimensions["A"].width = 28
ws_sum.column_dimensions["B"].width = 12
ws_sum.column_dimensions["C"].width = 28
ws_sum.freeze_panes = "A2"

total = 0
for row_idx, (title, cases) in enumerate(sheets, 2):
    fill = PatternFill("solid", fgColor="DCE6F1") if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for col, val in enumerate([title, len(cases), title], 1):
        cell = ws_sum.cell(row=row_idx, column=col, value=val)
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center" if col == 2 else "left", vertical="center")
    total += len(cases)

total_row = len(sheets) + 2
ws_sum.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
ws_sum.cell(row=total_row, column=2, value=total).font = Font(bold=True)
for col in range(1, 4):
    ws_sum.cell(row=total_row, column=col).border = BORDER
    ws_sum.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="FFF2CC")

out_path = "d:/code/docker-authz-proxy/docker-authz-proxy-testcases-v3.xlsx"
wb.save(out_path)
print(f"Generated: {out_path}")
print(f"Total test cases: {total}")
for title, cases in sheets:
    print(f"  {title}: {len(cases)}")
