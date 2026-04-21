# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "d:/code/docker-authz-proxy/docker-authz-proxy-testcases-v3.xlsx"
DST = "d:/code/docker-authz-proxy/docker-authz-proxy-testcases-v5.xlsx"
NEW_SHEET = "网络互通测试"
SUMMARY_SHEET = "测试用例汇总"

HEADERS = ["测试用例编号", "测试用例名称", "测试目的", "前提条件", "测试步骤", "预期结果"]

PRECONDITION = (
    "Linux 系统，docker-authz-proxy 服务运行中\n"
    "bob (uid=1003) 和 alice (uid=1004) 已创建，均为普通用户\n"
    "两用户均已配置 DOCKER_HOST 指向各自 socket\n"
    "已安装 docker-authz-proxy-ctl"
)

TESTCASES = [
    {
        "id": "TC-PEER-001",
        "name": "用户级互通 - 通过 uid 建立互通",
        "purpose": "验证管理员可通过 --uid-a/--uid-b 为 bob 和 alice 建立用户级网络互通",
        "steps": (
            "1. 确认当前无互通记录\n"
            "   # docker-authz-proxy-ctl peer list\n"
            "2. 建立用户级互通\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004\n"
            "3. 查看互通记录\n"
            "   # docker-authz-proxy-ctl peer list"
        ),
        "expected": (
            "• peer list 初始输出 \"no network peers configured\"\n"
            "• allow 命令输出 \"user-level peer allowed: uid=1003 <-> uid=1004 (all containers)\"\n"
            "• peer list 显示一条 uid_a=1003 uid_b=1004 type=user-level 的记录"
        ),
    },
    {
        "id": "TC-PEER-002",
        "name": "用户级互通 - 通过 username 建立互通",
        "purpose": "验证管理员可通过 --user-a/--user-b 为 bob 和 alice 建立用户级网络互通",
        "steps": (
            "1. 确认当前无互通记录（或先执行 deny 清除）\n"
            "   # docker-authz-proxy-ctl peer deny --uid-a 1003 --uid-b 1004\n"
            "2. 通过用户名建立互通\n"
            "   # docker-authz-proxy-ctl peer allow --user-a bob --user-b alice\n"
            "3. 查看互通记录\n"
            "   # docker-authz-proxy-ctl peer list"
        ),
        "expected": (
            "• allow 命令输出 \"user-level peer allowed: uid=1003 <-> uid=1004 (all containers)\"\n"
            "• peer list 显示 uid_a=1003 uid_b=1004 type=user-level 的记录"
        ),
    },
    {
        "id": "TC-PEER-003",
        "name": "用户级互通 - 撤销互通（uid）",
        "purpose": "验证 deny 命令与 allow 参数对称，可精确撤销用户级互通",
        "steps": (
            "1. 先建立互通\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004\n"
            "2. 确认互通存在\n"
            "   # docker-authz-proxy-ctl peer list\n"
            "3. 撤销互通（参数与 allow 完全一致）\n"
            "   # docker-authz-proxy-ctl peer deny --uid-a 1003 --uid-b 1004\n"
            "4. 再次查看互通记录\n"
            "   # docker-authz-proxy-ctl peer list"
        ),
        "expected": (
            "• 步骤2 显示 user-level 互通记录\n"
            "• deny 命令输出 \"user-level peer denied: uid=1003 <-> uid=1004\"\n"
            "• 步骤4 输出 \"no network peers configured\""
        ),
    },
    {
        "id": "TC-PEER-004",
        "name": "用户级互通 - 撤销互通（username）",
        "purpose": "验证通过用户名撤销与通过用户名建立的互通对称",
        "steps": (
            "1. 建立互通\n"
            "   # docker-authz-proxy-ctl peer allow --user-a bob --user-b alice\n"
            "2. 撤销互通\n"
            "   # docker-authz-proxy-ctl peer deny --user-a bob --user-b alice\n"
            "3. 查看互通记录\n"
            "   # docker-authz-proxy-ctl peer list"
        ),
        "expected": (
            "• deny 命令成功执行，输出 \"user-level peer denied: uid=1003 <-> uid=1004\"\n"
            "• peer list 输出 \"no network peers configured\""
        ),
    },
    {
        "id": "TC-PEER-005",
        "name": "容器级互通 - 建立指定容器互通",
        "purpose": "验证管理员可为 bob 和 alice 的指定容器建立容器级网络互通",
        "steps": (
            "1. bob 启动容器\n"
            "   $ su - bob -c \"docker run -d --name bob-web nginx:alpine\"\n"
            "2. alice 启动容器\n"
            "   $ su - alice -c \"docker run -d --name alice-app nginx:alpine\"\n"
            "3. 建立容器级互通\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004 \\\n"
            "       --container-a bob-web --container-b alice-app\n"
            "4. 查看互通记录\n"
            "   # docker-authz-proxy-ctl peer list"
        ),
        "expected": (
            "• allow 命令输出 \"container-level peer allowed: uid=1003 container=bob-web <-> uid=1004 container=alice-app\"\n"
            "• peer list 显示 type=container-level，container_a=bob-web，container_b=alice-app 的记录"
        ),
    },
    {
        "id": "TC-PEER-006",
        "name": "容器级互通 - 撤销指定容器互通",
        "purpose": "验证 deny 命令可精确撤销容器级互通，参数与 allow 完全对称",
        "steps": (
            "1. 建立容器级互通（容器已存在）\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004 \\\n"
            "       --container-a bob-web --container-b alice-app\n"
            "2. 确认互通存在\n"
            "   # docker-authz-proxy-ctl peer list\n"
            "3. 撤销容器级互通（参数与 allow 完全一致）\n"
            "   # docker-authz-proxy-ctl peer deny --uid-a 1003 --uid-b 1004 \\\n"
            "       --container-a bob-web --container-b alice-app\n"
            "4. 查看互通记录\n"
            "   # docker-authz-proxy-ctl peer list"
        ),
        "expected": (
            "• deny 命令输出 \"container-level peer denied: uid=1003 container=bob-web <-> uid=1004 container=alice-app\"\n"
            "• 步骤4 输出 \"no network peers configured\""
        ),
    },
    {
        "id": "TC-PEER-007",
        "name": "用户级互通 - bob 容器 ping alice 容器",
        "purpose": "验证用户级互通建立后，bob 的容器可以 ping 通 alice 的容器",
        "steps": (
            "1. bob 启动容器\n"
            "   $ su - bob -c \"docker run -d --name bob-net-test alpine sleep 3600\"\n"
            "2. alice 启动容器\n"
            "   $ su - alice -c \"docker run -d --name alice-net-test alpine sleep 3600\"\n"
            "3. 建立用户级互通\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004\n"
            "4. 获取 alice 容器 IP\n"
            "   # docker inspect alice-net-test --format '{{.NetworkSettings.Networks}}'\n"
            "5. 从 bob 容器 ping alice 容器 IP\n"
            "   # docker exec bob-net-test ping -c 3 <alice-net-test-ip>"
        ),
        "expected": (
            "• ping 命令返回 0% packet loss\n"
            "• 输出类似 \"3 packets transmitted, 3 received, 0% packet loss\""
        ),
    },
    {
        "id": "TC-PEER-008",
        "name": "用户级互通 - alice 容器 ping bob 容器",
        "purpose": "验证用户级互通是双向的，alice 的容器也可以 ping 通 bob 的容器",
        "steps": (
            "1. 确保 TC-PEER-007 中的容器和互通已建立\n"
            "2. 获取 bob 容器 IP\n"
            "   # docker inspect bob-net-test --format '{{.NetworkSettings.Networks}}'\n"
            "3. 从 alice 容器 ping bob 容器 IP\n"
            "   # docker exec alice-net-test ping -c 3 <bob-net-test-ip>"
        ),
        "expected": (
            "• ping 命令返回 0% packet loss\n"
            "• 互通为双向可达"
        ),
    },
    {
        "id": "TC-PEER-009",
        "name": "用户级互通 - 撤销后网络隔离恢复",
        "purpose": "验证撤销用户级互通后，bob 和 alice 的容器恢复网络隔离",
        "steps": (
            "1. 确保 TC-PEER-007 中的容器和互通已建立，且 ping 可通\n"
            "2. 撤销互通\n"
            "   # docker-authz-proxy-ctl peer deny --uid-a 1003 --uid-b 1004\n"
            "3. 再次从 bob 容器 ping alice 容器 IP\n"
            "   # docker exec bob-net-test ping -c 3 <alice-net-test-ip> || echo \"UNREACHABLE\"\n"
            "4. 再次从 alice 容器 ping bob 容器 IP\n"
            "   # docker exec alice-net-test ping -c 3 <bob-net-test-ip> || echo \"UNREACHABLE\""
        ),
        "expected": (
            "• 步骤3 ping 失败，输出 \"100% packet loss\" 或 \"UNREACHABLE\"\n"
            "• 步骤4 ping 失败，网络隔离已恢复"
        ),
    },
    {
        "id": "TC-PEER-010",
        "name": "容器级互通 - 指定容器间网络连通",
        "purpose": "验证容器级互通建立后，指定容器间可以网络通信",
        "steps": (
            "1. bob 启动容器 bob-web\n"
            "   $ su - bob -c \"docker run -d --name bob-web nginx:alpine\"\n"
            "2. alice 启动容器 alice-app\n"
            "   $ su - alice -c \"docker run -d --name alice-app alpine sleep 3600\"\n"
            "3. 建立容器级互通\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004 \\\n"
            "       --container-a bob-web --container-b alice-app\n"
            "4. 获取 bob-web 容器 IP\n"
            "   # docker inspect bob-web --format '{{range .NetworkSettings.Networks}}{{.IPAddress}}{{end}}'\n"
            "5. 从 alice-app 访问 bob-web\n"
            "   # docker exec alice-app wget -qO- http://<bob-web-ip> | head -5"
        ),
        "expected": (
            "• wget 成功返回 nginx 欢迎页面 HTML 内容\n"
            "• HTTP 状态码 200"
        ),
    },
    {
        "id": "TC-PEER-011",
        "name": "容器级互通 - 非互通容器不受影响",
        "purpose": "验证容器级互通只影响指定容器，同用户的其他容器仍然隔离",
        "steps": (
            "1. bob 启动两个容器：bob-web（互通）和 bob-other（非互通）\n"
            "   $ su - bob -c \"docker run -d --name bob-web nginx:alpine\"\n"
            "   $ su - bob -c \"docker run -d --name bob-other alpine sleep 3600\"\n"
            "2. alice 启动容器 alice-app\n"
            "   $ su - alice -c \"docker run -d --name alice-app alpine sleep 3600\"\n"
            "3. 只为 bob-web 和 alice-app 建立容器级互通\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004 \\\n"
            "       --container-a bob-web --container-b alice-app\n"
            "4. 获取 alice-app IP\n"
            "   # docker inspect alice-app --format '{{range .NetworkSettings.Networks}}{{.IPAddress}}{{end}}'\n"
            "5. 从 bob-other 尝试 ping alice-app\n"
            "   # docker exec bob-other ping -c 3 <alice-app-ip> || echo \"UNREACHABLE\""
        ),
        "expected": (
            "• bob-web 可以访问 alice-app（互通容器正常）\n"
            "• bob-other ping alice-app 失败，输出 \"UNREACHABLE\"（非互通容器仍隔离）"
        ),
    },
    {
        "id": "TC-PEER-012",
        "name": "列表过滤 - 按 uid 过滤",
        "purpose": "验证 peer list --uid 可正确过滤显示指定用户参与的互通记录",
        "steps": (
            "1. 建立多条互通记录（用户级 + 容器级）\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004 \\\n"
            "       --container-a bob-web --container-b alice-app\n"
            "2. 按 bob 的 uid 过滤\n"
            "   # docker-authz-proxy-ctl peer list --uid 1003\n"
            "3. 按 alice 的 uid 过滤\n"
            "   # docker-authz-proxy-ctl peer list --uid 1004"
        ),
        "expected": (
            "• 步骤2 显示所有包含 uid=1003 的记录（2条）\n"
            "• 步骤3 显示所有包含 uid=1004 的记录（2条）\n"
            "• 不显示不包含该 uid 的记录"
        ),
    },
    {
        "id": "TC-PEER-013",
        "name": "列表过滤 - 按 username 过滤",
        "purpose": "验证 peer list --user 可通过用户名过滤互通记录",
        "steps": (
            "1. 确保存在互通记录\n"
            "2. 按用户名 bob 过滤\n"
            "   # docker-authz-proxy-ctl peer list --user bob\n"
            "3. 按用户名 alice 过滤\n"
            "   # docker-authz-proxy-ctl peer list --user alice"
        ),
        "expected": (
            "• 步骤2 显示 bob 参与的所有互通记录\n"
            "• 步骤3 显示 alice 参与的所有互通记录\n"
            "• 结果与 --uid 1003 / --uid 1004 过滤结果一致"
        ),
    },
    {
        "id": "TC-PEER-014",
        "name": "列表过滤 - 按容器名过滤",
        "purpose": "验证 peer list --container 可按容器名过滤互通记录",
        "steps": (
            "1. 确保存在容器级互通记录（bob-web <-> alice-app）\n"
            "2. 按容器名过滤\n"
            "   # docker-authz-proxy-ctl peer list --container bob-web\n"
            "   # docker-authz-proxy-ctl peer list --container alice-app"
        ),
        "expected": (
            "• 两条命令均显示 bob-web <-> alice-app 的容器级互通记录\n"
            "• 不显示用户级互通记录（用户级记录 container_a/b 为空）"
        ),
    },
    {
        "id": "TC-PEER-015",
        "name": "列表过滤 - uid + container 双重过滤",
        "purpose": "验证同时指定 --uid 和 --container 时取交集过滤",
        "steps": (
            "1. 建立多条记录：用户级互通 + bob-web/alice-app 容器级互通\n"
            "2. 双重过滤\n"
            "   # docker-authz-proxy-ctl peer list --uid 1003 --container bob-web"
        ),
        "expected": (
            "• 只显示同时满足 uid=1003 且包含 bob-web 容器的记录（1条容器级互通）\n"
            "• 用户级互通记录不显示（无容器字段）"
        ),
    },
    {
        "id": "TC-PEER-016",
        "name": "异常 - 重复建立相同互通",
        "purpose": "验证对已存在的互通再次执行 allow 时的行为（幂等或报错）",
        "steps": (
            "1. 建立用户级互通\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004\n"
            "2. 再次执行相同命令\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004\n"
            "3. 查看互通记录\n"
            "   # docker-authz-proxy-ctl peer list"
        ),
        "expected": (
            "• 第二次 allow 不报错（幂等）或返回明确提示\n"
            "• peer list 中该用户对只有一条记录，不重复"
        ),
    },
    {
        "id": "TC-PEER-017",
        "name": "异常 - 撤销不存在的互通",
        "purpose": "验证对不存在的互通执行 deny 时不报错，命令安全退出",
        "steps": (
            "1. 确认无互通记录\n"
            "   # docker-authz-proxy-ctl peer list\n"
            "2. 执行 deny\n"
            "   # docker-authz-proxy-ctl peer deny --uid-a 1003 --uid-b 1004\n"
            "   echo \"exit: $?\""
        ),
        "expected": (
            "• deny 命令正常退出，exit code = 0\n"
            "• 无 panic 或错误堆栈输出"
        ),
    },
    {
        "id": "TC-PEER-018",
        "name": "异常 - 容器归属验证（容器不属于指定用户）",
        "purpose": "验证建立容器级互通时，若容器不属于指定用户则拒绝",
        "steps": (
            "1. bob 启动容器 bob-web\n"
            "   $ su - bob -c \"docker run -d --name bob-web nginx:alpine\"\n"
            "2. 尝试将 bob-web 指定为 alice 的容器建立互通\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1004 --uid-b 1003 \\\n"
            "       --container-a bob-web --container-b bob-web"
        ),
        "expected": (
            "• 命令返回错误，exit code != 0\n"
            "• 错误信息包含 \"belongs to uid=\" 或 \"not uid=1004\""
        ),
    },
    {
        "id": "TC-PEER-019",
        "name": "异常 - 缺少必要参数",
        "purpose": "验证 allow/deny 缺少 uid-a 或 uid-b 时给出明确错误提示",
        "steps": (
            "1. 只指定一个用户\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003\n"
            "   echo \"exit: $?\"\n"
            "2. 不指定任何用户\n"
            "   # docker-authz-proxy-ctl peer allow\n"
            "   echo \"exit: $?\"\n"
            "3. container-a 和 container-b 只指定一个\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004 --container-a bob-web\n"
            "   echo \"exit: $?\""
        ),
        "expected": (
            "• 步骤1 报错 \"必须指定 --uid-b 或 --user-b\"，exit code != 0\n"
            "• 步骤2 报错提示缺少用户参数，exit code != 0\n"
            "• 步骤3 报错 \"--container-a 和 --container-b 必须同时指定或同时省略\"，exit code != 0"
        ),
    },
    {
        "id": "TC-PEER-020",
        "name": "互通隔离性 - bob/alice 互通不影响其他用户",
        "purpose": "验证 bob 和 alice 建立互通后，其他用户（如 ywyh）的容器仍与 bob/alice 隔离",
        "steps": (
            "1. 建立 bob 和 alice 用户级互通\n"
            "   # docker-authz-proxy-ctl peer allow --uid-a 1003 --uid-b 1004\n"
            "2. ywyh 启动容器\n"
            "   $ su - ywyh -c \"docker run -d --name ywyh-test alpine sleep 3600\"\n"
            "3. bob 启动容器\n"
            "   $ su - bob -c \"docker run -d --name bob-isolation-test alpine sleep 3600\"\n"
            "4. 获取 ywyh 容器 IP\n"
            "   # docker inspect ywyh-test --format '{{range .NetworkSettings.Networks}}{{.IPAddress}}{{end}}'\n"
            "5. 从 bob 容器尝试 ping ywyh 容器\n"
            "   # docker exec bob-isolation-test ping -c 3 <ywyh-ip> || echo \"UNREACHABLE\""
        ),
        "expected": (
            "• bob 容器无法 ping 通 ywyh 容器，输出 \"UNREACHABLE\"\n"
            "• bob 和 alice 的互通不影响与其他用户的隔离策略"
        ),
    },
]


def make_header_fill():
    return PatternFill(fill_type="solid", fgColor="1F4E79")


def make_header_font():
    return Font(color="FFFFFF", bold=True, name="微软雅黑", size=11)


def make_data_font():
    return Font(name="微软雅黑", size=10)


def make_wrap_align():
    return Alignment(wrap_text=True, vertical="top")


def make_header_align():
    return Alignment(wrap_text=True, vertical="center", horizontal="center")


def apply_new_sheet(ws):
    # Column widths
    col_widths = {"A": 18, "B": 30, "C": 35, "D": 35, "E": 65, "F": 55}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Header row
    ws.append(HEADERS)
    header_row = ws.max_row
    ws.row_dimensions[header_row].height = 30
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = make_header_fill()
        cell.font = make_header_font()
        cell.alignment = make_header_align()

    # Data rows
    for tc in TESTCASES:
        row_data = [
            tc["id"],
            tc["name"],
            tc["purpose"],
            PRECONDITION,
            tc["steps"],
            tc["expected"],
        ]
        ws.append(row_data)
        row_idx = ws.max_row
        ws.row_dimensions[row_idx].height = 120
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = make_data_font()
            cell.alignment = make_wrap_align()


def update_summary(ws):
    # Find last data row (before 合计 row if present)
    last_row = ws.max_row
    # Check if last row is 合计
    last_val = ws.cell(row=last_row, column=1).value
    if last_val and "合计" in str(last_val):
        # Insert before 合计 row
        ws.insert_rows(last_row)
        target_row = last_row
    else:
        ws.append([None])
        target_row = ws.max_row

    ws.cell(row=target_row, column=1).value = NEW_SHEET
    ws.cell(row=target_row, column=2).value = len(TESTCASES)
    ws.cell(row=target_row, column=3).value = NEW_SHEET

    data_font = make_data_font()
    wrap_align = make_wrap_align()
    for col_idx in range(1, 4):
        cell = ws.cell(row=target_row, column=col_idx)
        cell.font = data_font
        cell.alignment = wrap_align


def main():
    wb = openpyxl.load_workbook(SRC)

    # Add new sheet
    ws_new = wb.create_sheet(title=NEW_SHEET)
    apply_new_sheet(ws_new)

    # Update summary sheet
    summary_name = None
    for name in wb.sheetnames:
        if "汇总" in name:
            summary_name = name
            break
    if summary_name:
        update_summary(wb[summary_name])
    else:
        print("WARNING: summary sheet not found")

    wb.save(DST)
    print(f"Saved to {DST}")
    print(f"Total test cases written: {len(TESTCASES)}")


if __name__ == "__main__":
    main()
