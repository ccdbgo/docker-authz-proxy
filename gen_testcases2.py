#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
docker-authz-proxy 完整测试用例生成器 v2
覆盖：5 种用户 × 全部 Docker 命令 × 所有模块 + 性能/异常测试
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# 样式常量
# ─────────────────────────────────────────────────────────────────────────────
HEADERS = ["测试用例编号", "测试用例名称", "测试目的", "前提条件", "测试步骤", "预期结果"]
COL_WIDTHS = [18, 36, 44, 48, 68, 56]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

# 每个 Sheet 的行背景色
SHEET_COLORS = {
    "TC-SETUP":  "E8F4FD",
    "TC-ROOT":   "D5E8D4",
    "TC-SUDO":   "FFF2CC",
    "TC-DGRP":   "FCE4D6",
    "TC-BOB":    "EAD1DC",
    "TC-ALICE":  "CFE2F3",
    "TC-CROSS":  "E2EFDA",
    "TC-PERF":   "F4CCCC",
    "TC-ERR":    "EAD9FF",
    "TC-AUDIT":  "D9EAD3",
    "TC-DEPLOY": "FFF9C4",
}

thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")


def row_fill(tc_id):
    for prefix, color in SHEET_COLORS.items():
        if tc_id.startswith(prefix):
            return PatternFill("solid", fgColor=color)
    return PatternFill("solid", fgColor="FFFFFF")


def add_sheet(wb, title, cases, first=False):
    ws = wb.active if first else wb.create_sheet(title)
    if first:
        ws.title = title
    ws.freeze_panes = "A2"
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30
    for r, case in enumerate(cases, 2):
        fill = row_fill(case[0])
        for col, val in enumerate(case, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill
            c.border = BORDER
            c.alignment = WRAP
        ws.row_dimensions[r].height = 90
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# 公共前提条件文本
# ─────────────────────────────────────────────────────────────────────────────
PRE_COMMON = (
    "系统环境：\n"
    "• Linux 主机已安装 Docker Engine\n"
    "• docker-authz-proxy 已编译并部署到 /usr/local/bin/\n"
    "• 已创建用户：root(uid=0)、test-sudo(uid=1001,sudo组)、"
    "test-docker-g(uid=1002,docker组)、bob(uid=1003)、alice(uid=1004)\n"
    "• policy.yaml 和 quota.yaml 已按测试策略配置并加载\n"
    "• 代理服务已启动，各用户 socket 已创建\n"
    "• 各用户已设置 DOCKER_HOST=unix:///run/docker-authz/<username>.sock"
)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 0: 环境搭建与策略配置
# ─────────────────────────────────────────────────────────────────────────────
setup_cases = [
    (
        "TC-SETUP-001",
        "创建测试用户 test-sudo",
        "创建 test-sudo 用户并加入 sudo 组，验证用户创建成功",
        "Linux 系统，具有 root 权限",
        "1. 执行：useradd -m -s /bin/bash test-sudo\n"
        "2. 执行：usermod -aG sudo test-sudo\n"
        "3. 执行：id test-sudo",
        "id 输出包含 sudo 组；用户主目录 /home/test-sudo 已创建"
    ),
    (
        "TC-SETUP-002",
        "创建测试用户 test-docker-g",
        "创建 test-docker-g 用户并加入 docker 组，验证用户创建成功",
        "Linux 系统，具有 root 权限",
        "1. 执行：useradd -m -s /bin/bash test-docker-g\n"
        "2. 执行：usermod -aG docker test-docker-g\n"
        "3. 执行：id test-docker-g",
        "id 输出包含 docker 组；用户主目录已创建"
    ),
    (
        "TC-SETUP-003",
        "创建普通测试用户 bob 和 alice",
        "创建普通用户 bob 和 alice，不加入任何特权组",
        "Linux 系统，具有 root 权限",
        "1. 执行：useradd -m -s /bin/bash bob\n"
        "2. 执行：useradd -m -s /bin/bash alice\n"
        "3. 执行：id bob && id alice",
        "两个用户均创建成功；id 输出不包含 sudo 或 docker 组"
    ),
    (
        "TC-SETUP-004",
        "配置 policy.yaml - root 用户策略",
        "为 root 用户配置访问控制策略：允许所有操作（无 deny_rules）",
        "policy.yaml 文件可写",
        "1. 在 policy.yaml 中为 root 配置：无 deny_rules（默认全允许）\n"
        "2. 重启代理或发送 SIGHUP\n"
        "3. 以 root 身份执行 docker info",
        "docker info 返回正常；代理日志显示策略已加载"
    ),
    (
        "TC-SETUP-005",
        "配置 policy.yaml - test-sudo 用户策略",
        "为 test-sudo 配置策略：禁止 push、swarm、plugin、secret、config 操作",
        "policy.yaml 文件可写",
        "1. 在 deny_rules 中添加：\n"
        "   - users: [test-sudo]\n"
        "     actions: [push, swarm, plugin, secret, config]\n"
        "2. 重载策略\n"
        "3. 以 test-sudo 身份执行 docker ps",
        "docker ps 正常返回；策略加载成功"
    ),
    (
        "TC-SETUP-006",
        "配置 policy.yaml - test-docker-g 用户策略",
        "为 test-docker-g 配置策略：禁止 build、push、commit、load、save、swarm、plugin、secret、config",
        "policy.yaml 文件可写",
        "1. 在 deny_rules 中添加：\n"
        "   - users: [test-docker-g]\n"
        "     actions: [build, push, commit, load, save, swarm, plugin, secret, config]\n"
        "2. 重载策略",
        "策略加载成功；代理日志无错误"
    ),
    (
        "TC-SETUP-007",
        "配置 policy.yaml - bob 用户策略",
        "为 bob 配置策略：禁止 exec、build、push、commit、load、save、swarm、plugin、secret、config",
        "policy.yaml 文件可写",
        "1. 在 deny_rules 中添加：\n"
        "   - users: [bob]\n"
        "     actions: [exec, build, push, commit, load, save, swarm, plugin, secret, config]\n"
        "2. 重载策略",
        "策略加载成功"
    ),
    (
        "TC-SETUP-008",
        "配置 policy.yaml - alice 用户策略",
        "为 alice 配置最严格策略：禁止 exec、build、push、commit、load、save、prune、swarm、plugin、secret、config",
        "policy.yaml 文件可写",
        "1. 在 deny_rules 中添加：\n"
        "   - users: [alice]\n"
        "     actions: [exec, build, push, commit, load, save, prune, swarm, plugin, secret, config]\n"
        "2. 重载策略",
        "策略加载成功"
    ),
    (
        "TC-SETUP-009",
        "配置 quota.yaml - 各用户资源配额",
        "为各用户配置差异化资源配额",
        "quota.yaml 文件可写",
        "1. 配置 quota.yaml：\n"
        "   defaults: cpu_cores:2, mem_mb:2048, max_containers:5\n"
        "   users:\n"
        "     root: cpu_cores:0, mem_mb:0, max_containers:0 (无限制)\n"
        "     test-sudo: cpu_cores:4, mem_mb:4096, max_containers:10\n"
        "     test-docker-g: cpu_cores:2, mem_mb:2048, max_containers:5\n"
        "     bob: cpu_cores:1, mem_mb:1024, max_containers:3\n"
        "     alice: cpu_cores:1, mem_mb:512, max_containers:2\n"
        "2. 重载代理",
        "配额配置加载成功；代理日志无错误"
    ),
    (
        "TC-SETUP-010",
        "验证各用户 socket 已创建",
        "确认代理为所有测试用户创建了独立的 Unix socket",
        "代理服务已启动",
        "1. 执行：ls -la /run/docker-authz/\n"
        "2. 检查每个 socket 的权限",
        "存在 root.sock、test-sudo.sock、test-docker-g.sock、bob.sock、alice.sock；"
        "每个 socket 权限为 srw------- (600)，所有者为对应用户"
    ),
    (
        "TC-SETUP-011",
        "验证 DOCKER_HOST 环境变量配置",
        "确认各用户登录后 DOCKER_HOST 自动指向其专属 socket",
        "profile.d 脚本已部署",
        "1. 以 alice 身份登录新 shell\n"
        "2. 执行：echo $DOCKER_HOST\n"
        "3. 对其他用户重复",
        "各用户的 DOCKER_HOST 分别为 unix:///run/docker-authz/<username>.sock"
    ),
]  # end setup_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: root 用户 - 全量 Docker 命令遍历
# ─────────────────────────────────────────────────────────────────────────────
PRE_ROOT = PRE_COMMON + "\n• 当前操作用户：root（uid=0）\n• root 无 deny_rules，配额无限制"

root_cases = [
    (
        "TC-ROOT-001",
        "root - docker pull 拉取镜像",
        "验证 root 可以拉取任意镜像",
        PRE_ROOT,
        "1. 执行：docker pull nginx:latest\n2. 执行：docker pull alpine:3.18",
        "两个镜像均拉取成功；操作日志记录 action=pull, result=allow"
    ),
    (
        "TC-ROOT-002",
        "root - docker images 查看所有镜像",
        "验证 root 可以看到系统中所有用户的镜像（无过滤）",
        PRE_ROOT + "\n• alice 和 bob 各已构建私有镜像",
        "1. 执行：docker images\n2. 执行：docker images --all",
        "返回所有用户的镜像（包括 alice、bob 的私有镜像）；无过滤"
    ),
    (
        "TC-ROOT-003",
        "root - docker run 创建并启动容器（无配额限制）",
        "验证 root 可以创建容器，且无资源配额限制",
        PRE_ROOT,
        "1. 执行：docker run -d --name root-nginx nginx:latest\n"
        "2. 执行：docker inspect root-nginx | grep -E 'NanoCPUs|Memory'",
        "容器成功创建并运行；NanoCPUs=0（无限制），Memory=0（无限制）"
    ),
    (
        "TC-ROOT-004",
        "root - docker ps 查看所有容器",
        "验证 root 的 docker ps 返回所有用户的容器（无过滤）",
        PRE_ROOT + "\n• alice、bob、test-sudo 各有运行中的容器",
        "1. 执行：docker ps\n2. 执行：docker ps -a",
        "返回所有用户的容器，包括其他用户创建的容器；无过滤"
    ),
    (
        "TC-ROOT-005",
        "root - docker exec 进入任意用户容器",
        "验证 root 可以 exec 进入任意用户的容器",
        PRE_ROOT + "\n• alice 已创建容器 user-1004-alice-app",
        "1. 执行：docker exec root-nginx echo hello\n"
        "2. 执行：docker exec user-1004-alice-app echo hello",
        "两个 exec 均成功执行；返回 hello"
    ),
    (
        "TC-ROOT-006",
        "root - docker stop/start/restart 容器",
        "验证 root 可以停止、启动、重启任意容器",
        PRE_ROOT + "\n• root-nginx 容器正在运行",
        "1. 执行：docker stop root-nginx\n"
        "2. 执行：docker start root-nginx\n"
        "3. 执行：docker restart root-nginx",
        "三个操作均成功；容器状态依次变为 Exited → Running → Running"
    ),
    (
        "TC-ROOT-007",
        "root - docker logs/stats/top",
        "验证 root 可以查看任意容器的日志、统计和进程",
        PRE_ROOT + "\n• root-nginx 容器正在运行",
        "1. 执行：docker logs root-nginx\n"
        "2. 执行：docker stats root-nginx --no-stream\n"
        "3. 执行：docker top root-nginx",
        "三个命令均返回正常数据"
    ),
    (
        "TC-ROOT-008",
        "root - docker inspect 检查容器/镜像",
        "验证 root 可以 inspect 任意容器和镜像，且系统标签正确",
        PRE_ROOT,
        "1. 执行：docker inspect root-nginx\n"
        "2. 执行：docker inspect nginx:latest\n"
        "3. 检查 Labels 中的 system.authz.owner.uid",
        "返回完整 JSON；system.authz.owner.uid=0，system.authz.owner=root"
    ),
    (
        "TC-ROOT-009",
        "root - docker cp 文件复制",
        "验证 root 可以在容器和宿主机之间复制文件",
        PRE_ROOT + "\n• root-nginx 容器正在运行",
        "1. 执行：echo 'test' > /tmp/test.txt\n"
        "2. 执行：docker cp /tmp/test.txt root-nginx:/tmp/\n"
        "3. 执行：docker cp root-nginx:/tmp/test.txt /tmp/out.txt",
        "文件复制成功；/tmp/out.txt 内容为 test"
    ),
    (
        "TC-ROOT-010",
        "root - docker commit 提交镜像",
        "验证 root 可以将容器提交为新镜像",
        PRE_ROOT + "\n• root-nginx 容器正在运行",
        "1. 执行：docker commit root-nginx root-nginx-custom:v1\n"
        "2. 执行：docker images root-nginx-custom",
        "镜像 root-nginx-custom:v1 创建成功"
    ),
    (
        "TC-ROOT-011",
        "root - docker build 构建镜像",
        "验证 root 可以构建 Docker 镜像",
        PRE_ROOT + "\n• 已准备 Dockerfile（FROM alpine; RUN echo hello）",
        "1. 创建 /tmp/testbuild/Dockerfile\n"
        "2. 执行：docker build -t root-test-img /tmp/testbuild/",
        "镜像构建成功；docker images 可见 root-test-img"
    ),
    (
        "TC-ROOT-012",
        "root - docker tag 标记镜像",
        "验证 root 可以为镜像打标签",
        PRE_ROOT + "\n• root-test-img 镜像已存在",
        "1. 执行：docker tag root-test-img root-test-img:v2",
        "标签创建成功；docker images 显示 root-test-img:v2"
    ),
    (
        "TC-ROOT-013",
        "root - docker save/load 镜像导出导入",
        "验证 root 可以导出和导入镜像",
        PRE_ROOT + "\n• root-test-img 镜像已存在",
        "1. 执行：docker save root-test-img -o /tmp/root-test-img.tar\n"
        "2. 执行：docker rmi root-test-img:v2\n"
        "3. 执行：docker load -i /tmp/root-test-img.tar",
        "save 生成 tar 文件；load 成功恢复镜像"
    ),
    (
        "TC-ROOT-014",
        "root - docker rm/rmi 删除容器和镜像",
        "验证 root 可以删除任意容器和镜像",
        PRE_ROOT + "\n• root-nginx 容器已停止",
        "1. 执行：docker rm root-nginx\n"
        "2. 执行：docker rmi root-test-img",
        "容器和镜像均删除成功"
    ),
    (
        "TC-ROOT-015",
        "root - docker network ls/create/inspect/rm",
        "验证 root 可以管理所有网络，且 network ls 无过滤",
        PRE_ROOT,
        "1. 执行：docker network ls\n"
        "2. 执行：docker network create root-net\n"
        "3. 执行：docker network inspect root-net\n"
        "4. 执行：docker network rm root-net",
        "所有操作成功；network ls 显示所有用户的网络"
    ),
    (
        "TC-ROOT-016",
        "root - docker network connect/disconnect",
        "验证 root 可以将容器连接/断开任意网络",
        PRE_ROOT + "\n• root-nginx 容器运行中",
        "1. 执行：docker network create root-net2\n"
        "2. 执行：docker run -d --name root-net-test nginx\n"
        "3. 执行：docker network connect root-net2 root-net-test\n"
        "4. 执行：docker network disconnect root-net2 root-net-test\n"
        "5. 执行：docker rm -f root-net-test && docker network rm root-net2",
        "连接和断开均成功"
    ),
    (
        "TC-ROOT-017",
        "root - docker volume ls/create/inspect/rm",
        "验证 root 可以管理所有卷，且 volume ls 无过滤",
        PRE_ROOT,
        "1. 执行：docker volume ls\n"
        "2. 执行：docker volume create root-vol\n"
        "3. 执行：docker volume inspect root-vol\n"
        "4. 执行：docker volume rm root-vol",
        "所有操作成功；volume ls 显示所有用户的卷"
    ),
    (
        "TC-ROOT-018",
        "root - docker info/version/system df",
        "验证 root 可以查看系统信息",
        PRE_ROOT,
        "1. 执行：docker info\n"
        "2. 执行：docker version\n"
        "3. 执行：docker system df",
        "三个命令均返回正常数据"
    ),
    (
        "TC-ROOT-019",
        "root - docker system prune 系统清理",
        "验证 root 可以执行系统清理",
        PRE_ROOT + "\n• 存在已停止的容器和悬空镜像",
        "1. 执行：docker system prune -f",
        "清理成功；已停止容器和悬空镜像被删除"
    ),
    (
        "TC-ROOT-020",
        "root - docker events 事件流",
        "验证 root 可以监听 Docker 事件流",
        PRE_ROOT,
        "1. 后台执行：docker events --since 1m &\n"
        "2. 执行：docker run --rm alpine echo hi\n"
        "3. 等待 2 秒后终止 events",
        "events 输出包含容器创建、启动、停止、删除事件"
    ),
    (
        "TC-ROOT-021",
        "root - docker pause/unpause 容器",
        "验证 root 可以暂停和恢复容器",
        PRE_ROOT,
        "1. 执行：docker run -d --name root-pause-test nginx\n"
        "2. 执行：docker pause root-pause-test\n"
        "3. 执行：docker unpause root-pause-test\n"
        "4. 执行：docker rm -f root-pause-test",
        "pause 后容器状态为 Paused；unpause 后恢复 Running"
    ),
    (
        "TC-ROOT-022",
        "root - docker rename 重命名容器",
        "验证 root 可以重命名容器",
        PRE_ROOT,
        "1. 执行：docker run -d --name root-old-name nginx\n"
        "2. 执行：docker rename root-old-name root-new-name\n"
        "3. 执行：docker ps | grep root-new-name\n"
        "4. 执行：docker rm -f root-new-name",
        "重命名成功；docker ps 显示新名称"
    ),
    (
        "TC-ROOT-023",
        "root - docker kill 强制终止容器",
        "验证 root 可以向容器发送信号",
        PRE_ROOT,
        "1. 执行：docker run -d --name root-kill-test nginx\n"
        "2. 执行：docker kill root-kill-test\n"
        "3. 执行：docker rm root-kill-test",
        "容器被终止；状态变为 Exited"
    ),
    (
        "TC-ROOT-024",
        "root - docker export/import 容器导出导入",
        "验证 root 可以导出容器文件系统并导入为镜像",
        PRE_ROOT,
        "1. 执行：docker run -d --name root-export-test alpine\n"
        "2. 执行：docker export root-export-test -o /tmp/root-export.tar\n"
        "3. 执行：docker import /tmp/root-export.tar root-imported:v1\n"
        "4. 执行：docker rm -f root-export-test",
        "export 生成 tar；import 创建镜像 root-imported:v1"
    ),
    (
        "TC-ROOT-025",
        "root - docker diff/history/port/wait",
        "验证 root 可以使用容器诊断命令",
        PRE_ROOT,
        "1. 执行：docker run -d --name root-diag -p 8081:80 nginx\n"
        "2. 执行：docker diff root-diag\n"
        "3. 执行：docker port root-diag\n"
        "4. 执行：docker history nginx:latest\n"
        "5. 执行：docker rm -f root-diag",
        "所有命令均返回正常数据"
    ),
    (
        "TC-ROOT-026",
        "root - docker update 更新容器资源",
        "验证 root 可以动态更新容器资源限制",
        PRE_ROOT,
        "1. 执行：docker run -d --name root-update-test nginx\n"
        "2. 执行：docker update --cpus=0.5 --memory=256m root-update-test\n"
        "3. 执行：docker inspect root-update-test | grep -E 'NanoCPUs|Memory'\n"
        "4. 执行：docker rm -f root-update-test",
        "update 成功；NanoCPUs=500000000，Memory=268435456"
    ),
    (
        "TC-ROOT-027",
        "root - docker search 搜索镜像",
        "验证 root 可以搜索 Docker Hub 镜像",
        PRE_ROOT + "\n• 网络可访问 Docker Hub",
        "1. 执行：docker search nginx --limit 5",
        "返回搜索结果列表"
    ),
    (
        "TC-ROOT-028",
        "root - 系统标签验证",
        "验证 root 创建的容器自动注入正确的系统标签",
        PRE_ROOT,
        "1. 执行：docker run -d --name root-label-test nginx\n"
        "2. 执行：docker inspect root-label-test | grep system.authz\n"
        "3. 执行：docker rm -f root-label-test",
        "Labels 包含：system.authz.owner.uid=0，system.authz.owner=root，system.authz.created_by=docker"
    ),
    (
        "TC-ROOT-029",
        "root - 操作审计日志验证",
        "验证 root 的所有操作均被记录到审计日志",
        PRE_ROOT,
        "1. 执行若干 docker 命令（ps、run、rm 等）\n"
        "2. 查看 /var/log/docker-authz/user-operation/root.log",
        "日志文件包含所有操作记录；每条记录为有效 JSON；包含 time、user=root、uid=0、action、result=allow"
    ),
    (
        "TC-ROOT-030",
        "root - 操作其他用户容器（所有权绕过）",
        "验证 root 可以操作任意用户的容器（不受所有权限制）",
        PRE_ROOT + "\n• bob 已创建容器 user-1003-bob-app",
        "1. 以 root 身份执行：docker stop user-1003-bob-app\n"
        "2. 执行：docker start user-1003-bob-app\n"
        "3. 执行：docker rm -f user-1003-bob-app",
        "所有操作成功；root 不受所有权限制"
    ),
]  # end root_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: test-sudo 用户
# 策略：禁止 push、swarm、plugin、secret、config
# 配额：cpu=4, mem=4096MB, max_containers=10
# 特性：sudo 用户，eUID=0 但 loginUID=1001，资源按原始用户隔离
# ─────────────────────────────────────────────────────────────────────────────
PRE_SUDO = (
    PRE_COMMON + "\n"
    "• 当前操作用户：test-sudo（uid=1001，sudo 组）\n"
    "• deny_rules：禁止 push、swarm、plugin、secret、config\n"
    "• 配额：cpu_cores=4, mem_mb=4096, max_containers=10\n"
    "• 以 sudo docker <cmd> 方式执行（eUID=0，loginUID=1001）"
)

sudo_cases = [
    (
        "TC-SUDO-001",
        "test-sudo - docker ps（允许）",
        "验证 test-sudo 可以执行 docker ps，且只看到自己的容器",
        PRE_SUDO + "\n• bob 已有运行中的容器",
        "1. 以 test-sudo 身份执行：sudo docker ps\n"
        "2. 执行：sudo docker ps -a",
        "只返回 test-sudo 自己的容器；bob 的容器不可见；result=allow"
    ),
    (
        "TC-SUDO-002",
        "test-sudo - docker pull（允许）",
        "验证 test-sudo 可以拉取镜像",
        PRE_SUDO,
        "1. 执行：sudo docker pull redis:7\n"
        "2. 执行：sudo docker images",
        "redis:7 拉取成功；docker images 可见"
    ),
    (
        "TC-SUDO-003",
        "test-sudo - docker run（允许，配额注入）",
        "验证 test-sudo 创建容器时自动注入 4CPU/4096MB 配额",
        PRE_SUDO,
        "1. 执行：sudo docker run -d --name sudo-app redis:7\n"
        "2. 执行：sudo docker inspect sudo-app | grep -E 'NanoCPUs|Memory'",
        "容器创建成功；NanoCPUs=4000000000，Memory=4294967296"
    ),
    (
        "TC-SUDO-004",
        "test-sudo - docker exec（允许）",
        "验证 test-sudo 可以 exec 进入自己的容器",
        PRE_SUDO + "\n• sudo-app 容器运行中",
        "1. 执行：sudo docker exec sudo-app redis-cli ping",
        "返回 PONG；exec 成功"
    ),
    (
        "TC-SUDO-005",
        "test-sudo - docker exec 进入他人容器（拒绝）",
        "验证 test-sudo 无法 exec 进入其他用户的容器",
        PRE_SUDO + "\n• bob 已创建容器 user-1003-bob-app",
        "1. 执行：sudo docker exec user-1003-bob-app echo hi",
        "返回 403 Forbidden；日志记录 deny_reason=ownership"
    ),
    (
        "TC-SUDO-006",
        "test-sudo - docker stop/start/restart（允许）",
        "验证 test-sudo 可以管理自己的容器生命周期",
        PRE_SUDO + "\n• sudo-app 容器运行中",
        "1. 执行：sudo docker stop sudo-app\n"
        "2. 执行：sudo docker start sudo-app\n"
        "3. 执行：sudo docker restart sudo-app",
        "三个操作均成功"
    ),
    (
        "TC-SUDO-007",
        "test-sudo - docker build（允许）",
        "验证 test-sudo 可以构建镜像",
        PRE_SUDO + "\n• 已准备 Dockerfile",
        "1. 创建 /tmp/sudobuild/Dockerfile（FROM alpine）\n"
        "2. 执行：sudo docker build -t sudo-test-img /tmp/sudobuild/",
        "镜像构建成功"
    ),
    (
        "TC-SUDO-008",
        "test-sudo - docker push（拒绝）",
        "验证 test-sudo 被策略禁止执行 docker push",
        PRE_SUDO + "\n• sudo-test-img 镜像已存在",
        "1. 执行：sudo docker tag sudo-test-img registry.example.com/sudo-test-img\n"
        "2. 执行：sudo docker push registry.example.com/sudo-test-img",
        "返回 403 Forbidden；日志记录 action=push, result=deny, deny_reason=policy"
    ),
    (
        "TC-SUDO-009",
        "test-sudo - docker commit（允许）",
        "验证 test-sudo 可以提交容器为镜像",
        PRE_SUDO + "\n• sudo-app 容器运行中",
        "1. 执行：sudo docker commit sudo-app sudo-committed:v1",
        "镜像创建成功"
    ),
    (
        "TC-SUDO-010",
        "test-sudo - docker save/load（允许）",
        "验证 test-sudo 可以导出和导入镜像",
        PRE_SUDO + "\n• sudo-test-img 镜像已存在",
        "1. 执行：sudo docker save sudo-test-img -o /tmp/sudo-img.tar\n"
        "2. 执行：sudo docker load -i /tmp/sudo-img.tar",
        "save 和 load 均成功"
    ),
    (
        "TC-SUDO-011",
        "test-sudo - docker images（只看自己的镜像）",
        "验证 test-sudo 的 docker images 只返回自己的镜像和公共镜像",
        PRE_SUDO + "\n• alice 已构建私有镜像 alice-private-img",
        "1. 执行：sudo docker images",
        "只显示 test-sudo 的镜像和公共镜像；alice-private-img 不可见"
    ),
    (
        "TC-SUDO-012",
        "test-sudo - docker rmi（允许，只能删自己的镜像）",
        "验证 test-sudo 可以删除自己的镜像，但不能删除他人镜像",
        PRE_SUDO + "\n• sudo-committed:v1 镜像已存在\n• alice 有私有镜像",
        "1. 执行：sudo docker rmi sudo-committed:v1（应成功）\n"
        "2. 尝试删除 alice 的镜像（应失败）",
        "步骤 1 成功；步骤 2 返回 403"
    ),
    (
        "TC-SUDO-013",
        "test-sudo - docker network 管理（允许）",
        "验证 test-sudo 可以管理自己的网络，且网络名自动加前缀",
        PRE_SUDO,
        "1. 执行：sudo docker network create mynet\n"
        "2. 执行：sudo docker network ls\n"
        "3. 以 root 查看实际网络名\n"
        "4. 执行：sudo docker network rm mynet",
        "test-sudo 看到的网络名为 mynet；实际名为 test-sudo_u1001_mynet；network ls 只显示自己的网络"
    ),
    (
        "TC-SUDO-014",
        "test-sudo - docker volume 管理（允许）",
        "验证 test-sudo 可以管理自己的卷",
        PRE_SUDO,
        "1. 执行：sudo docker volume create myvol\n"
        "2. 执行：sudo docker volume ls\n"
        "3. 执行：sudo docker volume rm myvol",
        "卷创建成功；volume ls 只显示自己的卷；删除成功"
    ),
    (
        "TC-SUDO-015",
        "test-sudo - docker swarm（拒绝）",
        "验证 test-sudo 被策略禁止执行 swarm 操作",
        PRE_SUDO,
        "1. 执行：sudo docker swarm init",
        "返回 403 Forbidden；日志记录 action=swarm, result=deny"
    ),
    (
        "TC-SUDO-016",
        "test-sudo - docker plugin（拒绝）",
        "验证 test-sudo 被策略禁止执行 plugin 操作",
        PRE_SUDO,
        "1. 执行：sudo docker plugin ls",
        "返回 403 Forbidden；日志记录 action=plugin, result=deny"
    ),
    (
        "TC-SUDO-017",
        "test-sudo - docker secret（拒绝）",
        "验证 test-sudo 被策略禁止执行 secret 操作",
        PRE_SUDO,
        "1. 执行：sudo docker secret ls",
        "返回 403 Forbidden；日志记录 action=secret, result=deny"
    ),
    (
        "TC-SUDO-018",
        "test-sudo - sudo 身份识别验证",
        "验证代理正确识别 sudo 用户的真实 UID（loginUID=1001，非 0）",
        PRE_SUDO,
        "1. 以 test-sudo 执行：sudo docker ps\n"
        "2. 查看 /var/log/docker-authz/auth.log",
        "auth.log 中记录：effective_uid=0，real_uid=1001，switched_identity=true，username=test-sudo"
    ),
    (
        "TC-SUDO-019",
        "test-sudo - 容器数量配额限制",
        "验证 test-sudo 创建容器超过 10 个时被拒绝",
        PRE_SUDO + "\n• test-sudo 已有 10 个运行中的容器",
        "1. 尝试执行：sudo docker run -d nginx",
        "返回 403 或 429；错误信息说明已达容器数量上限（max_containers=10）"
    ),
    (
        "TC-SUDO-020",
        "test-sudo - docker logs/stats/top/inspect（允许）",
        "验证 test-sudo 可以查看自己容器的详细信息",
        PRE_SUDO + "\n• sudo-app 容器运行中",
        "1. 执行：sudo docker logs sudo-app\n"
        "2. 执行：sudo docker stats sudo-app --no-stream\n"
        "3. 执行：sudo docker top sudo-app\n"
        "4. 执行：sudo docker inspect sudo-app",
        "四个命令均返回正常数据"
    ),
    (
        "TC-SUDO-021",
        "test-sudo - docker cp（允许）",
        "验证 test-sudo 可以在自己的容器和宿主机之间复制文件",
        PRE_SUDO + "\n• sudo-app 容器运行中",
        "1. 执行：echo 'sudo-test' > /tmp/sudo-test.txt\n"
        "2. 执行：sudo docker cp /tmp/sudo-test.txt sudo-app:/tmp/\n"
        "3. 执行：sudo docker cp sudo-app:/tmp/sudo-test.txt /tmp/sudo-out.txt",
        "文件复制成功"
    ),
    (
        "TC-SUDO-022",
        "test-sudo - docker system prune（允许）",
        "验证 test-sudo 可以执行系统清理",
        PRE_SUDO + "\n• 存在已停止的容器",
        "1. 执行：sudo docker system prune -f",
        "清理成功；只清理 test-sudo 自己的资源"
    ),
    (
        "TC-SUDO-023",
        "test-sudo - docker rm（允许，只能删自己的容器）",
        "验证 test-sudo 可以删除自己的容器，但不能删除他人容器",
        PRE_SUDO + "\n• sudo-app 容器已停止\n• bob 有容器 user-1003-bob-app",
        "1. 执行：sudo docker rm sudo-app（应成功）\n"
        "2. 执行：sudo docker rm user-1003-bob-app（应失败）",
        "步骤 1 成功；步骤 2 返回 403"
    ),
    (
        "TC-SUDO-024",
        "test-sudo - docker info/version/system df（允许）",
        "验证 test-sudo 可以查看系统信息",
        PRE_SUDO,
        "1. 执行：sudo docker info\n"
        "2. 执行：sudo docker version\n"
        "3. 执行：sudo docker system df",
        "三个命令均返回正常数据"
    ),
    (
        "TC-SUDO-025",
        "test-sudo - 系统标签验证",
        "验证 test-sudo 创建的容器系统标签显示真实用户身份",
        PRE_SUDO,
        "1. 执行：sudo docker run -d --name sudo-label-test nginx\n"
        "2. 执行：sudo docker inspect sudo-label-test | grep system.authz\n"
        "3. 执行：sudo docker rm -f sudo-label-test",
        "system.authz.owner.uid=1001（真实 UID），system.authz.owner=test-sudo"
    ),
]  # end sudo_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: test-docker-g 用户
# 策略：禁止 build、push、commit、load、save、swarm、plugin、secret、config
# 配额：cpu=2, mem=2048MB, max_containers=5
# 特性：docker 组成员，可直接访问 /var/run/docker.sock（但通过代理）
# ─────────────────────────────────────────────────────────────────────────────
PRE_DGRP = (
    PRE_COMMON + "\n"
    "• 当前操作用户：test-docker-g（uid=1002，docker 组）\n"
    "• deny_rules：禁止 build、push、commit、load、save、swarm、plugin、secret、config\n"
    "• 配额：cpu_cores=2, mem_mb=2048, max_containers=5"
)

dgrp_cases = [
    (
        "TC-DGRP-001",
        "test-docker-g - docker ps（允许，只看自己容器）",
        "验证 test-docker-g 可以执行 docker ps，且只看到自己的容器",
        PRE_DGRP + "\n• alice 已有运行中的容器",
        "1. 以 test-docker-g 身份执行：docker ps\n"
        "2. 执行：docker ps -a",
        "只返回 test-docker-g 自己的容器；alice 的容器不可见"
    ),
    (
        "TC-DGRP-002",
        "test-docker-g - docker pull（允许）",
        "验证 test-docker-g 可以拉取镜像",
        PRE_DGRP,
        "1. 执行：docker pull postgres:15",
        "postgres:15 拉取成功"
    ),
    (
        "TC-DGRP-003",
        "test-docker-g - docker run（允许，配额注入）",
        "验证 test-docker-g 创建容器时自动注入 2CPU/2048MB 配额",
        PRE_DGRP,
        "1. 执行：docker run -d --name dgrp-app postgres:15\n"
        "2. 执行：docker inspect dgrp-app | grep -E 'NanoCPUs|Memory'",
        "容器创建成功；NanoCPUs=2000000000，Memory=2147483648"
    ),
    (
        "TC-DGRP-004",
        "test-docker-g - docker exec（允许）",
        "验证 test-docker-g 可以 exec 进入自己的容器",
        PRE_DGRP + "\n• dgrp-app 容器运行中",
        "1. 执行：docker exec dgrp-app echo hello",
        "返回 hello；exec 成功"
    ),
    (
        "TC-DGRP-005",
        "test-docker-g - docker build（拒绝）",
        "验证 test-docker-g 被策略禁止执行 docker build",
        PRE_DGRP + "\n• 已准备 Dockerfile",
        "1. 创建 /tmp/dgrpbuild/Dockerfile\n"
        "2. 执行：docker build -t dgrp-img /tmp/dgrpbuild/",
        "返回 403 Forbidden；日志记录 action=build, result=deny, deny_reason=policy"
    ),
    (
        "TC-DGRP-006",
        "test-docker-g - docker push（拒绝）",
        "验证 test-docker-g 被策略禁止执行 docker push",
        PRE_DGRP,
        "1. 执行：docker push registry.example.com/dgrp-img",
        "返回 403 Forbidden；日志记录 action=push, result=deny"
    ),
    (
        "TC-DGRP-007",
        "test-docker-g - docker commit（拒绝）",
        "验证 test-docker-g 被策略禁止执行 docker commit",
        PRE_DGRP + "\n• dgrp-app 容器运行中",
        "1. 执行：docker commit dgrp-app dgrp-committed:v1",
        "返回 403 Forbidden；日志记录 action=commit, result=deny"
    ),
    (
        "TC-DGRP-008",
        "test-docker-g - docker save（拒绝）",
        "验证 test-docker-g 被策略禁止执行 docker save",
        PRE_DGRP + "\n• postgres:15 镜像已存在",
        "1. 执行：docker save postgres:15 -o /tmp/dgrp-pg.tar",
        "返回 403 Forbidden；日志记录 action=save, result=deny"
    ),
    (
        "TC-DGRP-009",
        "test-docker-g - docker load（拒绝）",
        "验证 test-docker-g 被策略禁止执行 docker load",
        PRE_DGRP + "\n• /tmp/some-img.tar 文件已存在",
        "1. 执行：docker load -i /tmp/some-img.tar",
        "返回 403 Forbidden；日志记录 action=load, result=deny"
    ),
    (
        "TC-DGRP-010",
        "test-docker-g - docker tag（允许）",
        "验证 test-docker-g 可以为自己的镜像打标签",
        PRE_DGRP + "\n• postgres:15 镜像已存在",
        "1. 执行：docker tag postgres:15 dgrp-pg:custom",
        "标签创建成功"
    ),
    (
        "TC-DGRP-011",
        "test-docker-g - docker images（只看自己和公共镜像）",
        "验证 test-docker-g 的 docker images 只返回自己的镜像和公共镜像",
        PRE_DGRP + "\n• test-sudo 已有私有镜像",
        "1. 执行：docker images",
        "只显示 test-docker-g 的镜像和公共镜像；test-sudo 的私有镜像不可见"
    ),
    (
        "TC-DGRP-012",
        "test-docker-g - docker rmi（允许，只能删自己的镜像）",
        "验证 test-docker-g 可以删除自己的镜像",
        PRE_DGRP + "\n• dgrp-pg:custom 镜像已存在",
        "1. 执行：docker rmi dgrp-pg:custom",
        "镜像删除成功"
    ),
    (
        "TC-DGRP-013",
        "test-docker-g - docker stop/rm（允许，只能操作自己的容器）",
        "验证 test-docker-g 可以停止和删除自己的容器",
        PRE_DGRP + "\n• dgrp-app 容器运行中",
        "1. 执行：docker stop dgrp-app\n"
        "2. 执行：docker rm dgrp-app",
        "两个操作均成功"
    ),
    (
        "TC-DGRP-014",
        "test-docker-g - docker network 管理（允许）",
        "验证 test-docker-g 可以管理自己的网络",
        PRE_DGRP,
        "1. 执行：docker network create dgrp-net\n"
        "2. 执行：docker network ls\n"
        "3. 执行：docker network rm dgrp-net",
        "网络创建成功；network ls 只显示自己的网络；删除成功"
    ),
    (
        "TC-DGRP-015",
        "test-docker-g - docker volume 管理（允许）",
        "验证 test-docker-g 可以管理自己的卷",
        PRE_DGRP,
        "1. 执行：docker volume create dgrp-vol\n"
        "2. 执行：docker volume ls\n"
        "3. 执行：docker volume rm dgrp-vol",
        "卷创建成功；volume ls 只显示自己的卷；删除成功"
    ),
    (
        "TC-DGRP-016",
        "test-docker-g - docker swarm（拒绝）",
        "验证 test-docker-g 被策略禁止执行 swarm 操作",
        PRE_DGRP,
        "1. 执行：docker swarm init",
        "返回 403 Forbidden"
    ),
    (
        "TC-DGRP-017",
        "test-docker-g - docker logs/stats/top/inspect（允许）",
        "验证 test-docker-g 可以查看自己容器的详细信息",
        PRE_DGRP + "\n• 重新创建 dgrp-app 容器",
        "1. 执行：docker run -d --name dgrp-app2 postgres:15\n"
        "2. 执行：docker logs dgrp-app2\n"
        "3. 执行：docker stats dgrp-app2 --no-stream\n"
        "4. 执行：docker inspect dgrp-app2\n"
        "5. 执行：docker rm -f dgrp-app2",
        "四个命令均返回正常数据"
    ),
    (
        "TC-DGRP-018",
        "test-docker-g - 容器数量配额限制",
        "验证 test-docker-g 创建容器超过 5 个时被拒绝",
        PRE_DGRP + "\n• test-docker-g 已有 5 个运行中的容器",
        "1. 尝试执行：docker run -d nginx",
        "返回 403 或 429；错误信息说明已达容器数量上限（max_containers=5）"
    ),
    (
        "TC-DGRP-019",
        "test-docker-g - docker info/version（允许）",
        "验证 test-docker-g 可以查看系统信息",
        PRE_DGRP,
        "1. 执行：docker info\n"
        "2. 执行：docker version",
        "两个命令均返回正常数据"
    ),
    (
        "TC-DGRP-020",
        "test-docker-g - docker system prune（允许）",
        "验证 test-docker-g 可以执行系统清理",
        PRE_DGRP,
        "1. 执行：docker system prune -f",
        "清理成功；只清理 test-docker-g 自己的资源"
    ),
    (
        "TC-DGRP-021",
        "test-docker-g - docker cp（允许）",
        "验证 test-docker-g 可以在自己的容器和宿主机之间复制文件",
        PRE_DGRP + "\n• dgrp-app2 容器运行中",
        "1. 执行：docker run -d --name dgrp-cp-test alpine sleep 60\n"
        "2. 执行：echo 'dgrp' > /tmp/dgrp.txt\n"
        "3. 执行：docker cp /tmp/dgrp.txt dgrp-cp-test:/tmp/\n"
        "4. 执行：docker rm -f dgrp-cp-test",
        "文件复制成功"
    ),
    (
        "TC-DGRP-022",
        "test-docker-g - 系统标签验证",
        "验证 test-docker-g 创建的容器系统标签正确",
        PRE_DGRP,
        "1. 执行：docker run -d --name dgrp-label-test nginx\n"
        "2. 执行：docker inspect dgrp-label-test | grep system.authz\n"
        "3. 执行：docker rm -f dgrp-label-test",
        "system.authz.owner.uid=1002，system.authz.owner=test-docker-g"
    ),
]  # end dgrp_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4: bob 用户
# 策略：禁止 exec、build、push、commit、load、save、swarm、plugin、secret、config
# 配额：cpu=1, mem=1024MB, max_containers=3
# ─────────────────────────────────────────────────────────────────────────────
PRE_BOB = (
    PRE_COMMON + "\n"
    "• 当前操作用户：bob（uid=1003，普通用户）\n"
    "• deny_rules：禁止 exec、build、push、commit、load、save、swarm、plugin、secret、config\n"
    "• 配额：cpu_cores=1, mem_mb=1024, max_containers=3"
)

bob_cases = [
    (
        "TC-BOB-001",
        "bob - docker ps（允许，只看自己容器）",
        "验证 bob 可以执行 docker ps，且只看到自己的容器",
        PRE_BOB + "\n• alice 已有运行中的容器",
        "1. 以 bob 身份执行：docker ps\n"
        "2. 执行：docker ps -a",
        "只返回 bob 自己的容器；alice 的容器不可见"
    ),
    (
        "TC-BOB-002",
        "bob - docker pull（允许）",
        "验证 bob 可以拉取镜像",
        PRE_BOB,
        "1. 执行：docker pull node:18-alpine",
        "node:18-alpine 拉取成功"
    ),
    (
        "TC-BOB-003",
        "bob - docker run（允许，配额注入）",
        "验证 bob 创建容器时自动注入 1CPU/1024MB 配额",
        PRE_BOB,
        "1. 执行：docker run -d --name bob-app node:18-alpine sleep 3600\n"
        "2. 执行：docker inspect bob-app | grep -E 'NanoCPUs|Memory'",
        "容器创建成功；NanoCPUs=1000000000，Memory=1073741824"
    ),
    (
        "TC-BOB-004",
        "bob - docker exec（拒绝）",
        "验证 bob 被策略禁止执行 docker exec",
        PRE_BOB + "\n• bob-app 容器运行中",
        "1. 执行：docker exec bob-app echo hello",
        "返回 403 Forbidden；日志记录 action=exec, result=deny, deny_reason=policy"
    ),
    (
        "TC-BOB-005",
        "bob - docker attach（拒绝）",
        "验证 bob 被策略禁止执行 docker attach（exec 类操作）",
        PRE_BOB + "\n• bob-app 容器运行中",
        "1. 执行：docker attach bob-app",
        "返回 403 Forbidden；日志记录 action=exec, result=deny"
    ),
    (
        "TC-BOB-006",
        "bob - docker build（拒绝）",
        "验证 bob 被策略禁止执行 docker build",
        PRE_BOB + "\n• 已准备 Dockerfile",
        "1. 创建 /tmp/bobbuild/Dockerfile\n"
        "2. 执行：docker build -t bob-img /tmp/bobbuild/",
        "返回 403 Forbidden；日志记录 action=build, result=deny"
    ),
    (
        "TC-BOB-007",
        "bob - docker push（拒绝）",
        "验证 bob 被策略禁止执行 docker push",
        PRE_BOB,
        "1. 执行：docker push registry.example.com/bob-img",
        "返回 403 Forbidden；日志记录 action=push, result=deny"
    ),
    (
        "TC-BOB-008",
        "bob - docker commit（拒绝）",
        "验证 bob 被策略禁止执行 docker commit",
        PRE_BOB + "\n• bob-app 容器运行中",
        "1. 执行：docker commit bob-app bob-committed:v1",
        "返回 403 Forbidden；日志记录 action=commit, result=deny"
    ),
    (
        "TC-BOB-009",
        "bob - docker save（拒绝）",
        "验证 bob 被策略禁止执行 docker save",
        PRE_BOB,
        "1. 执行：docker save node:18-alpine -o /tmp/bob-node.tar",
        "返回 403 Forbidden；日志记录 action=save, result=deny"
    ),
    (
        "TC-BOB-010",
        "bob - docker load（拒绝）",
        "验证 bob 被策略禁止执行 docker load",
        PRE_BOB,
        "1. 执行：docker load -i /tmp/some-img.tar",
        "返回 403 Forbidden；日志记录 action=load, result=deny"
    ),
    (
        "TC-BOB-011",
        "bob - docker stop/start/restart（允许）",
        "验证 bob 可以管理自己容器的生命周期",
        PRE_BOB + "\n• bob-app 容器运行中",
        "1. 执行：docker stop bob-app\n"
        "2. 执行：docker start bob-app\n"
        "3. 执行：docker restart bob-app",
        "三个操作均成功"
    ),
    (
        "TC-BOB-012",
        "bob - docker rm（允许，只能删自己的容器）",
        "验证 bob 可以删除自己的容器，但不能删除他人容器",
        PRE_BOB + "\n• bob-app 容器已停止\n• alice 有容器 user-1004-alice-app",
        "1. 执行：docker stop bob-app && docker rm bob-app（应成功）\n"
        "2. 执行：docker rm user-1004-alice-app（应失败）",
        "步骤 1 成功；步骤 2 返回 403"
    ),
    (
        "TC-BOB-013",
        "bob - docker images（只看自己和公共镜像）",
        "验证 bob 的 docker images 只返回自己的镜像和公共镜像",
        PRE_BOB + "\n• test-sudo 已有私有镜像",
        "1. 执行：docker images",
        "只显示 bob 的镜像和公共镜像；test-sudo 的私有镜像不可见"
    ),
    (
        "TC-BOB-014",
        "bob - docker rmi（允许，只能删自己的镜像）",
        "验证 bob 可以删除自己的镜像",
        PRE_BOB + "\n• node:18-alpine 镜像已存在（bob 拉取的）",
        "1. 执行：docker rmi node:18-alpine",
        "镜像删除成功"
    ),
    (
        "TC-BOB-015",
        "bob - docker tag（允许）",
        "验证 bob 可以为自己的镜像打标签",
        PRE_BOB + "\n• node:18-alpine 镜像已存在",
        "1. 执行：docker pull node:18-alpine\n"
        "2. 执行：docker tag node:18-alpine bob-node:custom",
        "标签创建成功"
    ),
    (
        "TC-BOB-016",
        "bob - docker network 管理（允许）",
        "验证 bob 可以管理自己的网络",
        PRE_BOB,
        "1. 执行：docker network create bob-net\n"
        "2. 执行：docker network ls\n"
        "3. 执行：docker network inspect bob-net\n"
        "4. 执行：docker network rm bob-net",
        "所有操作成功；network ls 只显示 bob 自己的网络"
    ),
    (
        "TC-BOB-017",
        "bob - docker volume 管理（允许）",
        "验证 bob 可以管理自己的卷",
        PRE_BOB,
        "1. 执行：docker volume create bob-vol\n"
        "2. 执行：docker volume ls\n"
        "3. 执行：docker volume inspect bob-vol\n"
        "4. 执行：docker volume rm bob-vol",
        "所有操作成功；volume ls 只显示 bob 自己的卷"
    ),
    (
        "TC-BOB-018",
        "bob - docker logs/stats/top/inspect（允许）",
        "验证 bob 可以查看自己容器的详细信息",
        PRE_BOB + "\n• 重新创建 bob-app 容器",
        "1. 执行：docker run -d --name bob-app2 node:18-alpine sleep 3600\n"
        "2. 执行：docker logs bob-app2\n"
        "3. 执行：docker stats bob-app2 --no-stream\n"
        "4. 执行：docker top bob-app2\n"
        "5. 执行：docker inspect bob-app2\n"
        "6. 执行：docker rm -f bob-app2",
        "四个命令均返回正常数据"
    ),
    (
        "TC-BOB-019",
        "bob - docker cp（允许）",
        "验证 bob 可以在自己的容器和宿主机之间复制文件",
        PRE_BOB,
        "1. 执行：docker run -d --name bob-cp-test alpine sleep 60\n"
        "2. 执行：echo 'bob' > /tmp/bob.txt\n"
        "3. 执行：docker cp /tmp/bob.txt bob-cp-test:/tmp/\n"
        "4. 执行：docker rm -f bob-cp-test",
        "文件复制成功"
    ),
    (
        "TC-BOB-020",
        "bob - docker info/version（允许）",
        "验证 bob 可以查看系统信息",
        PRE_BOB,
        "1. 执行：docker info\n"
        "2. 执行：docker version",
        "两个命令均返回正常数据"
    ),
    (
        "TC-BOB-021",
        "bob - docker swarm（拒绝）",
        "验证 bob 被策略禁止执行 swarm 操作",
        PRE_BOB,
        "1. 执行：docker swarm init",
        "返回 403 Forbidden"
    ),
    (
        "TC-BOB-022",
        "bob - 容器数量配额限制",
        "验证 bob 创建容器超过 3 个时被拒绝",
        PRE_BOB + "\n• bob 已有 3 个运行中的容器",
        "1. 尝试执行：docker run -d alpine sleep 60",
        "返回 403 或 429；错误信息说明已达容器数量上限（max_containers=3）"
    ),
    (
        "TC-BOB-023",
        "bob - CPU 配额超限被拒绝",
        "验证 bob 请求超过 1 CPU 时被拒绝",
        PRE_BOB,
        "1. 执行：docker run --cpus=2 alpine echo hi",
        "返回 403；错误信息说明 CPU 超出配额（max=1）"
    ),
    (
        "TC-BOB-024",
        "bob - 内存配额超限被拒绝",
        "验证 bob 请求超过 1024MB 内存时被拒绝",
        PRE_BOB,
        "1. 执行：docker run -m 2g alpine echo hi",
        "返回 403；错误信息说明内存超出配额（max=1024MB）"
    ),
    (
        "TC-BOB-025",
        "bob - docker pause/unpause（允许）",
        "验证 bob 可以暂停和恢复自己的容器",
        PRE_BOB,
        "1. 执行：docker run -d --name bob-pause-test alpine sleep 60\n"
        "2. 执行：docker pause bob-pause-test\n"
        "3. 执行：docker unpause bob-pause-test\n"
        "4. 执行：docker rm -f bob-pause-test",
        "pause 后状态为 Paused；unpause 后恢复 Running"
    ),
    (
        "TC-BOB-026",
        "bob - docker rename（允许）",
        "验证 bob 可以重命名自己的容器",
        PRE_BOB,
        "1. 执行：docker run -d --name bob-old alpine sleep 60\n"
        "2. 执行：docker rename bob-old bob-new\n"
        "3. 执行：docker rm -f bob-new",
        "重命名成功"
    ),
    (
        "TC-BOB-027",
        "bob - docker kill（允许）",
        "验证 bob 可以强制终止自己的容器",
        PRE_BOB,
        "1. 执行：docker run -d --name bob-kill-test alpine sleep 60\n"
        "2. 执行：docker kill bob-kill-test\n"
        "3. 执行：docker rm bob-kill-test",
        "容器被终止"
    ),
    (
        "TC-BOB-028",
        "bob - docker diff/history（允许）",
        "验证 bob 可以查看容器变更和镜像历史",
        PRE_BOB,
        "1. 执行：docker run -d --name bob-diff-test alpine sleep 60\n"
        "2. 执行：docker diff bob-diff-test\n"
        "3. 执行：docker history alpine\n"
        "4. 执行：docker rm -f bob-diff-test",
        "两个命令均返回正常数据"
    ),
    (
        "TC-BOB-029",
        "bob - docker system prune（允许）",
        "验证 bob 可以执行系统清理",
        PRE_BOB,
        "1. 执行：docker system prune -f",
        "清理成功；只清理 bob 自己的资源"
    ),
    (
        "TC-BOB-030",
        "bob - 系统标签验证",
        "验证 bob 创建的容器系统标签正确",
        PRE_BOB,
        "1. 执行：docker run -d --name bob-label-test alpine sleep 60\n"
        "2. 执行：docker inspect bob-label-test | grep system.authz\n"
        "3. 执行：docker rm -f bob-label-test",
        "system.authz.owner.uid=1003，system.authz.owner=bob"
    ),
]  # end bob_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5: alice 用户（最严格策略）
# 策略：禁止 exec、build、push、commit、load、save、prune、swarm、plugin、secret、config
# 配额：cpu=1, mem=512MB, max_containers=2
# ─────────────────────────────────────────────────────────────────────────────
PRE_ALICE = (
    PRE_COMMON + "\n"
    "• 当前操作用户：alice（uid=1004，普通用户）\n"
    "• deny_rules：禁止 exec、build、push、commit、load、save、prune、swarm、plugin、secret、config\n"
    "• 配额：cpu_cores=1, mem_mb=512, max_containers=2"
)

alice_cases = [
    (
        "TC-ALICE-001",
        "alice - docker ps（允许，只看自己容器）",
        "验证 alice 可以执行 docker ps，且只看到自己的容器",
        PRE_ALICE + "\n• bob 已有运行中的容器",
        "1. 以 alice 身份执行：docker ps\n"
        "2. 执行：docker ps -a",
        "只返回 alice 自己的容器；bob 的容器不可见"
    ),
    (
        "TC-ALICE-002",
        "alice - docker pull（允许）",
        "验证 alice 可以拉取镜像",
        PRE_ALICE,
        "1. 执行：docker pull python:3.11-slim",
        "python:3.11-slim 拉取成功"
    ),
    (
        "TC-ALICE-003",
        "alice - docker run（允许，配额注入）",
        "验证 alice 创建容器时自动注入 1CPU/512MB 配额",
        PRE_ALICE,
        "1. 执行：docker run -d --name alice-app python:3.11-slim sleep 3600\n"
        "2. 执行：docker inspect alice-app | grep -E 'NanoCPUs|Memory'",
        "容器创建成功；NanoCPUs=1000000000，Memory=536870912（512MB）"
    ),
    (
        "TC-ALICE-004",
        "alice - docker exec（拒绝）",
        "验证 alice 被策略禁止执行 docker exec",
        PRE_ALICE + "\n• alice-app 容器运行中",
        "1. 执行：docker exec alice-app echo hello",
        "返回 403 Forbidden；日志记录 action=exec, result=deny, deny_reason=policy"
    ),
    (
        "TC-ALICE-005",
        "alice - docker build（拒绝）",
        "验证 alice 被策略禁止执行 docker build",
        PRE_ALICE + "\n• 已准备 Dockerfile",
        "1. 创建 /tmp/alicebuild/Dockerfile\n"
        "2. 执行：docker build -t alice-img /tmp/alicebuild/",
        "返回 403 Forbidden；日志记录 action=build, result=deny"
    ),
    (
        "TC-ALICE-006",
        "alice - docker push（拒绝）",
        "验证 alice 被策略禁止执行 docker push",
        PRE_ALICE,
        "1. 执行：docker push registry.example.com/alice-img",
        "返回 403 Forbidden"
    ),
    (
        "TC-ALICE-007",
        "alice - docker commit（拒绝）",
        "验证 alice 被策略禁止执行 docker commit",
        PRE_ALICE + "\n• alice-app 容器运行中",
        "1. 执行：docker commit alice-app alice-committed:v1",
        "返回 403 Forbidden"
    ),
    (
        "TC-ALICE-008",
        "alice - docker save（拒绝）",
        "验证 alice 被策略禁止执行 docker save",
        PRE_ALICE,
        "1. 执行：docker save python:3.11-slim -o /tmp/alice-py.tar",
        "返回 403 Forbidden"
    ),
    (
        "TC-ALICE-009",
        "alice - docker load（拒绝）",
        "验证 alice 被策略禁止执行 docker load",
        PRE_ALICE,
        "1. 执行：docker load -i /tmp/some-img.tar",
        "返回 403 Forbidden"
    ),
    (
        "TC-ALICE-010",
        "alice - docker prune（拒绝）",
        "验证 alice 被策略禁止执行所有 prune 操作",
        PRE_ALICE,
        "1. 执行：docker system prune -f\n"
        "2. 执行：docker container prune -f\n"
        "3. 执行：docker image prune -f\n"
        "4. 执行：docker volume prune -f",
        "四个命令均返回 403 Forbidden；日志记录 action=prune, result=deny"
    ),
    (
        "TC-ALICE-011",
        "alice - docker stop/start/restart（允许）",
        "验证 alice 可以管理自己容器的生命周期",
        PRE_ALICE + "\n• alice-app 容器运行中",
        "1. 执行：docker stop alice-app\n"
        "2. 执行：docker start alice-app\n"
        "3. 执行：docker restart alice-app",
        "三个操作均成功"
    ),
    (
        "TC-ALICE-012",
        "alice - docker rm（允许，只能删自己的容器）",
        "验证 alice 可以删除自己的容器，但不能删除他人容器",
        PRE_ALICE + "\n• alice-app 容器已停止\n• bob 有容器 user-1003-bob-app",
        "1. 执行：docker stop alice-app && docker rm alice-app（应成功）\n"
        "2. 执行：docker rm user-1003-bob-app（应失败）",
        "步骤 1 成功；步骤 2 返回 403"
    ),
    (
        "TC-ALICE-013",
        "alice - docker images（只看自己和公共镜像）",
        "验证 alice 的 docker images 只返回自己的镜像和公共镜像",
        PRE_ALICE + "\n• bob 已有私有镜像",
        "1. 执行：docker images",
        "只显示 alice 的镜像和公共镜像；bob 的私有镜像不可见"
    ),
    (
        "TC-ALICE-014",
        "alice - docker rmi（允许，只能删自己的镜像）",
        "验证 alice 可以删除自己的镜像",
        PRE_ALICE + "\n• python:3.11-slim 镜像已存在（alice 拉取的）",
        "1. 执行：docker rmi python:3.11-slim",
        "镜像删除成功"
    ),
    (
        "TC-ALICE-015",
        "alice - docker tag（允许）",
        "验证 alice 可以为自己的镜像打标签",
        PRE_ALICE + "\n• python:3.11-slim 镜像已存在",
        "1. 执行：docker pull python:3.11-slim\n"
        "2. 执行：docker tag python:3.11-slim alice-py:custom",
        "标签创建成功"
    ),
    (
        "TC-ALICE-016",
        "alice - docker network 管理（允许）",
        "验证 alice 可以管理自己的网络",
        PRE_ALICE,
        "1. 执行：docker network create alice-net\n"
        "2. 执行：docker network ls\n"
        "3. 执行：docker network rm alice-net",
        "所有操作成功；network ls 只显示 alice 自己的网络"
    ),
    (
        "TC-ALICE-017",
        "alice - docker volume 管理（允许）",
        "验证 alice 可以管理自己的卷",
        PRE_ALICE,
        "1. 执行：docker volume create alice-vol\n"
        "2. 执行：docker volume ls\n"
        "3. 执行：docker volume rm alice-vol",
        "所有操作成功；volume ls 只显示 alice 自己的卷"
    ),
    (
        "TC-ALICE-018",
        "alice - docker logs/stats/top/inspect（允许）",
        "验证 alice 可以查看自己容器的详细信息",
        PRE_ALICE + "\n• alice-app 容器运行中",
        "1. 执行：docker run -d --name alice-app2 python:3.11-slim sleep 3600\n"
        "2. 执行：docker logs alice-app2\n"
        "3. 执行：docker stats alice-app2 --no-stream\n"
        "4. 执行：docker inspect alice-app2\n"
        "5. 执行：docker rm -f alice-app2",
        "四个命令均返回正常数据"
    ),
    (
        "TC-ALICE-019",
        "alice - docker cp（允许）",
        "验证 alice 可以在自己的容器和宿主机之间复制文件",
        PRE_ALICE,
        "1. 执行：docker run -d --name alice-cp-test alpine sleep 60\n"
        "2. 执行：echo 'alice' > /tmp/alice.txt\n"
        "3. 执行：docker cp /tmp/alice.txt alice-cp-test:/tmp/\n"
        "4. 执行：docker rm -f alice-cp-test",
        "文件复制成功"
    ),
    (
        "TC-ALICE-020",
        "alice - docker info/version（允许）",
        "验证 alice 可以查看系统信息",
        PRE_ALICE,
        "1. 执行：docker info\n"
        "2. 执行：docker version",
        "两个命令均返回正常数据"
    ),
    (
        "TC-ALICE-021",
        "alice - docker swarm（拒绝）",
        "验证 alice 被策略禁止执行 swarm 操作",
        PRE_ALICE,
        "1. 执行：docker swarm init",
        "返回 403 Forbidden"
    ),
    (
        "TC-ALICE-022",
        "alice - 容器数量配额限制",
        "验证 alice 创建容器超过 2 个时被拒绝",
        PRE_ALICE + "\n• alice 已有 2 个运行中的容器",
        "1. 尝试执行：docker run -d alpine sleep 60",
        "返回 403 或 429；错误信息说明已达容器数量上限（max_containers=2）"
    ),
    (
        "TC-ALICE-023",
        "alice - 内存配额超限被拒绝",
        "验证 alice 请求超过 512MB 内存时被拒绝",
        PRE_ALICE,
        "1. 执行：docker run -m 1g alpine echo hi",
        "返回 403；错误信息说明内存超出配额（max=512MB）"
    ),
    (
        "TC-ALICE-024",
        "alice - docker pause/unpause（允许）",
        "验证 alice 可以暂停和恢复自己的容器",
        PRE_ALICE,
        "1. 执行：docker run -d --name alice-pause-test alpine sleep 60\n"
        "2. 执行：docker pause alice-pause-test\n"
        "3. 执行：docker unpause alice-pause-test\n"
        "4. 执行：docker rm -f alice-pause-test",
        "pause 后状态为 Paused；unpause 后恢复 Running"
    ),
    (
        "TC-ALICE-025",
        "alice - docker kill（允许）",
        "验证 alice 可以强制终止自己的容器",
        PRE_ALICE,
        "1. 执行：docker run -d --name alice-kill-test alpine sleep 60\n"
        "2. 执行：docker kill alice-kill-test\n"
        "3. 执行：docker rm alice-kill-test",
        "容器被终止"
    ),
    (
        "TC-ALICE-026",
        "alice - 系统标签验证",
        "验证 alice 创建的容器系统标签正确",
        PRE_ALICE,
        "1. 执行：docker run -d --name alice-label-test alpine sleep 60\n"
        "2. 执行：docker inspect alice-label-test | grep system.authz\n"
        "3. 执行：docker rm -f alice-label-test",
        "system.authz.owner.uid=1004，system.authz.owner=alice"
    ),
    (
        "TC-ALICE-027",
        "alice - 标签防篡改验证",
        "验证 alice 无法通过指定标签覆盖 system.authz.owner.uid",
        PRE_ALICE,
        "1. 执行：docker run -d --name alice-tamper-test -l system.authz.owner.uid=9999 alpine sleep 60\n"
        "2. 执行：docker inspect alice-tamper-test | grep system.authz.owner.uid\n"
        "3. 执行：docker rm -f alice-tamper-test",
        "system.authz.owner.uid=1004（代理注入的值），用户指定的 9999 被覆盖"
    ),
    (
        "TC-ALICE-028",
        "alice - docker search（允许）",
        "验证 alice 可以搜索 Docker Hub 镜像",
        PRE_ALICE + "\n• 网络可访问 Docker Hub",
        "1. 执行：docker search python --limit 5",
        "返回搜索结果列表"
    ),
    (
        "TC-ALICE-029",
        "alice - docker diff/history（允许）",
        "验证 alice 可以查看容器变更和镜像历史",
        PRE_ALICE,
        "1. 执行：docker run -d --name alice-diff-test alpine sleep 60\n"
        "2. 执行：docker diff alice-diff-test\n"
        "3. 执行：docker history alpine\n"
        "4. 执行：docker rm -f alice-diff-test",
        "两个命令均返回正常数据"
    ),
    (
        "TC-ALICE-030",
        "alice - 操作审计日志验证",
        "验证 alice 的所有操作（包括被拒绝的）均被记录到审计日志",
        PRE_ALICE,
        "1. 执行若干允许和被拒绝的操作\n"
        "2. 查看 /var/log/docker-authz/user-operation/alice.log",
        "日志包含所有操作记录；被拒绝的操作记录 result=deny；被允许的记录 result=allow"
    ),
]  # end alice_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 6: 跨用户隔离验证
# ─────────────────────────────────────────────────────────────────────────────
cross_cases = [
    (
        "TC-CROSS-001",
        "容器可见性隔离-5 用户同时运行",
        "验证 5 个用户同时运行容器时，每个用户只能看到自己的容器",
        PRE_COMMON + "\n• 5 个用户各自创建了 1 个容器",
        "1. 各用户分别执行 docker ps\n"
        "2. 记录每个用户看到的容器列表",
        "每个用户只看到自己的容器；5 个用户的视图完全隔离"
    ),
    (
        "TC-CROSS-002",
        "镜像可见性隔离",
        "验证私有镜像只对所有者可见，公共镜像对所有人可见",
        PRE_COMMON + "\n• root 已将 ubuntu:22.04 标记为公共镜像\n• 各用户各有私有镜像",
        "1. 各用户分别执行 docker images\n"
        "2. 检查是否能看到 ubuntu:22.04 和其他用户的私有镜像",
        "所有用户都能看到 ubuntu:22.04；各用户只能看到自己的私有镜像"
    ),
    (
        "TC-CROSS-003",
        "网络可见性隔离",
        "验证用户只能看到自己的网络",
        PRE_COMMON + "\n• 各用户各创建了 1 个网络",
        "1. 各用户分别执行 docker network ls",
        "每个用户只看到自己的网络；其他用户的网络不可见"
    ),
    (
        "TC-CROSS-004",
        "卷可见性隔离",
        "验证用户只能看到自己的卷",
        PRE_COMMON + "\n• 各用户各创建了 1 个卷",
        "1. 各用户分别执行 docker volume ls",
        "每个用户只看到自己的卷；其他用户的卷不可见"
    ),
    (
        "TC-CROSS-005",
        "跨用户操作容器被拒绝（bob 操作 alice 的容器）",
        "验证 bob 无法停止、删除、exec alice 的容器",
        PRE_COMMON + "\n• alice 已创建容器 user-1004-alice-app",
        "1. bob 执行：docker stop user-1004-alice-app\n"
        "2. bob 执行：docker rm user-1004-alice-app\n"
        "3. bob 执行：docker exec user-1004-alice-app echo hi",
        "三个操作均返回 403 Forbidden；日志记录 deny_reason=ownership"
    ),
    (
        "TC-CROSS-006",
        "跨用户操作镜像被拒绝",
        "验证 alice 无法删除 bob 的私有镜像",
        PRE_COMMON + "\n• bob 已构建私有镜像 bob-private-img",
        "1. alice 执行：docker rmi bob-private-img",
        "返回 403 Forbidden；日志记录 deny_reason=ownership"
    ),
    (
        "TC-CROSS-007",
        "容器名称前缀隔离验证",
        "验证不同用户创建同名容器时，实际名称通过前缀区分",
        PRE_COMMON,
        "1. bob 执行：docker run -d --name myapp alpine sleep 60\n"
        "2. alice 执行：docker run -d --name myapp alpine sleep 60\n"
        "3. root 执行：docker ps -a | grep myapp",
        "root 看到 user-1003-myapp 和 user-1004-myapp；两个容器共存无冲突"
    ),
    (
        "TC-CROSS-008",
        "网络名称前缀隔离验证",
        "验证不同用户创建同名网络时，实际名称通过前缀区分",
        PRE_COMMON,
        "1. bob 执行：docker network create mynet\n"
        "2. alice 执行：docker network create mynet\n"
        "3. root 执行：docker network ls | grep mynet",
        "root 看到 bob_u1003_mynet 和 alice_u1004_mynet；两个网络共存无冲突"
    ),
    (
        "TC-CROSS-009",
        "root 操作其他用户资源（所有权绕过）",
        "验证 root 可以操作任意用户的容器、镜像、网络、卷",
        PRE_COMMON + "\n• alice 和 bob 各有容器、网络、卷",
        "1. root 执行：docker stop user-1004-alice-app\n"
        "2. root 执行：docker rm user-1003-bob-app\n"
        "3. root 执行：docker network rm alice_u1004_mynet",
        "所有操作成功；root 不受所有权限制"
    ),
    (
        "TC-CROSS-010",
        "test-sudo 操作他人容器被拒绝",
        "验证 test-sudo 即使有 sudo 权限也无法操作其他用户的容器",
        PRE_COMMON + "\n• bob 已创建容器 user-1003-bob-app",
        "1. test-sudo 执行：sudo docker stop user-1003-bob-app",
        "返回 403 Forbidden；代理以 loginUID=1001 识别 test-sudo，所有权检查失败"
    ),
    (
        "TC-CROSS-011",
        "端口映射冲突防护",
        "验证两个用户无法同时绑定同一宿主机端口",
        PRE_COMMON,
        "1. bob 执行：docker run -d -p 9090:80 --name bob-port nginx\n"
        "2. alice 尝试执行：docker run -d -p 9090:80 --name alice-port nginx",
        "步骤 1 成功；步骤 2 返回错误（端口冲突）"
    ),
    (
        "TC-CROSS-012",
        "用户存储目录隔离",
        "验证用户无法挂载其他用户的存储目录",
        PRE_COMMON,
        "1. alice 尝试执行：docker run -v /var/docker/user-storage/user-1003/data:/data alpine ls /data",
        "返回 403 Forbidden；错误信息说明不允许挂载该路径"
    ),
    (
        "TC-CROSS-013",
        "代理重启后隔离状态保持",
        "验证代理重启后所有权记录和隔离状态不丢失",
        PRE_COMMON + "\n• 各用户均有容器和镜像",
        "1. 重启代理服务\n"
        "2. 各用户分别执行 docker ps 和 docker images",
        "重启后各用户仍只能看到自己的资源；所有权记录完整保留"
    ),
]  # end cross_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 7: 性能测试
# ─────────────────────────────────────────────────────────────────────────────
perf_cases = [
    (
        "TC-PERF-001",
        "单用户高频请求-docker ps 吞吐量",
        "验证代理在单用户高频请求下的响应延迟和吞吐量",
        PRE_COMMON + "\n• 测试工具：ab 或 wrk 或 shell 循环",
        "1. 以 bob 身份连续执行 100 次 docker ps\n"
        "2. 记录总耗时和平均延迟\n"
        "3. 查看代理日志中的 latency_ms 字段",
        "100 次请求全部成功；平均延迟 < 100ms；无超时或错误"
    ),
    (
        "TC-PERF-002",
        "多用户并发请求-5 用户同时 docker ps",
        "验证 5 个用户同时发起请求时代理的并发处理能力",
        PRE_COMMON,
        "1. 5 个用户同时执行 docker ps（各 20 次，共 100 次）\n"
        "2. 记录总耗时和错误数",
        "100 次请求全部成功；无数据混淆；平均延迟 < 200ms"
    ),
    (
        "TC-PERF-003",
        "并发容器创建-多用户同时 docker run",
        "验证多用户同时创建容器时代理的并发处理能力",
        PRE_COMMON,
        "1. bob、alice、test-docker-g 同时执行 docker run -d alpine sleep 60\n"
        "2. 等待所有命令完成\n"
        "3. 各用户执行 docker ps 验证容器归属",
        "所有容器创建成功；各用户只看到自己的容器；无所有权混淆"
    ),
    (
        "TC-PERF-004",
        "最大并发请求限制测试",
        "验证超过 --max-concurrent 限制时代理的行为",
        PRE_COMMON + "\n• 代理以 --max-concurrent=10 启动",
        "1. 同时发起 20 个 docker ps 请求\n"
        "2. 观察响应情况",
        "前 10 个请求正常处理；超出部分被排队或返回 503；无请求丢失或崩溃"
    ),
    (
        "TC-PERF-005",
        "请求超时处理",
        "验证超过 --request-timeout 的请求被正确中断",
        PRE_COMMON + "\n• 代理以 --request-timeout=5s 启动",
        "1. 发起一个预期耗时超过 5 秒的请求（如大镜像 pull）\n"
        "2. 等待超时",
        "5 秒后请求被中断；返回超时错误；代理日志记录超时事件；代理不崩溃"
    ),
    (
        "TC-PERF-006",
        "长时间运行稳定性测试",
        "验证代理在长时间运行后无内存泄漏或性能下降",
        PRE_COMMON,
        "1. 持续运行代理 1 小时\n"
        "2. 每 5 分钟执行一次 docker ps（共 12 次）\n"
        "3. 监控代理进程的内存使用",
        "所有请求均成功；内存使用稳定（无持续增长）；响应延迟无明显下降"
    ),
    (
        "TC-PERF-007",
        "并发日志写入性能",
        "验证多用户并发操作时日志写入不成为性能瓶颈",
        PRE_COMMON,
        "1. 5 个用户同时执行 20 次 docker ps（共 100 次）\n"
        "2. 检查日志文件完整性和写入速度",
        "100 条日志记录均完整写入；每条记录为有效 JSON；无乱码或截断"
    ),
    (
        "TC-PERF-008",
        "SQLite 数据库并发写入性能",
        "验证多用户并发创建/删除容器时数据库无竞争错误",
        PRE_COMMON,
        "1. 5 个用户同时创建容器（各 3 个，共 15 个）\n"
        "2. 同时删除所有容器\n"
        "3. 检查数据库一致性",
        "所有操作成功完成；数据库中所有权记录准确；无 SQLITE_BUSY 错误"
    ),
    (
        "TC-PERF-009",
        "大量容器时 docker ps 响应时间",
        "验证系统中有大量容器时 docker ps 的响应时间",
        PRE_COMMON + "\n• 系统中共有 50 个容器（各用户分布）",
        "1. 各用户执行 docker ps\n"
        "2. 记录响应时间",
        "docker ps 响应时间 < 500ms；过滤结果正确"
    ),
    (
        "TC-PERF-010",
        "策略热重载期间请求处理",
        "验证策略热重载期间正在处理的请求不受影响",
        PRE_COMMON,
        "1. 发起持续的 docker ps 请求流\n"
        "2. 同时修改 policy.yaml 触发热重载\n"
        "3. 观察请求是否中断",
        "热重载期间请求正常处理；无请求失败；新策略在重载完成后生效"
    ),
]  # end perf_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 8: 异常与边界测试
# ─────────────────────────────────────────────────────────────────────────────
err_cases = [
    (
        "TC-ERR-001",
        "Docker daemon 不可用时代理行为",
        "验证上游 Docker daemon 不可用时代理返回合适的错误",
        PRE_COMMON,
        "1. 停止 Docker daemon（systemctl stop docker）\n"
        "2. 以 bob 身份执行：docker ps\n"
        "3. 重启 Docker daemon",
        "返回 502 Bad Gateway 或连接错误；代理不崩溃；日志记录上游连接失败"
    ),
    (
        "TC-ERR-002",
        "policy.yaml 文件不存在时代理行为",
        "验证 policy.yaml 文件被删除后代理的行为",
        PRE_COMMON,
        "1. 删除 policy.yaml 文件\n"
        "2. 等待 3 秒\n"
        "3. 以 bob 身份执行：docker ps",
        "代理继续使用上一次有效策略；日志记录策略文件不存在警告；docker ps 正常返回"
    ),
    (
        "TC-ERR-003",
        "policy.yaml 格式错误时代理行为",
        "验证 policy.yaml 格式错误时代理保持使用上一次有效策略",
        PRE_COMMON,
        "1. 将 policy.yaml 修改为无效 YAML 格式（如缺少冒号）\n"
        "2. 等待 3 秒\n"
        "3. 以 bob 身份执行：docker ps",
        "代理继续使用上一次有效策略；日志记录策略解析错误；docker ps 正常返回"
    ),
    (
        "TC-ERR-004",
        "quota.yaml 格式错误时代理行为",
        "验证 quota.yaml 格式错误时代理使用默认配额",
        PRE_COMMON,
        "1. 将 quota.yaml 修改为无效格式\n"
        "2. 重启代理\n"
        "3. 以 bob 身份执行：docker run -d alpine sleep 60",
        "代理使用默认配额（或无配额）；日志记录配额文件解析错误"
    ),
    (
        "TC-ERR-005",
        "SQLite 数据库文件损坏时代理行为",
        "验证数据库文件损坏时代理的容错行为",
        PRE_COMMON,
        "1. 停止代理\n"
        "2. 损坏 owners.db 文件（写入随机数据）\n"
        "3. 重启代理\n"
        "4. 以 bob 身份执行：docker ps",
        "代理重新初始化数据库；日志记录数据库错误；docker ps 正常返回（可能无历史所有权数据）"
    ),
    (
        "TC-ERR-006",
        "身份伪造攻击防护",
        "验证用户无法通过修改请求头伪造其他用户身份",
        PRE_COMMON,
        "1. alice 通过 alice.sock 发送请求，在 Header 中注入 X-User: root\n"
        "2. 执行：docker ps",
        "代理忽略 Header 中的用户信息；仍以 SO_PEERCRED 识别的 alice 身份处理；返回 alice 的容器"
    ),
    (
        "TC-ERR-007",
        "访问其他用户的 socket 被拒绝",
        "验证用户无法通过其他用户的 socket 发起请求",
        PRE_COMMON,
        "1. bob 尝试连接 alice.sock（权限为 600，所有者为 alice）\n"
        "2. 执行：DOCKER_HOST=unix:///run/docker-authz/alice.sock docker ps",
        "连接被拒绝（权限错误）；bob 无法使用 alice 的 socket"
    ),
    (
        "TC-ERR-008",
        "无效 JWT Token 被拒绝",
        "验证携带无效/过期 JWT Token 的请求被拒绝",
        PRE_COMMON + "\n• 代理以 --jwt-secret=mysecret 启动",
        "1. 向 JWT TCP 端口发送请求，Header 携带伪造 Token\n"
        "2. 执行：docker ps",
        "返回 401 Unauthorized；auth.log 记录认证失败原因"
    ),
    (
        "TC-ERR-009",
        "操作不存在的容器",
        "验证操作不存在的容器时返回合适的错误",
        PRE_COMMON,
        "1. bob 执行：docker stop nonexistent-container\n"
        "2. bob 执行：docker rm nonexistent-container",
        "返回 404 Not Found；代理正确转发 Docker daemon 的错误响应"
    ),
    (
        "TC-ERR-010",
        "操作不存在的镜像",
        "验证操作不存在的镜像时返回合适的错误",
        PRE_COMMON,
        "1. bob 执行：docker rmi nonexistent-image:latest",
        "返回 404 Not Found；代理正确转发 Docker daemon 的错误响应"
    ),
    (
        "TC-ERR-011",
        "代理进程崩溃后自动重启",
        "验证 systemd 配置了 Restart=always，代理崩溃后自动重启",
        PRE_COMMON + "\n• 代理以 systemd 服务运行",
        "1. 执行：kill -9 <proxy_pid>\n"
        "2. 等待 5 秒\n"
        "3. 检查服务状态\n"
        "4. 以 bob 身份执行：docker ps",
        "服务自动重启；状态恢复为 active(running)；docker ps 正常返回"
    ),
    (
        "TC-ERR-012",
        "日志目录不可写时代理行为",
        "验证日志目录权限不足时代理的容错行为",
        PRE_COMMON,
        "1. 修改日志目录权限为只读（chmod 444）\n"
        "2. 以 bob 身份执行：docker ps\n"
        "3. 恢复日志目录权限",
        "docker ps 正常返回（操作不受影响）；代理日志记录日志写入失败警告"
    ),
    (
        "TC-ERR-013",
        "磁盘空间不足时代理行为",
        "验证磁盘空间不足时代理的行为",
        PRE_COMMON + "\n• 模拟磁盘空间不足（填充磁盘至 99%）",
        "1. 以 bob 身份执行：docker run -d alpine sleep 60",
        "返回合适的错误（磁盘空间不足）；代理不崩溃；日志记录错误"
    ),
    (
        "TC-ERR-014",
        "超大请求体处理",
        "验证代理能正确处理超大请求体（如大型 Dockerfile）",
        PRE_COMMON,
        "1. 创建包含 10000 行 RUN echo 的 Dockerfile\n"
        "2. 以 root 身份执行：docker build -t large-build .",
        "代理正确转发请求；不因请求体过大而崩溃或截断"
    ),
    (
        "TC-ERR-015",
        "并发策略重载与请求处理",
        "验证策略重载期间并发请求不导致竞争条件",
        PRE_COMMON,
        "1. 启动 10 个并发 docker ps 请求\n"
        "2. 同时发送 SIGHUP 触发策略重载\n"
        "3. 观察所有请求的结果",
        "所有请求正常完成；无 panic 或数据竞争；策略重载成功"
    ),
    (
        "TC-ERR-016",
        "新用户首次连接时 socket 不存在",
        "验证新用户在 socket 创建前尝试连接时的错误处理",
        PRE_COMMON + "\n• 新用户 newuser 刚创建，socket 尚未生成",
        "1. 以 newuser 身份执行：DOCKER_HOST=unix:///run/docker-authz/newuser.sock docker ps\n"
        "2. 等待 15 秒（代理检测新用户间隔）\n"
        "3. 再次执行 docker ps",
        "步骤 1 返回连接错误；步骤 3 成功（socket 已自动创建）"
    ),
    (
        "TC-ERR-017",
        "容器名称冲突处理",
        "验证同一用户创建同名容器时的错误处理",
        PRE_COMMON,
        "1. bob 执行：docker run -d --name bob-dup alpine sleep 60\n"
        "2. bob 再次执行：docker run -d --name bob-dup alpine sleep 60",
        "步骤 1 成功；步骤 2 返回 409 Conflict（容器名已存在）"
    ),
    (
        "TC-ERR-018",
        "网络名称冲突处理",
        "验证同一用户创建同名网络时的错误处理",
        PRE_COMMON,
        "1. bob 执行：docker network create bob-dup-net\n"
        "2. bob 再次执行：docker network create bob-dup-net",
        "步骤 1 成功；步骤 2 返回 409 Conflict"
    ),
]  # end err_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 9: 审计日志专项测试
# ─────────────────────────────────────────────────────────────────────────────
audit_cases = [
    (
        "TC-AUDIT-001",
        "操作日志-允许操作完整字段验证",
        "验证被允许的操作日志包含所有必要字段",
        PRE_COMMON,
        "1. bob 执行：docker ps\n"
        "2. 查看 /var/log/docker-authz/user-operation/bob.log 最新一条记录",
        "日志记录包含：time（RFC3339）、user=bob、uid=1003、action=ps、"
        "uri=/containers/json、result=allow、status_code=200、latency_ms>0"
    ),
    (
        "TC-AUDIT-002",
        "操作日志-拒绝操作完整字段验证",
        "验证被拒绝的操作日志包含拒绝原因",
        PRE_COMMON,
        "1. bob 执行：docker exec <container> echo hi（被策略禁止）\n"
        "2. 查看 bob.log 最新一条记录",
        "日志记录包含：result=deny、deny_reason=policy、status_code=403"
    ),
    (
        "TC-AUDIT-003",
        "认证日志-成功认证记录",
        "验证成功认证事件被记录到 auth.log",
        PRE_COMMON,
        "1. bob 连接代理并执行 docker ps\n"
        "2. 查看 /var/log/docker-authz/auth.log",
        "auth.log 包含记录：pid、effective_uid=1003、real_uid=1003、username=bob"
    ),
    (
        "TC-AUDIT-004",
        "认证日志-sudo 用户双重身份记录",
        "验证 sudo 用户的认证日志记录真实 UID 和有效 UID",
        PRE_COMMON,
        "1. test-sudo 执行：sudo docker ps\n"
        "2. 查看 auth.log",
        "auth.log 记录：effective_uid=0、real_uid=1001、switched_identity=true、username=test-sudo"
    ),
    (
        "TC-AUDIT-005",
        "认证日志-认证失败记录",
        "验证认证失败事件被记录到 auth.log",
        PRE_COMMON + "\n• 代理以 JWT 模式启动",
        "1. 使用无效 JWT Token 发起请求\n"
        "2. 查看 auth.log",
        "auth.log 记录认证失败事件，包含 failure_reason 字段"
    ),
    (
        "TC-AUDIT-006",
        "容器事件日志-容器生命周期记录",
        "验证容器创建、启动、停止、删除事件被记录",
        PRE_COMMON,
        "1. bob 执行：docker run -d --name bob-event-test alpine sleep 60\n"
        "2. bob 执行：docker stop bob-event-test\n"
        "3. bob 执行：docker rm bob-event-test\n"
        "4. 查看 /var/log/docker-authz/container-run/ 目录",
        "日志文件记录容器创建、启动、停止、删除事件，包含容器 ID、用户、时间戳"
    ),
    (
        "TC-AUDIT-007",
        "日志查询-按用户过滤",
        "验证 --query log --user 参数能正确过滤指定用户的日志",
        PRE_COMMON + "\n• alice 和 bob 均有操作日志",
        "1. 执行：./docker-authz-proxy --query log --user alice",
        "只返回 alice 的操作日志记录；不包含 bob 的记录"
    ),
    (
        "TC-AUDIT-008",
        "日志查询-按操作和结果过滤",
        "验证 --query log --action --result 参数能正确过滤日志",
        PRE_COMMON + "\n• 存在多种操作的日志记录",
        "1. 执行：./docker-authz-proxy --query log --action exec --result deny",
        "只返回 action=exec 且 result=deny 的日志记录"
    ),
    (
        "TC-AUDIT-009",
        "日志查询-按时间范围过滤",
        "验证 --since/--until 参数能正确过滤时间范围内的日志",
        PRE_COMMON + "\n• 存在不同时间段的日志记录",
        "1. 执行：./docker-authz-proxy --query log --since 2025-01-01T00:00:00Z --until 2025-12-31T23:59:59Z",
        "只返回指定时间范围内的日志记录"
    ),
    (
        "TC-AUDIT-010",
        "日志轮转-SIGUSR1 信号",
        "验证发送 SIGUSR1 信号后代理重新打开日志文件",
        PRE_COMMON,
        "1. 重命名当前日志文件（模拟 logrotate）\n"
        "2. 执行：kill -USR1 <proxy_pid>\n"
        "3. bob 执行：docker ps\n"
        "4. 检查是否创建了新日志文件",
        "代理重新打开日志文件；新操作记录写入新文件"
    ),
    (
        "TC-AUDIT-011",
        "各用户日志文件权限隔离",
        "验证用户无法读取其他用户的操作日志",
        PRE_COMMON,
        "1. 以 alice 身份尝试读取：cat /var/log/docker-authz/user-operation/bob.log",
        "返回权限拒绝错误；alice 只能读取自己的日志文件"
    ),
    (
        "TC-AUDIT-012",
        "延迟时间记录准确性",
        "验证操作日志中的 latency_ms 字段准确反映请求处理时间",
        PRE_COMMON,
        "1. bob 执行：docker ps\n"
        "2. 查看日志中的 latency_ms 字段",
        "latency_ms 为正整数，合理反映请求处理耗时（通常 < 1000ms）"
    ),
    (
        "TC-AUDIT-013",
        "代理运行日志-启动事件记录",
        "验证代理启动事件被记录到 proxy-run 日志",
        PRE_COMMON,
        "1. 重启代理服务\n"
        "2. 查看 /var/log/docker-authz/proxy-run/ 目录",
        "日志文件记录代理启动事件，包含启动时间、配置参数、策略加载结果"
    ),
    (
        "TC-AUDIT-014",
        "代理运行日志-策略重载事件记录",
        "验证策略重载事件被记录到 proxy-run 日志",
        PRE_COMMON,
        "1. 修改 policy.yaml\n"
        "2. 等待 3 秒（自动重载）\n"
        "3. 查看 proxy-run 日志",
        "日志记录策略重载事件，包含重载时间和新策略摘要"
    ),
]  # end audit_cases

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 10: 部署与运维测试
# ─────────────────────────────────────────────────────────────────────────────
deploy_cases = [
    (
        "TC-DEPLOY-001",
        "代理服务启动验证",
        "验证代理服务能以完整配置正常启动",
        "Linux 系统已安装 Docker；代理二进制已部署",
        "1. 执行：systemctl start docker-authz-proxy\n"
        "2. 执行：systemctl status docker-authz-proxy\n"
        "3. 检查 /run/docker-authz/ 目录",
        "服务状态为 active(running)；各用户 socket 已创建；日志目录已初始化"
    ),
    (
        "TC-DEPLOY-002",
        "新用户 socket 动态创建",
        "验证代理运行期间新增系统用户后自动创建其 socket",
        "代理已运行",
        "1. 执行：useradd -m newuser2\n"
        "2. 等待 15 秒\n"
        "3. 检查 /run/docker-authz/newuser2.sock",
        "newuser2.sock 已自动创建；权限为 600"
    ),
    (
        "TC-DEPLOY-003",
        "代理服务优雅停止",
        "验证代理服务能优雅停止，不丢失进行中的请求",
        "代理已运行，有活跃请求",
        "1. 发起一个长时间请求（如 docker pull）\n"
        "2. 执行：systemctl stop docker-authz-proxy\n"
        "3. 观察进行中的请求",
        "进行中的请求正常完成后服务停止；无请求被强制中断"
    ),
    (
        "TC-DEPLOY-004",
        "配置文件路径自定义",
        "验证通过命令行参数可以自定义各配置文件路径",
        "已准备自定义路径的配置文件",
        "1. 以 --policy=/custom/policy.yaml --quota-file=/custom/quota.yaml 启动代理\n"
        "2. 验证策略和配额生效",
        "代理使用自定义路径的配置文件；策略和配额正常生效"
    ),
    (
        "TC-DEPLOY-005",
        "日志级别配置验证",
        "验证 --log-level 参数能控制日志详细程度",
        "代理已停止",
        "1. 以 --log-level=debug 启动代理\n"
        "2. 执行 docker ps\n"
        "3. 检查日志输出",
        "日志包含 debug 级别的详细信息（请求解析、策略匹配等）"
    ),
    (
        "TC-DEPLOY-006",
        "代理崩溃后自动重启",
        "验证 systemd 配置了 Restart=always，代理崩溃后自动重启",
        "代理以 systemd 服务运行",
        "1. 执行：kill -9 <proxy_pid>\n"
        "2. 等待 5 秒\n"
        "3. 执行：systemctl status docker-authz-proxy",
        "服务自动重启；状态恢复为 active(running)"
    ),
    (
        "TC-DEPLOY-007",
        "孤立卷定期清理",
        "验证代理定期清理不再关联任何容器的孤立卷",
        "存在孤立卷（无容器使用）；清理间隔配置为 1 分钟",
        "1. 创建孤立卷：docker volume create orphan-vol\n"
        "2. 等待超过清理间隔时间\n"
        "3. 检查孤立卷是否被删除",
        "孤立卷已被自动删除；代理日志记录清理事件"
    ),
    (
        "TC-DEPLOY-008",
        "用户存储目录自动创建",
        "验证新用户首次连接时代理自动创建用户存储目录",
        "新用户 newuser3 首次使用代理",
        "1. newuser3 连接代理并执行 docker ps\n"
        "2. 检查 /var/docker/user-storage/ 目录",
        "目录 /var/docker/user-storage/user-<uid>/ 已自动创建"
    ),
    (
        "TC-DEPLOY-009",
        "DOCKER_HOST 环境变量自动配置",
        "验证 /etc/profile.d/docker-authz.sh 正确设置 DOCKER_HOST",
        "profile.d 脚本已部署",
        "1. 以 alice 身份登录新 shell\n"
        "2. 执行：echo $DOCKER_HOST",
        "输出 unix:///run/docker-authz/alice.sock"
    ),
    (
        "TC-DEPLOY-010",
        "代理升级后数据兼容性",
        "验证代理升级后历史所有权数据仍然有效",
        "代理已运行，各用户有容器和镜像",
        "1. 记录当前各用户的容器和镜像列表\n"
        "2. 升级代理二进制（替换文件并重启）\n"
        "3. 各用户重新执行 docker ps 和 docker images",
        "升级后各用户仍只能看到自己的资源；历史所有权数据完整保留"
    ),
]  # end deploy_cases

# ─────────────────────────────────────────────────────────────────────────────
# Build workbook
# ─────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

sheets = [
    ("环境搭建与策略配置", setup_cases,  True),
    ("root用户测试",       root_cases,   False),
    ("test-sudo用户测试",  sudo_cases,   False),
    ("test-docker-g用户测试", dgrp_cases, False),
    ("bob用户测试",        bob_cases,    False),
    ("alice用户测试",      alice_cases,  False),
    ("跨用户隔离验证",     cross_cases,  False),
    ("性能测试",           perf_cases,   False),
    ("异常与边界测试",     err_cases,    False),
    ("审计日志测试",       audit_cases,  False),
    ("部署与运维测试",     deploy_cases, False),
]

for title, cases, first in sheets:
    add_sheet(wb, title, cases, first)

# ── Summary sheet (insert at position 0) ─────────────────────────────────────
ws_sum = wb.create_sheet("测试概览", 0)
ws_sum.freeze_panes = "A2"

sum_hdrs = ["模块", "Sheet 名称", "用例数量", "测试重点"]
sum_widths = [24, 28, 12, 60]
for col, (h, w) in enumerate(zip(sum_hdrs, sum_widths), 1):
    c = ws_sum.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.column_dimensions[get_column_letter(col)].width = w
ws_sum.row_dimensions[1].height = 30

summary_rows = [
    ("环境搭建与策略配置", "环境搭建与策略配置", len(setup_cases),
     "用户创建、policy.yaml/quota.yaml 配置、socket 验证"),
    ("root 用户测试",      "root用户测试",       len(root_cases),
     "全量 Docker 命令遍历、无配额限制、全资源可见、系统标签"),
    ("test-sudo 用户测试", "test-sudo用户测试",  len(sudo_cases),
     "sudo 身份识别、允许/拒绝操作、配额注入、资源隔离"),
    ("test-docker-g 用户测试", "test-docker-g用户测试", len(dgrp_cases),
     "docker 组用户、build/push/commit/save/load 禁止、资源隔离"),
    ("bob 用户测试",       "bob用户测试",        len(bob_cases),
     "exec/build/push 禁止、配额限制、资源隔离"),
    ("alice 用户测试",     "alice用户测试",      len(alice_cases),
     "最严格策略、prune 禁止、最小配额、标签防篡改"),
    ("跨用户隔离验证",     "跨用户隔离验证",     len(cross_cases),
     "容器/镜像/网络/卷可见性、名称前缀、端口冲突、存储隔离"),
    ("性能测试",           "性能测试",           len(perf_cases),
     "高频请求、并发处理、超时、长时间稳定性、数据库并发"),
    ("异常与边界测试",     "异常与边界测试",     len(err_cases),
     "daemon 不可用、配置错误、身份伪造、资源不存在、崩溃恢复"),
    ("审计日志测试",       "审计日志测试",       len(audit_cases),
     "日志字段完整性、查询过滤、日志轮转、权限隔离"),
    ("部署与运维测试",     "部署与运维测试",     len(deploy_cases),
     "服务启停、动态 socket、配置自定义、升级兼容性"),
]

row_colors = [
    "E8F4FD","D5E8D4","FFF2CC","FCE4D6","EAD1DC",
    "CFE2F3","E2EFDA","F4CCCC","EAD9FF","D9EAD3","FFF9C4"
]
total = 0
for r, (mod, sheet, cnt, focus) in enumerate(summary_rows, 2):
    fill = PatternFill("solid", fgColor=row_colors[(r-2) % len(row_colors)])
    for col, val in enumerate([mod, sheet, cnt, focus], 1):
        c = ws_sum.cell(row=r, column=col, value=val)
        c.fill = fill; c.border = BORDER
        c.alignment = Alignment(horizontal="left" if col in (1,2,4) else "center",
                                vertical="center", wrap_text=True)
    ws_sum.row_dimensions[r].height = 28
    total += cnt

r_total = len(summary_rows) + 2
for col, val in enumerate(["合计", "", total, ""], 1):
    c = ws_sum.cell(row=r_total, column=col, value=val)
    c.font = Font(bold=True, size=12)
    c.fill = PatternFill("solid", fgColor="D9D9D9")
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[r_total].height = 28

out = "d:/code/docker-authz-proxy/docker-authz-proxy-testcases-v2.xlsx"
wb.save(out)
print(f"Generated: {out}  ({total} test cases across {len(sheets)} sheets)")


